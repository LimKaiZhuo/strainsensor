import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from own_package.others import print_array_to_excel, print_df_to_excel
from own_package.active_learning.acquisition import load_model_ensemble, model_ensemble_prediction
from own_package.features_labels_setup import load_data_to_fl
import os, itertools
from collections import deque

def combine_excel_results(results_excel_dir, end_excel_dir, plot_dir, sheets, fn):
    df = pd.read_excel(end_excel_dir, index_col=0)
    end_store = df.values[:,-2].tolist()
    p_end_store = df.values[:,-1].tolist()

    xls = pd.ExcelFile(results_excel_dir)

    p_y_store_store = []
    for sheet in sheets:
        df = pd.read_excel(xls, sheet_name=sheet, index_col=0)
        df = df.sort_index()
        p_y_store = np.sinh(df.iloc[:, fn + 20:fn + 39].values)
        p_y_store_store.append(np.concatenate((np.zeros((np.shape(p_y_store)[0], 1)), p_y_store), axis=1).tolist())

    y_store = np.sinh(df.iloc[:, fn + 1:fn + 20].values)
    y_store = np.concatenate((np.zeros((np.shape(y_store)[0], 1)), y_store), axis=1).tolist()

    numel = len(y_store[0])

    p_y_store_store.append(np.sinh(np.mean(np.arcsinh(np.array(p_y_store_store)), axis=0)))
    p_y_store_store = np.array(p_y_store_store)
    combine_mse = np.mean((np.arcsinh(p_y_store_store[-1,:,:])-np.arcsinh(np.array(y_store)))**2)

    for idx, [end, p_end, y] in enumerate(zip(end_store, p_end_store, y_store)):
        x = np.linspace(0, end, num=numel)
        p_x = np.linspace(0, p_end, num=numel)
        temp_p_y_store = p_y_store_store[:, idx, :]

        plt.plot(x, y, c='b', label='Actual Spline Fit')
        plt.scatter(x, y, c='b', marker='+')
        name_store = sheets + ['Combined']

        plt.plot(p_x, temp_p_y_store.tolist()[-1], c='r', label=name_store[-1])
        plt.scatter(p_x, temp_p_y_store.tolist()[-1], c='r', marker='x')

        for p_y, name in zip(temp_p_y_store.tolist()[:-1], name_store[:-1]):
            plt.plot(p_x, p_y, label=name)
            plt.scatter(p_x, p_y,  marker='x')

        plt.legend(loc='upper left')
        plt.title('Expt. ' + str(idx + 1))
        plt.savefig('{}/Expt_{}.png'.format(plot_dir, idx+1),
                    bbox_inches='tight')
        plt.close()

    wb = openpyxl.load_workbook(results_excel_dir)
    ws = wb['Results']
    ws.cell(1,1).value = 'mse'
    ws.cell(1,2).value = combine_mse
    wb.save(results_excel_dir)


def cutoff_combine_excel_results(dir_store, results_excel_dir,  plot_dir, sheets, fn, numel, plot_mode):
    def get_best_df(dir, name, wb):
        hparam_df = pd.read_excel('{}/hparam_results.xlsx'.format(dir), index_col=None)
        mse = hparam_df.iloc[:,-1].values
        min_idx = int(hparam_df.iloc[np.argmin(mse),0])

        xls = pd.ExcelFile('{}/skf_results.xlsx'.format(dir))
        skf_df = pd.read_excel(xls, sheet_name='{}_{}_0'.format(name, min_idx), index_col=0)

        df1 = skf_df.iloc[:, :fn + 1 + 2 * numel].sort_index()
        y_store = df1.iloc[:, fn + 1:fn + 1 + numel].values
        p_y = df1.iloc[:, fn + 1 + numel:fn + 1 + 2 * numel].values
        rc = np.mean(np.abs(y_store-p_y)/y_store)
        mse = np.mean((y_store-p_y)**2)


        df2 = skf_df.iloc[:, fn + 1 + 2 * numel:].reset_index(drop=True)
        best_name = '{}_{}'.format(name, min_idx)
        df2.iloc[0, 2] = best_name
        skf_df = pd.concat([df1, df2], axis=1, sort=False)

        sheet_names = wb.sheetnames
        if name in sheet_names:
            ws = wb[name]
        else:
            wb.create_sheet(name)
            ws = wb[name]

        print_df_to_excel(df=skf_df, ws=ws, index=True, header=True)

        return [best_name, mse, rc]

    while os.path.isfile(results_excel_dir):
        expand = 1
        while True:
            expand += 1
            new_file_name = results_excel_dir.split('.xlsx')[0] + ' - ' + str(expand) + '.xlsx'
            if os.path.isfile(new_file_name):
                continue
            else:
                results_excel_dir = new_file_name
                break

    best_store = []
    wb = openpyxl.Workbook()
    for dir, sheet in zip(dir_store,sheets):
        best_store.append(get_best_df(dir, sheet, wb))
    wb.save(results_excel_dir)

    cutoff = [10, 100]
    xls = pd.ExcelFile(results_excel_dir)

    p_y_store = []
    for sheet in sheets:
        df = pd.read_excel(xls, sheet_name=sheet, index_col=0)
        df = df.sort_index()

        p_y = df.iloc[:, fn + 1 + numel:fn + 1 + 2 * numel].values.tolist()
        p_y_store.append(p_y)

    y_store = df.iloc[:, fn + 1:fn + 1 + numel].values
    p_y_store_mean = np.mean(np.array(p_y_store), axis=0)

    combine_mse = np.mean((y_store-p_y_store_mean)**2)
    p_y_store.append(p_y_store_mean.tolist())

    rc = np.mean(np.abs(y_store-p_y_store_mean)/y_store)

    se = (y_store-p_y_store_mean)**2
    cumulative_mse = []
    for idx in range(np.shape(se)[0]):
        cumulative_mse.append(np.mean(se[0:idx+1,:]))

    sheets.append('Combined')

    if plot_mode:
        for idx, [x, p_x_store] in enumerate(zip(y_store.tolist(), np.swapaxes(np.array(p_y_store),0,1).tolist())):
            plt.plot([0, x[0], x[1], x[2]],
                     [0, 0, 10 * (x[1] - x[0]), cutoff[0] * (x[1] - x[0]) + cutoff[1] * (x[2] - x[1])], c='r',
                     label='Actual Spline Fit')
            for idx1, p_x in enumerate(p_x_store):
                if idx1==3:
                    plt.plot([0, p_x[0], p_x[1], p_x[2]],
                             [0, 0, 10 * (p_x[1] - p_x[0]), cutoff[0] * (p_x[1] - p_x[0]) + cutoff[1] * (p_x[2] - p_x[1])],
                             label=sheets[idx1])
            plt.legend(loc='upper left')
            plt.title('Expt. ' + str(idx + 1))
            plt.savefig('{}/Expt_{}.png'.format(plot_dir, idx+1),
                        bbox_inches='tight')
            plt.close()

    df.iloc[:, fn + 1 + numel:fn + 1 + 2 * numel] = np.array(p_y_store[-1])
    df = df.iloc[:, :fn + 1 + 2 * numel]
    df['Cumulative MSE'] = cumulative_mse

    wb = openpyxl.load_workbook(results_excel_dir)
    wb.create_sheet('Results')
    names = wb.sheetnames
    ws = wb[names[-1]]
    print_df_to_excel(df=df, ws=ws, index=True, header=True)

    best_store = np.array(best_store).T.tolist()
    best_store[0].append('Combined')
    best_store[1].append(combine_mse)
    best_store[2].append(rc)

    col = fn + 1 + 1 + 2 * numel + 3
    ws.cell(1, col).value = 'models'
    print_array_to_excel(best_store[0], (1, col+1), ws, axis=1)
    ws.cell(2, col+0).value = 'mse'
    print_array_to_excel([[float(x) for x in y] for y in best_store[1:]], (2, col + 1), ws, axis=2)
    ws.cell(3, col+0).value = 'RC'
    wb.save(results_excel_dir)


def mse_tracker(excel_store, write_excel, rounds, headers, fn, numel):
    while os.path.isfile(write_excel):
        expand = 1
        while True:
            expand += 1
            new_file_name = write_excel.split('.xlsx')[0] + ' - ' + str(expand) + '.xlsx'
            if os.path.isfile(new_file_name):
                continue
            else:
                write_excel = new_file_name
                break
    print('Writing into' + write_excel + '\n')

    mse_store = []
    rc_store = []
    se_store = []
    re_store = []
    are_store = []
    last_expt_store = []
    for excel in excel_store:
        df = pd.read_excel(excel, sheet_name='Results', index_col=0)
        mse_store.append(df.iloc[0, 16:].values.tolist())
        rc_store.append(df.iloc[1, 16:].values.tolist())
        y_store = df.iloc[:, fn + 1:fn + 1 + numel].values
        p_y = df.iloc[:, fn + 1 + numel:fn + 1 + 2 * numel].values
        se = np.square(p_y-y_store)
        se_store.append(se)
        re = np.abs(p_y-y_store) / y_store
        are = np.arctan(re)
        re_store.append(re)
        are_store.append(are)
        last_expt_store.append(np.shape(se)[0])

    wb = openpyxl.Workbook()
    wb.create_sheet('MSE Results')
    ws = wb[wb.sheetnames[-1]]
    mse_df = pd.DataFrame(data=mse_store, index=rounds, columns=headers)
    print_df_to_excel(df=mse_df, ws=ws)

    wb.create_sheet('RE Results')
    ws = wb[wb.sheetnames[-1]]
    rc_df = pd.DataFrame(data=rc_store, index=rounds, columns=headers)
    print_df_to_excel(df=rc_df, ws=ws)

    wb.create_sheet('ARE Results')
    ws = wb[wb.sheetnames[-1]]
    are_df = pd.DataFrame(data=[np.mean(x) for x in are_store], index=rounds, columns=['Combined'])
    print_df_to_excel(df=are_df, ws=ws)

    wb.create_sheet('Batch MSE Results')
    batch_store = []
    re_batch_store = []
    are_batch_store = []
    last_expt_store0 = [0]+last_expt_store[:-1]
    for idx, (last_expt_idx0, last_expt_idx) in enumerate(zip(last_expt_store0, last_expt_store)):
        batch = []
        re_batch = []
        are_batch = []
        for se, re, are in zip(se_store[idx:], re_store[idx:], are_store[idx:]):
            batch.append(np.mean(se[last_expt_idx0:last_expt_idx, :]))
            re_batch.append(np.mean(re[last_expt_idx0:last_expt_idx, :]))
            are_batch.append(np.mean(are[last_expt_idx0:last_expt_idx, :]))
        batch_store.append(batch)
        re_batch_store.append(re_batch)
        are_batch_store.append(are_batch)
    batch_store = [['']*idx + batch for idx,batch in enumerate(batch_store)]
    ws = wb[wb.sheetnames[-1]]
    df = pd.DataFrame(data=batch_store, index=last_expt_store, columns=rounds)
    print_df_to_excel(df=df, ws=ws)

    wb.create_sheet('Batch RE Results')
    re_batch_store = [[''] * idx + re_batch for idx, re_batch in enumerate(re_batch_store)]
    ws = wb[wb.sheetnames[-1]]
    df = pd.DataFrame(data=re_batch_store, index=last_expt_store, columns=rounds)
    print_df_to_excel(df=df, ws=ws)

    wb.create_sheet('Batch ARE Results')
    are_batch_store = [[''] * idx + are_batch for idx, are_batch in enumerate(are_batch_store)]
    ws = wb[wb.sheetnames[-1]]
    df = pd.DataFrame(data=are_batch_store, index=last_expt_store, columns=rounds)
    print_df_to_excel(df=df, ws=ws)

    wb.save(write_excel)


def final_prediction_results(write_excel, model_dir_store, combined_excel_store, rounds, excel_loader_dir_store, fn, numel):
    wb = openpyxl.load_workbook(write_excel)
    results_col = fn + 1 + 1 + 2 * numel + 3
    mse_store = []
    mre_store = []
    mare_store = []
    final_excel_loader_dir = excel_loader_dir_store[-1]
    final_features = pd.read_excel(final_excel_loader_dir, sheet_name='features',
                                   index_col=0).sort_index().values
    final_df = pd.read_excel(final_excel_loader_dir, sheet_name='cutoff', index_col=0).sort_index()
    for idx, (model_dir, combined_excel, loader_excel, round) in enumerate(zip(model_dir_store, combined_excel_store, excel_loader_dir_store, rounds)):
        wb.create_sheet('Round {}'.format(round))
        ws = wb[wb.sheetnames[-1]]

        combined_df = pd.read_excel(combined_excel, sheet_name='Results', index_col=0).sort_index()
        numel_expt = combined_df.shape[0]
        total_expt = final_df.shape[0]

        y_store = combined_df.iloc[:, fn + 1:fn + 1 + numel].values
        p_y_store = combined_df.iloc[:, fn + 1 + numel:fn + 1 + 2 * numel].values

        if total_expt>numel_expt:
            model_store = load_model_ensemble(model_dir)
            fl = load_data_to_fl(data_loader_excel_file=loader_excel, norm_mask=[0,1,3,4,5],
                                 normalise_labels=True, label_type='cutoff')

            p_y, _ = model_ensemble_prediction(model_store, fl.apply_scaling(final_features[numel_expt:,:]))

            y_store = np.concatenate((y_store, final_df.values[numel_expt:,:]),axis=0)
            p_y_store = np.concatenate((p_y_store, p_y),axis=0)

        se_store = (y_store-p_y_store)**2
        re_store = np.abs(y_store - p_y_store) / y_store
        are_store = np.arctan(re_store)

        column_headers = final_df.columns.values.tolist()

        df = pd.DataFrame(data=np.concatenate((y_store, p_y_store, se_store, re_store, are_store), axis=1),
                          index=list(final_df.index),
                          columns=column_headers +  ['P_{}'.format(col) for col in column_headers]
                                  + ['SE_{}'.format(col) for col in column_headers] + ['RE_{}'.format(col) for col in column_headers]
                          + ['ARE_{}'.format(col) for col in column_headers])
        print_df_to_excel(df=df, ws=ws)
        
        col = fn + 1 + 1 + 2 * numel + 3
        mse_store.append(np.mean(se_store))
        mre_store.append(np.mean(re_store))
        mare_store.append(np.mean(are_store))
        ws.cell(1, col).value = 'MSE'
        ws.cell(1, col+1).value = mse_store[-1]
        ws.cell(2, col).value = 'MRE'
        ws.cell(2, col+1).value = mre_store[-1]
        ws.cell(3, col).value = 'ARE'
        ws.cell(3, col+1).value = mare_store[-1]

    wb.create_sheet('Final_results')
    ws = wb[wb.sheetnames[-1]]
    df = pd.DataFrame(data=np.array([mse_store, mre_store, mare_store]), index=['mse', 're', 'are'], columns=rounds)
    print_df_to_excel(df=df, ws=ws)
    wb.save(write_excel)

