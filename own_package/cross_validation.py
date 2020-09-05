import tensorflow as tf
from tensorflow.python.keras import backend as K
import gc
import numpy as np
import pandas as pd
from pandas import Series
import openpyxl
from openpyxl import load_workbook
from sklearn.metrics import mean_squared_error
import time, os, pickle

# Own Scripts
from own_package.models.models import MTmodel, Kmodel, Pmodel
from own_package.svr import SVRmodel, MIMOSVRmodel, DTRmodel, Predict_SVR_DTR, XGBmodel
from own_package.active_learning.acquisition import model_ensemble_prediction
from .others import print_array_to_excel, print_df_to_excel


def run_skf(model_mode, loss_mode, fl, fl_store, hparams,
            skf_file, label_type='cutoff', scoring='mse',
            skf_sheet=None,
            k_folds=10, k_shuffle=True,
            save_model=False, save_model_name=None, save_model_dir=None,
            plot_name=None):
    '''
    Stratified k fold cross validation for training and evaluating model 2 only. Model 1 data is trained before hand.
    :param model_mode: Choose between using SNN or cDNN (non_smiles) and SNN_smiles or cDNN_smiles
    :param cv_mode: Cross validation mode. Either 'skf' or 'loocv'.
    :param hparams: hparams dict containing hyperparameters information
    :param loader_file: data_loader excel file location
    :param skf_file: skf_file name to save excel file as
    :param skf_sheet: name of sheet to save inside the skf_file excel. If None, will default to SNN or cDNN as name
    :param k_folds: Number of k folds. Used only for skf cv_mode
    :param k_shuffle: Whether to shuffle the given examples to split into k folds if using skf
    :return:
    '''

    # Run k model instance to perform skf
    predicted_labels_store = []
    mse_store = []
    mse_norm_store = []
    folds = []
    val_idx = []
    val_features_c = []
    val_labels = []
    for fold, fl_tuple in enumerate(fl_store):
        instance_start = time.time()
        (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl

        # Set up model
        if loss_mode == 'normal':
            sess = tf.compat.v1.Session()
            # sess = tf.Session()
            K.set_session(sess)
            model = MTmodel(fl=ss_fl, mode=model_mode, hparams=hparams)
        elif loss_mode == 'hul':
            model = HULMTmodel(fl=ss_fl, mode=model_mode, hparams=hparams)
            print('HUL Standard Deviation Values:')
            print([np.exp(K.get_value(log_var[0])) ** 0.5 for log_var in model.model.layers[-1].log_vars])
        elif loss_mode == 'ann':
            sess = tf.compat.v1.Session()
            # sess = tf.Session()
            K.set_session(sess)
            model = Kmodel(fl=ss_fl, mode=model_mode, hparams=hparams)
        elif loss_mode == 'p_model':
            model = Pmodel(fl=ss_fl, mode=model_mode, hparams=hparams)
        elif loss_mode == 'svr':
            if not fl.normalise_labels:
                raise TypeError('fl labels are not normalised. For SVR, the labels must be normalised.')
            model = SVRmodel(fl=ss_fl, epsilon=hparams['epsilon'], c=hparams['c'])
        elif loss_mode == 'dtr':
            #if not fl.normalise_labels:
            #    raise TypeError('fl labels are not normalised. For SVR, the labels must be normalised.')
            model = DTRmodel(fl=ss_fl, max_depth=hparams['max_depth'], num_est=hparams['num_est'])
        elif loss_mode == 'mimosvr':
            if not fl.normalise_labels:
                raise TypeError('fl labels are not normalised. For SVR, the labels must be normalised.')
            model = MIMOSVRmodel(fl=ss_fl, gamma=hparams['gamma'])

        else:
            raise KeyError('loss_mode ' + loss_mode + 'is not a valid selection for loss mode.')

        # Train model and save model training loss vs epoch plot if plot_name is given, else no plot will be saved
        if plot_name:
            model.train_model(ss_fl, i_ss_fl, plot_name='{}_fold_{}.png'.format(plot_name, fold))
        else:
            model.train_model(ss_fl, i_ss_fl)

        # Evaluation
        predicted_labels, mse, mse_norm = model.eval(i_ss_fl)
        if fl.normalise_labels:
            predicted_labels = fl.labels_scaler.inverse_transform(predicted_labels)

        if label_type=='cutoff':
            for row, p_label in enumerate(predicted_labels.tolist()):
                if p_label[1]>p_label[2]:
                    predicted_labels[row,1]=predicted_labels[row,2]
                if p_label[0]>predicted_labels[row,1]:
                    predicted_labels[row,0]=predicted_labels[row,1]
            pass

        predicted_labels_store.extend(predicted_labels)
        mse_store.append(mse)
        mse_norm_store.append(mse_norm)
        '''
        if fold == k_folds-1:
            stringlist = []
            model.model.summary(print_fn=lambda x: stringlist.append(x))
            short_model_summary = "\n".join(stringlist)
            print(short_model_summary)
        '''
        # Saving model
        if save_model:
            # Set save_model_name
            if isinstance(save_model_name, str):
                save_model_name1 = save_model_name + '_' + model_mode + '_' + str(fold + 1)
            else:
                save_model_name1 = model_mode + '_' + str(fold + 1)

            # Save model
            print('Saving instance {} model in {}'.format(fold + 1, save_model_dir + save_model_name1 + '.h5'))
            if loss_mode == 'normal' or loss_mode == 'ann':
                model.model.save(save_model_dir + save_model_name1 + '.h5')
            elif loss_mode == 'hul':
                model.prediction_model.save(save_model_dir + save_model_name1 + '.h5')
            elif loss_mode == 'svr' or loss_mode == 'dtr':
                pickle.dump(Predict_SVR_DTR(model=model.model, labels_scaler=model.labels_scaler),
                            open(save_model_dir + save_model_name1 + '.pkl', 'wb'))

        # Need to put the next 3 lines if not memory will run out
        del model
        if loss_mode == 'normal' or loss_mode == 'ann':
            K.clear_session()
            sess.close()
        gc.collect()

        # Preparing data to put into new_df that consists of all the validation dataset and its predicted labels
        folds.extend([fold] * i_ss_fl.count)  # Make a col that contains the fold number for each example
        if len(val_features_c):
            val_features_c = np.concatenate((val_features_c, i_ss_fl.features_c),
                                            axis=0)
        else:
            val_features_c = i_ss_fl.features_c

        val_labels.extend(i_ss_fl.labels)
        val_idx.extend(i_ss_fl.idx)

        # Printing one instance summary.
        instance_end = time.time()
        print(
            '\nFor k-fold run {} out of {}. Each fold has {} examples. Model is {} with {} loss. Time taken for '
            'instance = {}\n'
            'Post-training results: \nmse = {}, mse_norm = {}. Scoring is {}\n'
            '####################################################################################################'
                .format(fold + 1, k_folds, i_ss_fl.count, model_mode, loss_mode, instance_end - instance_start, mse,
                        mse_norm, scoring))

    mse_avg = np.average(mse_store)
    mse_norm_avg = np.average(mse_norm_store)
    re = np.average(np.abs(np.array(val_labels) - np.array(predicted_labels_store))/np.array(val_labels))

    # Calculating metrics based on complete validation prediction
    mse_full = mean_squared_error(val_labels, predicted_labels_store)
    try:
        mse_norm_full = mean_squared_error(fl.labels_scaler.transform(val_labels),
                                           fl.labels_scaler.transform(predicted_labels_store))
    except AttributeError:
        mse_norm_full=mse_full

    # Creating dataframe to print into excel later.
    new_df = np.concatenate((np.array(folds)[:, None],  # Convert 1d list to col. vector
                             val_features_c,
                             np.array(val_labels),
                             np.array(predicted_labels_store))
                            , axis=1)
    if fl.label_type == 'points':
        predicted_labels_name = list(map(str, np.arange(2,21)))
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = ['folds'] + \
                  list(map(str, fl.features_c_names)) + \
                  list(map(str, np.arange(2,21))) + \
                  predicted_labels_name
    elif fl.label_type == 'cutoff':
        predicted_labels_name = list(fl.labels_names)
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = ['folds'] + \
                  list(map(str, fl.features_c_names)) + \
                  list(fl.labels_names) + \
                  predicted_labels_name
    elif fl.label_type == 'gf20':
        predicted_labels_name = list(map(str, np.arange(1,21)))
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = ['folds'] + \
                  list(map(str, fl.features_c_names)) + \
                  list(map(str, np.arange(1,21))) + \
                  predicted_labels_name

    # val_idx is the original position of the example in the data_loader
    new_df = pd.DataFrame(data=new_df, columns=headers, index=val_idx)

    print('Writing into' + skf_file)
    wb = load_workbook(skf_file)

    # Creating new worksheet. Even if SNN worksheet already exists, a new SNN1 ws will be created and so on
    if skf_sheet is None:
        wb.create_sheet(model_mode)
    else:
        wb.create_sheet(model_mode + skf_sheet)
    sheet_name = wb.sheetnames[-1]  # Taking the ws name from the back ensures that if SNN1 is the new ws, it works

    # Writing hparam dataframe first
    pd_writer = pd.ExcelWriter(skf_file, engine='openpyxl')
    pd_writer.book = wb
    pd_writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    new_df.to_excel(pd_writer, sheet_name)
    start_col = len(new_df.columns) + 4
    hparams = pd.DataFrame(dict([(k, Series(v)) for k, v in hparams.items()]))
    hparams.to_excel(pd_writer, sheet_name, startrow=0, startcol=start_col - 1)
    start_row = 5

    # Writing other subset split, instance per run, and bounds
    ws = wb[sheet_name]
    headers = ['mse', 'mse_norm', 're']
    values = [mse_avg, mse_norm_avg]
    values_full = [mse_full, mse_norm_full, re]
    print_array_to_excel(np.array(headers), (1 + start_row, start_col + 1), ws, axis=1)
    print_array_to_excel(np.array(values), (2 + start_row, start_col + 1), ws, axis=1)
    print_array_to_excel(np.array(values_full), (3 + start_row, start_col + 1), ws, axis=1)
    ws.cell(2 + start_row, start_col).value = 'Folds avg'
    ws.cell(3 + start_row, start_col).value = 'Overall'
    pd_writer.save()
    pd_writer.close()
    wb.close()

    if scoring == 'mse':
        return mse_full
    elif scoring == 're':
        return re
    else:
        raise KeyError('Scoring function {} is not valid'.format(scoring))


def run_skf_with_training_error(model_mode, loss_mode, fl, fl_store, hparams,
            skf_file, label_type='cutoff', scoring='mse',
            skf_sheet=None, te_sheet = None,
            k_folds=10, k_shuffle=True,
            save_model=False, save_model_name=None, save_model_dir=None,
            plot_name=None):
    '''
    Stratified k fold cross validation for training and evaluating model 2 only. Model 1 data is trained before hand.
    :param model_mode: Choose between using SNN or cDNN (non_smiles) and SNN_smiles or cDNN_smiles
    :param cv_mode: Cross validation mode. Either 'skf' or 'loocv'.
    :param hparams: hparams dict containing hyperparameters information
    :param loader_file: data_loader excel file location
    :param skf_file: skf_file name to save excel file as
    :param skf_sheet: name of sheet to save inside the skf_file excel. If None, will default to SNN or cDNN as name
    :param k_folds: Number of k folds. Used only for skf cv_mode
    :param k_shuffle: Whether to shuffle the given examples to split into k folds if using skf
    :return:
    '''
    fn=6
    numel=3
    # Run k model instance to perform skf
    predicted_labels_store = []
    mse_store = []
    mse_norm_store = []
    folds = []
    val_idx = []
    val_features_c = []
    val_labels = []
    column_headers = fl.labels_names

    wb = openpyxl.load_workbook(te_sheet)
    msee_store = []
    mre_store = []

    for fold, fl_tuple in enumerate(fl_store):
        instance_start = time.time()
        (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl

        wb.create_sheet('{}'.format(fold))
        ws = wb[wb.sheetnames[-1]]

        # Set up model
        if loss_mode == 'normal':
            sess = tf.compat.v1.Session()
            # sess = tf.Session()
            K.set_session(sess)
            model = MTmodel(fl=ss_fl, mode=model_mode, hparams=hparams, labels_norm=labels_norm)
        elif loss_mode == 'hul':
            model = HULMTmodel(fl=ss_fl, mode=model_mode, hparams=hparams, labels_norm=labels_norm)
            print('HUL Standard Deviation Values:')
            print([np.exp(K.get_value(log_var[0])) ** 0.5 for log_var in model.model.layers[-1].log_vars])
        elif loss_mode == 'ann':
            sess = tf.compat.v1.Session()
            # sess = tf.Session()
            K.set_session(sess)
            model = Kmodel(fl=ss_fl, mode=model_mode, hparams=hparams)
        elif loss_mode == 'p_model':
            model = Pmodel(fl=ss_fl, mode=model_mode, hparams=hparams)
        elif loss_mode == 'svr':
            if not fl.normalise_labels:
                raise TypeError('fl labels are not normalised. For SVR, the labels must be normalised.')
            model = SVRmodel(fl=ss_fl, epsilon=hparams['epsilon'], c=hparams['c'])
        elif loss_mode == 'dtr':
            #if not fl.normalise_labels:
                #raise TypeError('fl labels are not normalised. For SVR, the labels must be normalised.')
            model = DTRmodel(fl=ss_fl, max_depth=hparams['max_depth'], num_est=hparams['num_est'])
        elif loss_mode == 'mimosvr':
            if not fl.normalise_labels:
                raise TypeError('fl labels are not normalised. For SVR, the labels must be normalised.')
            model = MIMOSVRmodel(fl=ss_fl, gamma=hparams['gamma'])

        else:
            raise KeyError('loss_mode ' + loss_mode + 'is not a valid selection for loss mode.')

        # Train model and save model training loss vs epoch plot if plot_name is given, else no plot will be saved
        if plot_name:
            model.train_model(ss_fl, i_ss_fl, plot_name='{}_fold_{}.png'.format(plot_name, fold))
        else:
            model.train_model(ss_fl, i_ss_fl)

        p_y, _, _ = model.eval(fl)
        if fl.normalise_labels:
            p_y = fl.labels_scaler.inverse_transform(p_y)
        for row, p_label in enumerate(p_y.tolist()):
            if p_label[1] > p_label[2]:
                p_y[row, 1] = p_y[row, 2]
            if p_label[0] > p_y[row, 1]:
                p_y[row, 0] = p_y[row, 1]
        se_store = (fl.labels - p_y) ** 2
        re_store = np.abs(fl.labels - p_y) / fl.labels

        df = pd.DataFrame(data=np.concatenate((fl.labels, p_y, se_store, re_store), axis=1),
                          index=list(range(1, 1 + fl.count)),
                          columns=list(column_headers)
                                  + ['P_{}'.format(col) for col in column_headers]
                                  + ['SE_{}'.format(col) for col in column_headers]
                                  + ['RE_{}'.format(col) for col in column_headers])
        print_df_to_excel(df=df, ws=ws)

        col = fn + 1 + 1 + 2 * numel + 3
        msee_store.append(np.mean(se_store))
        mre_store.append(np.mean(re_store))
        ws.cell(1, col).value = 'MSE'
        ws.cell(1, col + 1).value = msee_store[-1]
        ws.cell(2, col).value = 'MRE'
        ws.cell(2, col + 1).value = mre_store[-1]
        ws.cell(3, col).value = 'ARE'
        ws.cell(3, col + 1).value = mare_store[-1]

        # Evaluation
        predicted_labels, mse, mse_norm = model.eval(i_ss_fl)
        if fl.normalise_labels:
            predicted_labels = fl.labels_scaler.inverse_transform(predicted_labels)

        if label_type=='cutoff':
            for row, p_label in enumerate(predicted_labels.tolist()):
                if p_label[1]>p_label[2]:
                    predicted_labels[row,1]=predicted_labels[row,2]
                if p_label[0]>predicted_labels[row,1]:
                    predicted_labels[row,0]=predicted_labels[row,1]
            pass

        predicted_labels_store.extend(predicted_labels)
        mse_store.append(mse)
        mse_norm_store.append(mse_norm)
        '''
        if fold == k_folds-1:
            stringlist = []
            model.model.summary(print_fn=lambda x: stringlist.append(x))
            short_model_summary = "\n".join(stringlist)
            print(short_model_summary)
        '''
        # Saving model
        if save_model:
            # Set save_model_name
            if isinstance(save_model_name, str):
                save_model_name1 = save_model_name + '_' + model_mode + '_' + str(fold + 1)
            else:
                save_model_name1 = model_mode + '_' + str(fold + 1)

            # Save model
            print('Saving instance {} model in {}'.format(fold + 1, save_model_dir + save_model_name1 + '.h5'))
            if loss_mode == 'normal' or loss_mode == 'ann':
                model.model.save(save_model_dir + save_model_name1 + '.h5')
            elif loss_mode == 'hul':
                model.prediction_model.save(save_model_dir + save_model_name1 + '.h5')
            elif loss_mode == 'svr' or loss_mode == 'dtr':
                pickle.dump(Predict_SVR_DTR(model=model.model, labels_scaler=model.labels_scaler),
                            open(save_model_dir + save_model_name1 + '.pkl', 'wb'))

        # Need to put the next 3 lines if not memory will run out
        del model
        if loss_mode == 'normal' or loss_mode == 'ann':
            K.clear_session()
            sess.close()
        gc.collect()

        # Preparing data to put into new_df that consists of all the validation dataset and its predicted labels
        folds.extend([fold] * i_ss_fl.count)  # Make a col that contains the fold number for each example
        if len(val_features_c):
            val_features_c = np.concatenate((val_features_c, i_ss_fl.features_c),
                                            axis=0)
        else:
            val_features_c = i_ss_fl.features_c

        val_labels.extend(i_ss_fl.labels)
        val_idx.extend(i_ss_fl.idx)

        # Printing one instance summary.
        instance_end = time.time()
        print(
            '\nFor k-fold run {} out of {}. Each fold has {} examples. Model is {} with {} loss. Time taken for '
            'instance = {}\n'
            'Post-training results: \nmse = {}, mse_norm = {}. Scoring is {}\n'
            '####################################################################################################'
                .format(fold + 1, k_folds, i_ss_fl.count, model_mode, loss_mode, instance_end - instance_start, mse,
                        mse_norm, scoring))

    ws = wb[wb.sheetnames[0]]
    df = pd.DataFrame(data=np.array([msee_store, mre_store]).T, columns=['mse', 're'],
                      index=range(1, 1 + len(msee_store)))
    df.insert(0, 'Fold', list(range(len(fl_store))))
    print_df_to_excel(df=df, ws=ws)
    wb.save(te_sheet)

    mse_avg = np.average(mse_store)
    mse_norm_avg = np.average(mse_norm_store)
    re = np.average(np.abs(np.array(val_labels) - np.array(predicted_labels_store))/np.array(val_labels))

    # Calculating metrics based on complete validation prediction
    mse_full = mean_squared_error(val_labels, predicted_labels_store)
    try:
        mse_norm_full = mean_squared_error(fl.labels_scaler.transform(val_labels),
                                           fl.labels_scaler.transform(predicted_labels_store))
    except AttributeError:
        mse_norm_full=mse_full

    # Creating dataframe to print into excel later.
    new_df = np.concatenate((np.array(folds)[:, None],  # Convert 1d list to col. vector
                             val_features_c,
                             np.array(val_labels),
                             np.array(predicted_labels_store))
                            , axis=1)
    if fl.label_type == 'points':
        predicted_labels_name = list(map(str, np.arange(2,101)))
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = ['folds'] + \
                  list(map(str, fl.features_c_names)) + \
                  list(map(str, np.arange(2,101))) + \
                  predicted_labels_name
    elif fl.label_type == 'cutoff':
        predicted_labels_name = list(fl.labels_names)
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = ['folds'] + \
                  list(map(str, fl.features_c_names)) + \
                  list(fl.labels_names) + \
                  predicted_labels_name

    # val_idx is the original position of the example in the data_loader
    new_df = pd.DataFrame(data=new_df, columns=headers, index=val_idx)

    print('Writing into' + skf_file)
    wb = load_workbook(skf_file)

    # Creating new worksheet. Even if SNN worksheet already exists, a new SNN1 ws will be created and so on
    if skf_sheet is None:
        wb.create_sheet(model_mode)
    else:
        wb.create_sheet(model_mode + skf_sheet)
    sheet_name = wb.sheetnames[-1]  # Taking the ws name from the back ensures that if SNN1 is the new ws, it works

    # Writing hparam dataframe first
    pd_writer = pd.ExcelWriter(skf_file, engine='openpyxl')
    pd_writer.book = wb
    pd_writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    new_df.to_excel(pd_writer, sheet_name)
    start_col = len(new_df.columns) + 4
    hparams = pd.DataFrame(dict([(k, Series(v)) for k, v in hparams.items()]))
    hparams.to_excel(pd_writer, sheet_name, startrow=0, startcol=start_col - 1)
    start_row = 5

    # Writing other subset split, instance per run, and bounds
    ws = wb[sheet_name]
    headers = ['mse', 'mse_norm', 're']
    values = [mse_avg, mse_norm_avg]
    values_full = [mse_full, mse_norm_full, re]
    print_array_to_excel(np.array(headers), (1 + start_row, start_col + 1), ws, axis=1)
    print_array_to_excel(np.array(values), (2 + start_row, start_col + 1), ws, axis=1)
    print_array_to_excel(np.array(values_full), (3 + start_row, start_col + 1), ws, axis=1)
    ws.cell(2 + start_row, start_col).value = 'Folds avg'
    ws.cell(3 + start_row, start_col).value = 'Overall'
    pd_writer.save()
    pd_writer.close()
    wb.close()

    if scoring == 'mse':
        return mse_full
    elif scoring == 're':
        return re
    else:
        raise KeyError('Scoring function {} is not valid'.format(scoring))


def eval_model_on_fl(model, fl, return_df=True, label_scaler=None):
    p_y, _, _ = model.eval(fl)
    if fl.label_type == 'cutoff':
        if label_scaler:
            p_y = label_scaler(p_y)
        for row, p_label in enumerate(p_y.tolist()):
            if p_label[1] > p_label[2]:
                p_y[row, 1] = p_y[row, 2]
            if p_label[0] > p_y[row, 1]:
                p_y[row, 0] = p_y[row, 1]
    se_store = (fl.labels - p_y) ** 2
    re_store = (np.abs(fl.labels - p_y).T/fl.labels[:,-1]).T
    if return_df:
        df = pd.DataFrame(data=np.concatenate((fl.labels, p_y), axis=1),
                                  index=list(range(1, 1 + fl.count)),
                                  columns=list(fl.labels_names)
                                          + ['P_{}'.format(col) for col in fl.labels_names])
        return p_y, df, np.mean(se_store), np.mean(re_store)
    else:
        return p_y, np.mean(se_store), np.mean(re_store)


def run_skf_train_val_test_error(model_mode, loss_mode, fl, fl_store, test_fl, ett_fl_store,
                                 hparams, model_name, scoring='mse',
            k_folds=10,
            save_model=False, save_model_name=None, save_model_dir=None,
            plot_name=None):
    '''
    Stratified k fold cross validation for training and evaluating model 2 only. Model 1 data is trained before hand.
    :param model_mode: Choose between using SNN or cDNN (non_smiles) and SNN_smiles or cDNN_smiles
    :param cv_mode: Cross validation mode. Either 'skf' or 'loocv'.
    :param hparams: hparams dict containing hyperparameters information
    :param loader_file: data_loader excel file location
    :param skf_file: skf_file name to save excel file as
    :param skf_sheet: name of sheet to save inside the skf_file excel. If None, will default to SNN or cDNN as name
    :param k_folds: Number of k folds. Used only for skf cv_mode
    :param k_shuffle: Whether to shuffle the given examples to split into k folds if using skf
    :return:
    '''
    fn=6
    numel=3
    # Run k model instance to perform skf
    predicted_labels_store = []
    mse_store = []
    mre_store = []
    folds = []
    val_idx = []
    val_features_c = []
    val_labels = []
    column_headers = fl.labels_names

    str_p_y_store = []
    str_df_store = []
    str_mse_store = []
    str_mre_store = []

    st_p_y_store = []
    st_df_store = []
    st_mse_store = []
    st_mre_store = []

    stt_p_y_store = []
    stt_df_store = []
    stt_mse_store = []
    stt_mre_store = []

    sett_p_y_store = []
    sett_df_store = []
    sett_mse_store = []
    sett_mre_store = []


    for fold, fl_tuple in enumerate(fl_store):
        instance_start = time.time()
        (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl

        # Set up model
        if loss_mode == 'ann':
            sess = tf.compat.v1.Session()
            # sess = tf.Session()
            K.set_session(sess)
            model = Kmodel(fl=ss_fl, mode=model_mode, hparams=hparams)
        elif loss_mode == 'svr':
            if not fl.normalise_labels:
                raise TypeError('fl labels are not normalised. For SVR, the labels must be normalised.')
            model = SVRmodel(fl=ss_fl, epsilon=hparams['epsilon'], c=hparams['c'])
        elif loss_mode == 'dtr':
            #if not fl.normalise_labels:
                #raise TypeError('fl labels are not normalised. For SVR, the labels must be normalised.')
            model = DTRmodel(fl=ss_fl, max_depth=hparams['max_depth'], num_est=hparams['num_est'])
        elif loss_mode == 'xgb':
            model = XGBmodel(fl=ss_fl,  hparams=hparams)
        else:
            raise KeyError('loss_mode ' + loss_mode + 'is not a valid selection for loss mode.')

        # Train model and save model training loss vs epoch plot if plot_name is given, else no plot will be saved
        if plot_name:
            model.train_model(ss_fl, i_ss_fl, plot_name='{}_fold_{}.png'.format(plot_name, fold))
        else:
            model.train_model(ss_fl, i_ss_fl)

        str_p_y, str_df, str_mse, str_mre = eval_model_on_fl(model, ss_fl, return_df=True)
        str_p_y_store.append(str_p_y)
        str_df_store.append(str_df)
        str_mse_store.append(str_mse)
        str_mre_store.append(str_mre)
        st_p_y, st_df, st_mse, st_mre = eval_model_on_fl(model, fl, return_df=True)
        st_p_y_store.append(st_p_y)
        st_df_store.append(st_df)
        st_mse_store.append(st_mse)
        st_mre_store.append(st_mre)
        stt_p_y, stt_df, stt_mse, stt_mre = eval_model_on_fl(model, test_fl, return_df=True)
        stt_p_y_store.append(stt_p_y)
        stt_df_store.append(stt_df)
        stt_mse_store.append(stt_mse)
        stt_mre_store.append(stt_mre)

        p_y_store = []
        df_store = []
        mse_store = []
        mre_store = []

        for ett_fl in ett_fl_store:
            p_y, df, mse, mre = eval_model_on_fl(model, ett_fl, return_df=True)
            p_y_store.append(p_y)
            df_store.append(df)
            mse_store.append(mse)
            mre_store.append(mre)

        sett_p_y_store.append(p_y_store)
        sett_df_store.append(df_store)
        sett_mse_store.append(mse_store)
        sett_mre_store.append(mre_store)

        p_y, mse, mre = eval_model_on_fl(model, i_ss_fl, return_df=False)
        predicted_labels_store.extend(p_y)
        mse_store.append(mse)
        mre_store.append(mre)

        # Saving model
        if save_model:
            # Set save_model_name
            if isinstance(save_model_name, str):
                save_model_name1 = save_model_name + '_' + model_mode + '_' + str(fold + 1)
            else:
                save_model_name1 = model_mode + '_' + str(fold + 1)

            # Save model
            print('Saving instance {} model in {}'.format(fold + 1, save_model_dir + save_model_name1 + '.h5'))
            if loss_mode == 'normal' or loss_mode == 'ann':
                model.model.save(save_model_dir + save_model_name1 + '.h5')
            elif loss_mode == 'svr' or loss_mode == 'dtr':
                pickle.dump(Predict_SVR_DTR(model=model.model, labels_scaler=model.labels_scaler),
                            open(save_model_dir + save_model_name1 + '.pkl', 'wb'))

        # Need to put the next 3 lines if not memory will run out
        del model
        if loss_mode == 'normal' or loss_mode == 'ann':
            K.clear_session()
            sess.close()
        gc.collect()

        # Preparing data to put into new_df that consists of all the validation dataset and its predicted labels
        folds.extend([fold] * i_ss_fl.count)  # Make a col that contains the fold number for each example
        if len(val_features_c):
            val_features_c = np.concatenate((val_features_c, i_ss_fl.features_c),
                                            axis=0)
        else:
            val_features_c = i_ss_fl.features_c

        val_labels.extend(i_ss_fl.labels)
        val_idx.extend(i_ss_fl.idx)

        # Printing one instance summary.
        instance_end = time.time()
        print(
            '\nFor k-fold run {} out of {}. Each fold has {} examples. Model is {} with {} loss. Time taken for '
            'instance = {}\n'
            'Post-training results: \nmse = {}, mre = {}. Scoring is {}\n'
            '####################################################################################################'
                .format(fold + 1, k_folds, i_ss_fl.count, model_mode, loss_mode, instance_end - instance_start, mse,
                        mre, scoring))

    # Calculating metrics based on complete validation prediction
    val_labels = np.array(val_labels)
    predicted_labels_store = np.array(predicted_labels_store)
    mse_full = mean_squared_error(val_labels, predicted_labels_store)
    mre_full = np.mean(np.abs(val_labels-predicted_labels_store).T / val_labels[:,-1])

    # Train set
    p_y = np.mean(np.array(st_p_y_store), axis=0)
    train_mse = mean_squared_error(fl.labels, p_y)
    train_mre = np.mean(np.abs(fl.labels-p_y).T / fl.labels[:,-1])
    new_df = np.concatenate((fl.labels, p_y) , axis=1)
    predicted_labels_name = list(fl.labels_names)
    predicted_labels_name = ['P_' + x for x in predicted_labels_name]
    headers = list(fl.labels_names) + predicted_labels_name
    train_df = pd.DataFrame(data=new_df, columns=headers)

    # Test set
    p_y = np.mean(np.array(stt_p_y_store), axis=0)
    test_mse = mean_squared_error(test_fl.labels, p_y)
    test_mre = np.mean(np.abs(test_fl.labels-p_y).T / test_fl.labels[:,-1])
    new_df = np.concatenate((test_fl.labels, p_y) , axis=1)
    predicted_labels_name = list(fl.labels_names)
    predicted_labels_name = ['P_' + x for x in predicted_labels_name]
    headers = list(fl.labels_names) + predicted_labels_name
    test_df = pd.DataFrame(data=new_df, columns=headers)

    # Extra testset
    ett_df_store = []
    ett_mse_store = []
    ett_mre_store = []
    sett_p_y_store = [[x[idx] for x in sett_p_y_store] for idx in range(len(ett_fl_store))]


    for ett_fl,ett_p_y in zip(ett_fl_store, sett_p_y_store):
        p_y = np.mean(np.array(ett_p_y), axis=0)
        ett_mse_store.append(mean_squared_error(ett_fl.labels, p_y))
        ett_mre_store.append(np.mean(np.abs(ett_fl.labels - p_y).T / ett_fl.labels[:,-1]))
        new_df = np.concatenate((ett_fl.labels, p_y), axis=1)
        predicted_labels_name = list(fl.labels_names)
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = list(fl.labels_names) + predicted_labels_name
        ett_df_store.append(pd.DataFrame(data=new_df, columns=headers))


    # Creating dataframe to print into excel later.
    new_df = np.concatenate((np.array(folds)[:, None],  # Convert 1d list to col. vector
                             val_features_c,
                             np.array(val_labels),
                             np.array(predicted_labels_store))
                            , axis=1)
    if fl.label_type == 'points':
        predicted_labels_name = list(map(str, np.arange(2,101)))
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = ['folds'] + \
                  list(map(str, fl.features_c_names)) + \
                  list(map(str, np.arange(2,101))) + \
                  predicted_labels_name
    elif fl.label_type == 'cutoff':
        predicted_labels_name = list(fl.labels_names)
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = ['folds'] + \
                  list(map(str, fl.features_c_names)) + \
                  list(fl.labels_names) + \
                  predicted_labels_name

    # val_idx is the original position of the example in the data_loader
    val_df = pd.DataFrame(data=new_df, columns=headers, index=val_idx).sort_index()

    model_names = ['{}_{}'.format(model_name, x) for x in range(1, fold+2)]
    data = [[model_names, st_mse_store, st_mre_store, stt_mse_store, stt_mre_store],
            [model_name, train_mse, train_mre, mse_full, mre_full, test_mse, test_mre],
            st_df_store,
            stt_df_store,
            train_df,
            val_df,
            test_df,
            [sett_mse_store, sett_mre_store],
            [ett_mse_store, ett_mre_store],
            sett_df_store,
            ett_df_store,
            [str_p_y_store, str_df_store, str_mse_store, str_mre_store]]

    if scoring == 'mse':
        return mse_full, train_mse, data
    elif scoring == 're':
        return mre_full, train_mre, data
    else:
        raise KeyError('Scoring function {} is not valid'.format(scoring))


def run_eval_model_on_train_val_test_error(model, fl, fl_store, test_fl, ett_fl_store, model_name, scoring='mse',):
    # Run k model instance to perform skf
    predicted_labels_store = []
    mse_store = []
    mre_store = []
    folds = []
    val_idx = []
    val_features_c = []
    val_labels = []
    column_headers = fl.labels_names

    str_p_y_store = []
    str_df_store = []
    str_mse_store = []
    str_mre_store = []

    st_p_y_store = []
    st_df_store = []
    st_mse_store = []
    st_mre_store = []

    stt_p_y_store = []
    stt_df_store = []
    stt_mse_store = []
    stt_mre_store = []

    sett_p_y_store = []
    sett_df_store = []
    sett_mse_store = []
    sett_mre_store = []

    def eval_model_ensemble_on_fl(model, fl, return_df=True, label_scaler=None):
        p_y = model.predict(fl.features_c_norm)
        if label_scaler:
            p_y = label_scaler(p_y)
        for row, p_label in enumerate(p_y.tolist()):
            if p_label[1] > p_label[2]:
                p_y[row, 1] = p_y[row, 2]
            if p_label[0] > p_y[row, 1]:
                p_y[row, 0] = p_y[row, 1]
        se_store = (fl.labels - p_y) ** 2
        re_store = (np.abs(fl.labels - p_y).T / fl.labels[:, -1]).T
        if return_df:
            df = pd.DataFrame(data=np.concatenate((fl.labels, p_y), axis=1),
                              index=list(range(1, 1 + fl.count)),
                              columns=list(fl.labels_names)
                                      + ['P_{}'.format(col) for col in fl.labels_names])
            return p_y, df, np.mean(se_store), np.mean(re_store)
        else:
            return p_y, np.mean(se_store), np.mean(re_store)

    for fold, fl_tuple in enumerate(fl_store):
        instance_start = time.time()
        (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl



        str_p_y, str_df, str_mse, str_mre = eval_model_ensemble_on_fl(model, ss_fl, return_df=True)
        str_p_y_store.append(str_p_y)
        str_df_store.append(str_df)
        str_mse_store.append(str_mse)
        str_mre_store.append(str_mre)
        st_p_y, st_df, st_mse, st_mre = eval_model_ensemble_on_fl(model, fl, return_df=True)
        st_p_y_store.append(st_p_y)
        st_df_store.append(st_df)
        st_mse_store.append(st_mse)
        st_mre_store.append(st_mre)
        stt_p_y, stt_df, stt_mse, stt_mre = eval_model_ensemble_on_fl(model, test_fl, return_df=True)
        stt_p_y_store.append(stt_p_y)
        stt_df_store.append(stt_df)
        stt_mse_store.append(stt_mse)
        stt_mre_store.append(stt_mre)

        p_y_store = []
        df_store = []
        mse_store = []
        mre_store = []

        for ett_fl in ett_fl_store:
            p_y, df, mse, mre = eval_model_ensemble_on_fl(model, ett_fl, return_df=True)
            p_y_store.append(p_y)
            df_store.append(df)
            mse_store.append(mse)
            mre_store.append(mre)

        sett_p_y_store.append(p_y_store)
        sett_df_store.append(df_store)
        sett_mse_store.append(mse_store)
        sett_mre_store.append(mre_store)

        p_y, mse, mre = eval_model_ensemble_on_fl(model, i_ss_fl, return_df=False)
        predicted_labels_store.extend(p_y)
        mse_store.append(mse)
        mre_store.append(mre)

        # Preparing data to put into new_df that consists of all the validation dataset and its predicted labels
        folds.extend([fold] * i_ss_fl.count)  # Make a col that contains the fold number for each example
        if len(val_features_c):
            val_features_c = np.concatenate((val_features_c, i_ss_fl.features_c),
                                            axis=0)
        else:
            val_features_c = i_ss_fl.features_c

        val_labels.extend(i_ss_fl.labels)
        val_idx.extend(i_ss_fl.idx)

    # Calculating metrics based on complete validation prediction
    val_labels = np.array(val_labels)
    predicted_labels_store = np.array(predicted_labels_store)
    mse_full = mean_squared_error(val_labels, predicted_labels_store)
    mre_full = np.mean(np.abs(val_labels - predicted_labels_store).T / val_labels[:, -1])

    # Train set
    p_y = np.mean(np.array(st_p_y_store), axis=0)
    train_mse = mean_squared_error(fl.labels, p_y)
    train_mre = np.mean(np.abs(fl.labels - p_y).T / fl.labels[:, -1])
    new_df = np.concatenate((fl.labels, p_y), axis=1)
    predicted_labels_name = list(fl.labels_names)
    predicted_labels_name = ['P_' + x for x in predicted_labels_name]
    headers = list(fl.labels_names) + predicted_labels_name
    train_df = pd.DataFrame(data=new_df, columns=headers)

    # Test set
    p_y = np.mean(np.array(stt_p_y_store), axis=0)
    test_mse = mean_squared_error(test_fl.labels, p_y)
    test_mre = np.mean(np.abs(test_fl.labels - p_y).T / test_fl.labels[:, -1])
    new_df = np.concatenate((test_fl.labels, p_y), axis=1)
    predicted_labels_name = list(fl.labels_names)
    predicted_labels_name = ['P_' + x for x in predicted_labels_name]
    headers = list(fl.labels_names) + predicted_labels_name
    test_df = pd.DataFrame(data=new_df, columns=headers)

    # Extra testset
    ett_df_store = []
    ett_mse_store = []
    ett_mre_store = []
    sett_p_y_store = [[x[idx] for x in sett_p_y_store] for idx in range(len(ett_fl_store))]

    for ett_fl, ett_p_y in zip(ett_fl_store, sett_p_y_store):
        p_y = np.mean(np.array(ett_p_y), axis=0)
        ett_mse_store.append(mean_squared_error(ett_fl.labels, p_y))
        ett_mre_store.append(np.mean(np.abs(ett_fl.labels - p_y).T / ett_fl.labels[:, -1]))
        new_df = np.concatenate((ett_fl.labels, p_y), axis=1)
        predicted_labels_name = list(fl.labels_names)
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = list(fl.labels_names) + predicted_labels_name
        ett_df_store.append(pd.DataFrame(data=new_df, columns=headers))

    # Creating dataframe to print into excel later.
    new_df = np.concatenate((np.array(folds)[:, None],  # Convert 1d list to col. vector
                             val_features_c,
                             np.array(val_labels),
                             np.array(predicted_labels_store))
                            , axis=1)
    if fl.label_type == 'points':
        predicted_labels_name = list(map(str, np.arange(2, 101)))
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = ['folds'] + \
                  list(map(str, fl.features_c_names)) + \
                  list(map(str, np.arange(2, 101))) + \
                  predicted_labels_name
    elif fl.label_type == 'cutoff':
        predicted_labels_name = list(fl.labels_names)
        predicted_labels_name = ['P_' + x for x in predicted_labels_name]
        headers = ['folds'] + \
                  list(map(str, fl.features_c_names)) + \
                  list(fl.labels_names) + \
                  predicted_labels_name

    # val_idx is the original position of the example in the data_loader
    val_df = pd.DataFrame(data=new_df, columns=headers, index=val_idx).sort_index()

    model_names = ['{}_{}'.format(model_name, x) for x in range(1, fold + 2)]
    data = [[model_names, st_mse_store, st_mre_store, stt_mse_store, stt_mre_store],
            [model_name, train_mse, train_mre, mse_full, mre_full, test_mse, test_mre],
            st_df_store,
            stt_df_store,
            train_df,
            val_df,
            test_df,
            [sett_mse_store, sett_mre_store],
            [ett_mse_store, ett_mre_store],
            sett_df_store,
            ett_df_store,
            [str_p_y_store, str_df_store, str_mse_store, str_mre_store]]

    if scoring == 'mse':
        return mse_full, train_mse, data
    elif scoring == 're':
        return mre_full, train_mre, data
    else:
        raise KeyError('Scoring function {} is not valid'.format(scoring))

