from tensorflow.python.keras import backend as K

from tensorflow.python.keras.models import load_model
from keras.utils import CustomObjectScope
from keras.initializers import glorot_uniform

import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os, time, gc, pickle, itertools, math
from typing import List
from collections import Counter

from skopt import gp_minimize, gbrt_minimize, forest_minimize, dummy_minimize
from skopt.space import Real, Integer, Categorical
from skopt.utils import use_named_args
from skopt.plots import plot_convergence
from sklearn.metrics import pairwise_distances, pairwise_distances_chunked
from own_package.features_labels_setup import load_data_to_fl
from own_package.others import print_array_to_excel, create_results_directory, print_df_to_excel



def load_svm_ensemble(model_directory) -> List:
    """
    Load list of trained svm models from a pickle saved file that can be used for prediction later
    :param model_directory: model directory where the h5 models are saved in. NOTE: All the models in the directory will
     be loaded. Hence, make sure all the models in there are part of the ensemble and no unwanted models are in the
     directory
    :return: [List: svm models]
    """

    # Loading model names into a list
    model_name_store = []
    directory = model_directory
    for idx, file in enumerate(os.listdir(directory)):
        filename = os.fsdecode(file)
        model_name_store.append(directory + '/' + filename)
    print('Loading the following models from {}. Total models = {}'.format(directory, len(model_name_store)))

    # Loading model class object into a list
    model_store = []
    for name in model_name_store:
        with open(name, "rb") as input_file:
            model_store.append(pickle.load(input_file))
        print('Model {} has been loaded'.format(name))

    return model_store


def svm_ensemble_prediction(model_store, composition):
    """
    Run prediction given one set of feactures_c_norm input, using all the models in model store.
    :param model_store: List of keras models returned by the def load_model_ensemble
    :param Composition: ndarray of shape (-1, 2). The columns represents the CNT and PVA composition.
    :return: List of metrics.
    """
    predictions_store = []
    distance_store = []
    if len(composition.shape) == 1:
        composition = composition.reshape(1, 2)
    for model in model_store:
        predictions_store.append(model.model.predict(composition))
        distance_store.append(model.model.decision_function(composition))

    predictions = np.round(np.average(np.array(predictions_store), axis=0), decimals=0)
    distance = np.mean(np.array(distance_store), axis=0)

    return predictions, distance


def load_model_ensemble(model_directory) -> List:
    """
    Load list of trained keras models from a .h5 saved file that can be used for prediction later
    :param model_directory: model directory where the h5 models are saved in. NOTE: All the models in the directory will
     be loaded. Hence, make sure all the models in there are part of the ensemble and no unwanted models are in the
     directory
    :return: [List: keras models]
    """

    # Loading model names into a list
    model_name_store = []
    directory = model_directory
    for idx, file in enumerate(os.listdir(directory)):
        filename = os.fsdecode(file)
        model_name_store.append(directory + '/' + filename)
    print('Loading the following models from {}. Total models = {}'.format(directory, len(model_name_store)))

    # Loading model class object into a list
    model_store = []
    for name in model_name_store:
        if name.endswith(".pkl"):
            model_store.append(pickle.load(open(name, 'rb')))
        elif name.endswith('.h5'):
            with CustomObjectScope({'GlorotUniform': glorot_uniform()}):
                model_store.append(load_model(name))

        else:
            raise TypeError('{} found that does not end with .pkl or .h5'.format(name))
        print('Model {} has been loaded'.format(name))

    return model_store


def load_model_chunks(chunks):
    model_store = []
    for name in chunks:
        if name.endswith(".pkl"):
            try:
                model_store.append(pickle.load(open(name, 'rb')))
            except EOFError:
                print('EOFError with {}'.format(name))
                model_store.append(None)
        elif name.endswith('.h5'):
            with CustomObjectScope({'GlorotUniform': glorot_uniform()}):
                model_store.append(load_model(name))
        print('Model {} has been loaded'.format(name))
    return model_store

def model_ensemble_prediction(model_store, features_c_norm):
    """
    Run prediction given one set of feactures_c_norm input, using all the models in model store.
    :param model_store: List of keras models returned by the def load_model_ensemble
    :param features_c_norm: ndarray of shape (1, -1). The columns represents the different features.
    :return: List of metrics.
    """
    predictions_store = []
    for model in model_store:
        predictions = model.predict(features_c_norm).tolist()
        predictions_store.append(predictions)
    predictions_store = np.array(predictions_store).squeeze()
    predictions_mean = np.mean(predictions_store, axis=0)
    predictions_std = np.std(predictions_store, axis=0)

    return predictions_mean, predictions_std


def features_to_features_input(fl, features_c, features_d) -> np.ndarray:
    """

    :param fl:
    :param features_c:
    :param features_d:
    :return:
    """
    lookup_df_store = fl.lookup_df_store
    # Filling up features descriptors for each discrete feature
    features_dc_store = []
    for item, lookup_df in zip(features_d, lookup_df_store):
        features_dc_store.extend(lookup_df.loc[item, :])

    return np.array(features_c.tolist() + features_dc_store)


def acquisition_opt(bounds, model_directory, svm_directory, loader_file, total_run, normalise_labels, batch_runs=1,
                    norm_mask=None,
                    acquisition_file='./excel/acquisition_opt.xlsx'):
    """
    bounds = [[5, 200, ],
              [0, 200, ],
              [5, 200, ],
              [0, 200, ],
              [10, 2000],
              [0, 0.3]]
    :param model_mode:
    :param loader_file:
    :param total_run:
    :param instance_per_run:
    :param hparam_file:
    :return:
    """
    while os.path.isfile(acquisition_file):
        expand = 1
        while True:
            expand += 1
            new_file_name = acquisition_file.split('.xlsx')[0] + ' - ' + str(expand) + '.xlsx'
            if os.path.isfile(new_file_name):
                continue
            else:
                acquisition_file = new_file_name
                break
    print('Writing into' + acquisition_file + '\n')
    wb = openpyxl.Workbook()
    wb.save(acquisition_file)

    model_store = load_model_ensemble(model_directory)
    svm_store = load_svm_ensemble(svm_directory)

    fl = load_data_to_fl(loader_file, norm_mask=norm_mask, normalise_labels=normalise_labels, label_type='cutoff')

    space = [Real(low=bounds[0][0], high=bounds[0][1], name='CNT'),
             Real(low=bounds[1][0], high=bounds[1][1], name='PVA'),
             Real(low=bounds[2][0], high=bounds[2][1], name='Thickness'),
             Categorical(categories=[0, 1, 2], name='Dimension')]

    for batch in range(batch_runs):
        instance_start = time.time()
        prediction_mean_store = []
        prediction_std_store = []
        l2_dist_store = []
        disagreement_store = []
        global iter_count
        iter_count = 0

        @use_named_args(space)
        def fitness(**params):
            global iter_count
            iter_count += 1
            if iter_count % 50 == 0:
                print('Current Iteration: {} out of {} for batch {} '.format(iter_count, total_run, batch + 1))

            features = np.array([x for x in params.values()])
            x = features[0]
            y = features[1]
            if x + y > 1:
                u = -y + 1
                v = -x + 1
                features[0:2] = np.array([u, v])

            # SVM Check
            p_class, distance = svm_ensemble_prediction(svm_store, features[0:2])

            if distance.item() < 0:
                # Distance should be negative value when SVM assigns class 0. Hence a_score will be negative.
                # The more negative the a_score is, the further the composition is from the hyperplane,
                # hence, the less likely the optimizer will select examples with class 0.
                a_score = 10e5 * distance.item()
                prediction_mean_store.append([-1] * fl.labels_dim)
                prediction_std_store.append([-1] * fl.labels_dim)
                l2_dist_store.append(-1)
                disagreement_store.append(-1)
            elif features[0] + features[1] > 1:
                # Distance should be negative value when SVM assigns class 0. Hence a_score will be negative.
                # The more negative the a_score is, the further the composition is from the hyperplane,
                # hence, the less likely the optimizer will select examples with class 0.
                a_score = 10e5 * (1 - (features[0] + features[1]))
                prediction_mean_store.append([-1] * fl.labels_dim)
                prediction_std_store.append([-1] * fl.labels_dim)
                l2_dist_store.append(-1)
                disagreement_store.append(-1)
            else:
                features_c = features[:-1]
                onehot = features[-1].item()
                if onehot == 0:
                    features = np.concatenate((features_c, np.array([1, 0, 0])))
                elif onehot == 1:
                    features = np.concatenate((features_c, np.array([0, 1, 0])))
                elif onehot == 2:
                    features = np.concatenate((features_c, np.array([0, 0, 1])))

                features_input_norm = fl.apply_scaling(features)
                prediction_mean, prediction_std = model_ensemble_prediction(model_store, features_input_norm)
                # Greedy Sampling
                # Get L2 distance of sampled example to all existing example in fl class object
                # Note: L2 distance is calculated using the normalised features so that all feature have the same weight
                l2_distance = np.linalg.norm(x=fl.features_c_norm - features_input_norm.reshape((1, -1)), ord=2, axis=1)
                l2_distance = np.min(l2_distance)  # Take the minimum L2 dist.

                # Overall Acquisition Score. Higher score if l2 distance is larger and uncertainty (std) is larger.
                disagreement = np.sum(prediction_std)
                a_score = l2_distance * disagreement

                # Storing intermediate results into list to print into excel later
                l2_dist_store.append(l2_distance)
                disagreement_store.append(disagreement)
                prediction_mean_store.append(prediction_mean.flatten().tolist())
                prediction_std_store.append(prediction_std.flatten().tolist())
            return -a_score

        '''
        search_result = forest_minimize(func=fitness,
                                        dimensions=space,
                                        acq_func='EI',  # Expected Improvement.
                                        n_calls=total_run,
                                        verbose=False)
        '''

        search_result = dummy_minimize(func=fitness,
                                       dimensions=space,
                                       n_calls=total_run,
                                       verbose=False)

        plot_convergence(search_result)
        x_iters = search_result.x_iters
        for idx, row in enumerate(x_iters):
            x, y = row[0:2]
            if x + y > 1:
                u = -y + 1
                v = -x + 1
                x_iters[idx][0:2] = [u, v]

        func_val = -search_result.func_vals
        best_x = search_result.x

        p_mean_name = np.array(['Pmean_' + str(x) for x in list(map(str, np.arange(1, 4)))])
        p_std_name = np.array(['Pstd_' + str(x) for x in list(map(str, np.arange(1, 4)))])
        p_sm_name = np.array(['Ps/m_' + str(x) for x in list(map(str, np.arange(1, 4)))])

        mean = np.array(prediction_mean_store)
        std = np.array(prediction_std_store)
        sm = np.array(prediction_std_store) / np.array(prediction_mean_store)
        a2 = (np.array(l2_dist_store) * np.sum(sm, axis=1)).reshape((-1, 1))

        data = np.concatenate((np.array(x_iters),
                               func_val.reshape((-1, 1)),
                               np.array(disagreement_store).reshape((-1, 1)),
                               np.array(l2_dist_store).reshape((-1, 1)),
                               mean,
                               std,
                               sm,
                               a2
                               ), axis=1)

        columns = np.concatenate((np.array(fl.features_c_names[:-2]),
                                  np.array(['A_score']),
                                  np.array(['Disagreement']),
                                  np.array(['L2']),
                                  p_mean_name,
                                  p_std_name,
                                  p_sm_name,
                                  np.array(['A_sm_score'])
                                  ))

        iter_df = pd.DataFrame(data=data,
                               columns=columns)

        iter_df = iter_df.sort_values(by=['A_score'], ascending=False)

        # Creating new worksheet. Even if SNN worksheet already exists, a new SNN1 ws will be created and so on
        wb.create_sheet(title='Acquisition_opt')
        sheet_name = wb.sheetnames[-1]  # Taking the ws name from the back ensures that if SNN1 is the new ws, it works

        # Writing hparam dataframe first
        pd_writer = pd.ExcelWriter(acquisition_file, engine='openpyxl')
        pd_writer.book = wb
        pd_writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        iter_df.to_excel(pd_writer, sheet_name)
        pd_writer.save()
        pd_writer.close()
        wb.close()

        # If batch_runs > 1, next batch will be calculated. The only difference is that the previous best trial point
        # with the highest a_score will be added to fl.features_c_norm such that the L2 greedy distance will
        # account for the fact that the previous batch would had contained the best example already.
        features = np.array(best_x)
        features_c = features[:-1]
        onehot = features[-1].item()
        if onehot == 0:
            features = np.concatenate((features_c, np.array([1, 0, 0])))
        elif onehot == 1:
            features = np.concatenate((features_c, np.array([0, 1, 0])))
        elif onehot == 2:
            features = np.concatenate((features_c, np.array([0, 0, 1])))

        fl.features_c_norm = np.concatenate((fl.features_c_norm, fl.apply_scaling(features)), axis=0)

        instance_end = time.time()
        print('Batch {} completed. Time taken: {}'.format(batch + 1, instance_end - instance_start))


def l2_points_opt(numel, write_dir, svm_directory, seed_number_of_expt, total_expt):
    write_dir = create_results_directory(results_directory=write_dir, excels=['l2_acq'])
    svm_store = load_svm_ensemble(svm_directory)
    base = [x / (numel * 2 - 1) for x in list(range(numel * 2))]
    compositions = [[x, y] if x + y <= 1 else [-x + 1, -y + 1] for x, y in
                    list(itertools.product(base[::2], base[1::2]))]

    distance_store = []
    for model in svm_store:
        distance_store.append(model.model.decision_function(compositions))

    distance = np.mean(np.array(distance_store), axis=0)
    valid_compositions = [x for x, dist in zip(compositions, distance) if dist >= 0]

    print('Number of compositions = {}. % valid = {}%'.format(len(valid_compositions),
                                                             len(valid_compositions) / len(compositions)*100))

    number_valid_compositions = round(math.sqrt(len(valid_compositions)))
    compositions_thickness = list(itertools.product(valid_compositions,
                                                    [x / (number_valid_compositions- 1)
                                                     for x in list(range(number_valid_compositions))]))

    print('Number of permutations = {}'.format(len(compositions_thickness*3)))

    all_permutations = np.array([x[0] + [x[1]] + y
                                 for x in compositions_thickness for y in [[1, 0, 0], [0, 1, 0], [0, 0, 1]]])

    expt_idx = np.random.randint(0, len(all_permutations), seed_number_of_expt)

    expt_store = all_permutations[expt_idx,:]

    for i in range(total_expt-seed_number_of_expt):
        start = time.time()
        d = pairwise_distances(expt_store, all_permutations, metric='euclidean')
        next_expt = np.argmax(np.min(d, axis=0))
        expt_store = np.concatenate((expt_store, all_permutations[next_expt, None, :]), axis=0)
        end = time.time()
        print('{} out of {} completed. Time taken = {}.'.format(i+1,total_expt-seed_number_of_expt, end-start))

    expt_store[:,2] = expt_store[:,2]*2000

    write_excel = '{}/l2_acq.xlsx'.format(write_dir)
    wb = openpyxl.load_workbook(write_excel)
    wb.create_sheet('l2_acq')
    ws = wb[wb.sheetnames[-1]]
    ws.cell(1,1).value = 'Valid Combinations'
    ws.cell(1,2).value = len(all_permutations)
    ws.cell(1,3).value = 'Seed Expt'
    ws.cell(1,4).value = seed_number_of_expt
    df = pd.DataFrame(data=expt_store, columns=['CNT', 'PVA', 'Thickness', '0D', '1D', '2D'],
                      index=list(range(1,total_expt+1)))
    print_df_to_excel(df=df, ws=ws, start_row=2)

    wb.save(write_excel)
    pass
