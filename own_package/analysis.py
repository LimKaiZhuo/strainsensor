import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl
import seaborn as sns
import openpyxl
import scipy.stats as stat
import tensorflow as tf
from tensorflow.python.keras import backend as K
import gc

from own_package.models.models import Kmodel
from own_package.others import print_array_to_excel, print_df_to_excel, create_excel_file, create_results_directory
from own_package.active_learning.acquisition import load_model_ensemble, model_ensemble_prediction, load_model_chunks
from own_package.features_labels_setup import load_data_to_fl, load_testset_to_fl
import os, itertools, pickle
from collections import defaultdict
from deap import algorithms, base, creator, tools


def l2_tracker(write_excel, final_excel_loader, last_idx_store):
    wb = openpyxl.load_workbook(write_excel)
    wb.create_sheet('L2 Results')
    ws = wb[wb.sheetnames[-1]]
    fl = load_data_to_fl(data_loader_excel_file=final_excel_loader, normalise_labels=True,
                         label_type='cutoff',
                         norm_mask=[0, 1, 3, 4, 5])
    final_features = fl.features_c_norm

    suggestions_store = [y2 - y1 for y2, y1 in zip(last_idx_store[1:], last_idx_store[:-1])] + [0]

    batch_l2_store = []
    batch_l2_suggestions_store = []
    for last_idx, suggestions_numel in zip(last_idx_store, suggestions_store):
        features = final_features[:last_idx, :].tolist()

        l2_store = []
        for idx, x in enumerate(features):
            other_features = np.array(features[:idx] + features[idx + 1:])
            l2_distance = np.linalg.norm(x=other_features - np.array(x).reshape((1, -1)), ord=2, axis=1)
            l2_store.append(np.min(l2_distance))
        batch_l2_store.append(np.mean(l2_store))

        if suggestions_numel == 0:
            batch_l2_suggestions_store.append(np.NaN)
        else:
            l2_suggestions_store = []
            suggestions_features = final_features[last_idx:last_idx + suggestions_numel].tolist()
            for sf in suggestions_features:
                l2_distance = np.linalg.norm(x=features - np.array(sf).reshape((1, -1)), ord=2, axis=1)
                l2_suggestions_store.append(np.min(l2_distance))
            batch_l2_suggestions_store.append(np.mean(l2_suggestions_store))

    df = pd.DataFrame(data=np.concatenate((np.array(last_idx_store).reshape(-1, 1),
                                           np.array(batch_l2_store).reshape(-1, 1),
                                           np.array(batch_l2_suggestions_store).reshape(-1, 1)), axis=1),
                      columns=['Expt Batches', 'Mean Min L2', 'Suggestions Mean Min L2'],
                      index=range(1, len(last_idx_store) + 1))
    print_df_to_excel(df=df, ws=ws)
    wb.save(write_excel)


def testset_prediction_results(write_excel, model_dir_store, excel_loader_dir_store, testset_excel_dir, rounds, fn,
                               numel):
    wb = openpyxl.load_workbook(write_excel)
    results_col = fn + 1 + 1 + 2 * numel + 3
    mse_store = []
    mre_store = []
    mare_store = []
    testset_fl = load_data_to_fl(testset_excel_dir, normalise_labels=True, label_type='cutoff',
                                 norm_mask=[0, 1, 3, 4, 5])
    column_headers = testset_fl.labels_names
    for idx, (model_dir, loader_excel, round) in enumerate(zip(model_dir_store, excel_loader_dir_store, rounds)):
        wb.create_sheet('Round {}'.format(round))
        ws = wb[wb.sheetnames[-1]]
        fl = load_data_to_fl(data_loader_excel_file=loader_excel, norm_mask=[0, 1, 3, 4, 5],
                             normalise_labels=True, label_type='cutoff')
        model_store = load_model_ensemble(model_dir)
        # Must use the round's fl to scale, not the testset scaler as it might be different
        p_y, _ = model_ensemble_prediction(model_store, fl.apply_scaling(testset_fl.features_c))

        for row, p_label in enumerate(p_y.tolist()):
            if p_label[1] > p_label[2]:
                p_y[row, 1] = p_y[row, 2]
            if p_label[0] > p_y[row, 1]:
                p_y[row, 0] = p_y[row, 1]

        se_store = (testset_fl.labels - p_y) ** 2
        re_store = np.abs(testset_fl.labels - p_y) / testset_fl.labels
        are_store = np.arctan(re_store)

        df = pd.DataFrame(data=np.concatenate((testset_fl.labels, p_y, se_store, re_store, are_store), axis=1),
                          index=list(range(1, 1 + testset_fl.count)),
                          columns=list(column_headers)
                                  + ['P_{}'.format(col) for col in column_headers]
                                  + ['SE_{}'.format(col) for col in column_headers]
                                  + ['RE_{}'.format(col) for col in column_headers]
                                  + ['ARE_{}'.format(col) for col in column_headers])
        print_df_to_excel(df=df, ws=ws)

        col = fn + 1 + 1 + 2 * numel + 3
        mse_store.append(np.mean(se_store))
        mre_store.append(np.mean(re_store))
        mare_store.append(np.mean(are_store))
        ws.cell(1, col).value = 'MSE'
        ws.cell(1, col + 1).value = mse_store[-1]
        ws.cell(2, col).value = 'MRE'
        ws.cell(2, col + 1).value = mre_store[-1]
        ws.cell(3, col).value = 'ARE'
        ws.cell(3, col + 1).value = mare_store[-1]

    wb.create_sheet('Final_results')
    ws = wb[wb.sheetnames[-1]]
    df = pd.DataFrame(data=np.array([mse_store, mre_store, mare_store]), index=['mse', 're', 'are'], columns=rounds)
    print_df_to_excel(df=df, ws=ws)
    wb.save(write_excel)


def testset_model_results_to_excel(write_excel, model_dir_store, loader_excel, testset_excel_dir, fn, numel, chunks):
    wb = openpyxl.load_workbook(write_excel)
    mse_store = []
    mre_store = []
    mare_store = []
    testset_fl = load_data_to_fl(testset_excel_dir, normalise_labels=True, label_type='cutoff',
                                 norm_mask=[0, 1, 3, 4, 5])
    column_headers = testset_fl.labels_names

    fl = load_data_to_fl(data_loader_excel_file=loader_excel, norm_mask=[0, 1, 3, 4, 5],
                         normalise_labels=True, label_type='cutoff')

    model_name_store = []
    for model_dir in model_dir_store:
        for idx, file in enumerate(os.listdir(model_dir)):
            filename = os.fsdecode(file)
            model_name_store.append(model_dir + '/' + filename)
        print('Loading the following models from {}. Total models = {}'.format(model_dir, len(model_name_store)))
    model_chunks = [model_name_store[x:x + chunks] for x in range(0, len(model_name_store), chunks)]
    testset_features_c_norm = fl.apply_scaling(testset_fl.features_c)
    model_idx = 1
    for single_model_chunk in model_chunks:
        model_store = load_model_chunks(single_model_chunk)
        for idx, model in enumerate(model_store):
            wb.create_sheet('{}'.format(model_idx))
            model_idx += 1
            ws = wb[wb.sheetnames[-1]]
            if model:
                # Must use the round's fl to scale, not the testset scaler as it might be different
                p_y = model.predict(testset_features_c_norm)
                print(np.std(p_y, axis=0))

                for row, p_label in enumerate(p_y.tolist()):
                    if p_label[1] > p_label[2]:
                        p_y[row, 1] = p_y[row, 2]
                    if p_label[0] > p_y[row, 1]:
                        p_y[row, 0] = p_y[row, 1]

                se_store = (testset_fl.labels - p_y) ** 2
                re_store = np.abs(testset_fl.labels - p_y) / testset_fl.labels
                are_store = np.arctan(re_store)

                df = pd.DataFrame(data=np.concatenate((testset_fl.labels, p_y, se_store, re_store, are_store), axis=1),
                                  index=list(range(1, 1 + testset_fl.count)),
                                  columns=list(column_headers)
                                          + ['P_{}'.format(col) for col in column_headers]
                                          + ['SE_{}'.format(col) for col in column_headers]
                                          + ['RE_{}'.format(col) for col in column_headers]
                                          + ['ARE_{}'.format(col) for col in column_headers])
                print_df_to_excel(df=df, ws=ws)

                col = fn + 1 + 1 + 2 * numel + 3
                mse_store.append(np.mean(se_store))
                mre_store.append(np.mean(re_store))
                mare_store.append(np.mean(are_store))
                ws.cell(1, col).value = 'MSE'
                ws.cell(1, col + 1).value = mse_store[-1]
                ws.cell(2, col).value = 'MRE'
                ws.cell(2, col + 1).value = mre_store[-1]
                ws.cell(3, col).value = 'ARE'
                ws.cell(3, col + 1).value = mare_store[-1]
            else:
                ws.cell(1, 1).value = 'EOF error'
                ws.cell(2, 1).value = model_name_store[idx - 1]
                mse_store.append(np.nan)
                mre_store.append(np.nan)
                mare_store.append(np.nan)

    ws = wb[wb.sheetnames[0]]
    df = pd.DataFrame(data=np.array([mse_store, mre_store, mare_store]).T, columns=['mse', 're', 'are'],
                      index=range(1, 1 + len(mse_store)))
    df.insert(0, 'Name', model_name_store)
    print_df_to_excel(df=df, ws=ws)
    wb.save(write_excel)


def eval_combination_on_testset(av_excel, y_dat, combination_dat):
    with open(y_dat, "rb") as f:
        y = pickle.load(f)
    with open(combination_dat, "rb") as f:
        p_y_store = pickle.load(f)
        p_y_store = np.array([x[1] for x in p_y_store])
    if av_excel:
        av = pd.read_excel(av_excel, sheet_name='av', index_col=None)
        selected_mask = [idx for idx, value in enumerate(av.iloc[:, -1].values) if value == 1]
    else:
        selected_mask = [1] * len(p_y_store)

    p_y_selected_mean = np.mean(p_y_store[selected_mask, :, :], axis=0)
    re = np.mean(np.abs(y - p_y_selected_mean) / y)

    data = np.concatenate((y, p_y_selected_mean), axis=1)
    df = pd.DataFrame(data=data, columns=['cut=10', 'cut=100', 'End', 'P_cut=10', 'P_cut=100', 'P_End'])

    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[-1]]
    print_df_to_excel(df=df, ws=ws)

    wb.create_sheet('Models')
    ws = wb[wb.sheetnames[-1]]
    ws.cell(1, 1).value = 'Names'
    try:
        print_array_to_excel(array=av.iloc[:, 0].values[selected_mask], first_cell=(2, 1), ws=ws, axis=0)
    except:
        pass
    ws.cell(1, 2).value = 'RE'
    ws.cell(1, 3).value = re
    excel_dir = create_excel_file('./results/eval_combi.xlsx')
    wb.save(excel_dir)


def save_testset_prediction(combination_excel):
    xls = pd.ExcelFile(combination_excel)
    sheetnames = xls.sheet_names[1:]
    model_names = pd.read_excel(xls, 'Sheet', index_col=0).iloc[:, 0].values
    df_store = [pd.read_excel(xls, sheet, index_col=0) for sheet in sheetnames]
    p_y_store = [[model_names[idx], x.iloc[:, 3:6].values] for idx, x in enumerate(df_store) if
                 (x.shape[1] == 17 and x.iloc[0, 16] < 0.45)]
    with open('./results/valtestset_prediction.dat', "wb") as f:
        pickle.dump(p_y_store, f)

    df = pd.read_excel(xls, '1', index_col=0)
    y = df.iloc[:, :3].values
    with open('./results/valtestset_y.dat', "wb") as f:
        pickle.dump(y, f)


def testset_optimal_combination(results_dir, y_dat, combination_dat, hparams):
    with open(y_dat, "rb") as f:
        y = pickle.load(f)
    with open(combination_dat, "rb") as f:
        p_y_store = pickle.load(f)
    p_y_names = [x[0] for x in p_y_store]
    p_y_store = np.array([x[1] for x in p_y_store])
    total_models = len(p_y_store)
    creator.create("FitnessMax", base.Fitness, weights=(-1,))
    creator.create("Individual", list, fitness=creator.FitnessMax)

    def eval(individual):
        selected_mask = [idx for idx, value in enumerate(individual) if value == 1]
        p_y_selected_mean = np.mean(p_y_store[selected_mask, :, :], axis=0)
        re = np.mean(np.abs(y - p_y_selected_mean) / y)
        return (re,)

    toolbox = base.Toolbox()
    toolbox.register("attr_bool", np.random.choice, np.arange(0, 2), p=hparams['init'])
    toolbox.register("individual", tools.initRepeat, creator.Individual, toolbox.attr_bool, n=total_models)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    toolbox.register("evaluate", eval)
    toolbox.register("mate", tools.cxTwoPoint)
    toolbox.register("mutate", tools.mutFlipBit, indpb=0.05)
    toolbox.register("select", tools.selTournament, tournsize=3)
    # Logging
    stats = tools.Statistics(key=lambda ind: ind.fitness.values)
    stats.register("avg", np.mean, axis=0)
    stats.register("std", np.std, axis=0)
    stats.register("min", np.min, axis=0)
    stats.register("max", np.max, axis=0)
    pop = toolbox.population(n=hparams['n_pop'])
    hof = tools.HallOfFame(1)
    pop, logbook = algorithms.eaSimple(toolbox=toolbox, population=pop,
                                       cxpb=0.5, mutpb=0.2,
                                       ngen=hparams['n_gen'], halloffame=hof, stats=stats,
                                       verbose=True)

    # Plotting
    gen = logbook.select("gen")
    fit_min = [x.item() for x in logbook.select("min")]
    fit_avg = [x.item() for x in logbook.select("avg")]
    fit_max = [x.item() for x in logbook.select("max")]

    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, fit_min, label="Min MRE")
    line2 = ax1.plot(gen, fit_avg, label="Avg MRE")
    line3 = ax1.plot(gen, fit_max, label="Max MRE")
    plt.legend()
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Relative Error")
    plt.savefig('{}/plots/GA_opt_MRE_all.png'.format(results_dir), bbox_inches="tight")

    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, fit_min, label="Min MRE")
    plt.legend()
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Total Generation Cost")
    plt.savefig('{}/plots/GA_opt_min_only.png'.format(results_dir), bbox_inches="tight")

    # Printing to excel
    excel_name = results_dir + '/results.xlsx'
    wb = openpyxl.Workbook()
    sheetname = wb.sheetnames[-1]
    ws = wb[sheetname]

    # Writing other subset split, instance per run, and bounds
    print_array_to_excel(['n_gen', 'n_pop'], (1, 1), ws, axis=1)
    print_array_to_excel([hparams['n_gen'], hparams['n_pop']], (2, 1), ws, axis=1)
    row = 2
    ws.cell(row + 1, 1).value = 'Best Allocation Value'
    ws.cell(row + 1, 2).value = hof[-1].fitness.values[-1]

    wb.create_sheet('av')
    ws = wb['av']
    ws.cell(1, 1).value = 'Names'
    ws.cell(1, 2).value = 'av'
    print_array_to_excel(p_y_names, (2, 1), ws=ws, axis=0)
    print_array_to_excel(list(hof[-1]), (2, 2), ws=ws, axis=0)

    wb.save(excel_name)


def save_valset_prediction(skf_excel_store):
    p_y_store = []
    for skf_excel in skf_excel_store:
        xls = pd.ExcelFile(skf_excel)
        sheetnames = xls.sheet_names[1:]
        df_store = [pd.read_excel(xls, sheet, index_col=0).sort_index() for sheet in sheetnames]
        p_y_store.extend(
            [['{}_{}'.format(skf_excel, x[1]), x[0].iloc[:, 10:13].values] for x in zip(df_store, sheetnames)])

    with open('./results/valset_prediction.dat', "wb") as f:
        pickle.dump(p_y_store, f)

    df = pd.read_excel(xls, sheetnames[0], index_col=0).sort_index()
    y = df.iloc[:, 7:10].values
    with open('./results/valset_y.dat', "wb") as f:
        pickle.dump(y, f)


def features_correlation_analysis(data_excel):
    write_dir = create_results_directory('./results/cross_correlation_analysis')
    try:
        del mpl.font_manager.weight_dict['roman']
        mpl.font_manager._rebuild()
    except KeyError:
        pass
    sns.set(style='ticks')
    mpl.rc('font', family='Times New Roman')

    df = pd.read_excel(data_excel, index_col=0, sheet_name='features')
    df_labels = pd.read_excel(data_excel, index_col=0, sheet_name='cutoff')
    working_range = df_labels.iloc[:, -1].values - df_labels.iloc[:, -2].values
    df.insert(loc=df.shape[-1] - 3, column='Working Range', value=working_range)
    df1 = df[df.iloc[:, -3] == 1].iloc[:, :-3]
    df2 = df[df.iloc[:, -2] == 1].iloc[:, :-3]
    df3 = df[df.iloc[:, -1] == 1].iloc[:, :-3]

    x_store = ['CNT Mass Percentage', 'PVA Mass Percentage', 'Thickness nm', 'Mxene Mass Percentage']
    mypal = sns.hls_palette(4, l=.3, s=.8)

    for dimension, df in enumerate([df1, df2, df3]):
        df['Mxene Mass Percentage'] = 1-df.iloc[:,0]-df.iloc[:,1]
        for x, color in zip(x_store, mypal):
            plt.close()
            sns.jointplot(x=x, y='Working Range', data=df, alpha=0.3, color=color, stat_func=stat.pearsonr)
            plt.savefig('{}/{}_dim_{}.png'.format(write_dir, x, dimension), bbox_inches='tight')


def training_curve_comparision(train_excel, val_excel, write_dir, hparams):
    fl = load_data_to_fl(data_loader_excel_file=train_excel, normalise_labels=False,
                         label_type='cutoff',
                         norm_mask=[0, 1, 3, 4, 5])
    val_fl = load_testset_to_fl(val_excel, scaler=fl.scaler, norm_mask=[0, 1, 3, 4, 5])

    data = defaultdict(list)

    def run_model(loss, pre, epoch):
        sess = tf.compat.v1.Session()
        K.set_session(sess)
        hparams['loss'] = loss
        hparams['pre'] = pre
        hparams['epochs'] = epoch
        model = Kmodel(fl=fl, mode='ann3', hparams=hparams)
        _, history = model.train_model(fl, val_fl, plot_name='{}/{}_{}_{}.png'.format(write_dir, loss, pre, epoch))

        data['train_error'].append(history.history['loss'])
        data['test_error'].append(history.history['val_loss'])
        data['names'].append('{}_{}_{}'.format(loss, pre, epoch))

        # Need to put the next 3 lines if not memory will run out
        del model
        K.clear_session()
        sess.close()
        gc.collect()

    #run_model(loss='mse', pre=100, epoch=500)
    #run_model(loss='haitao', pre=100, epoch=500)
    #run_model(loss='mse', pre=500, epoch=500)
    #run_model(loss='haitao', pre=500, epoch=500)
#
    #run_model(loss='mse', pre=100, epoch=1000)
    #run_model(loss='haitao', pre=100, epoch=1000)
    #run_model(loss='mse', pre=500, epoch=1000)
    #run_model(loss='haitao', pre=500, epoch=1000)
#
    #run_model(loss='mse', pre=100, epoch=5000)
    #run_model(loss='haitao', pre=100, epoch=5000)
    run_model(loss='mse', pre=500, epoch=5000)
    run_model(loss='haitao', pre=500, epoch=5000)
    #run_model(loss='mse', pre=100, epoch=500)
    #run_model(loss='mse', pre=100, epoch=500)
    #run_model(loss='mse', pre=100, epoch=500)
    #run_model(loss='mse', pre=100, epoch=500)
    #run_model(loss='mse', pre=100, epoch=500)
    #run_model(loss='mse', pre=100, epoch=500)

    wb = openpyxl.Workbook()
    wb.create_sheet('125train_loss')
    wb.create_sheet('30test_loss')

    train_df = pd.DataFrame(data=list(itertools.zip_longest(*data['train_error'], fillvalue="")), columns=data['names'])
    test_df = pd.DataFrame(data=list(itertools.zip_longest(*data['test_error'], fillvalue="")), columns=data['names'])

    ws = wb['125train_loss']
    print_df_to_excel(df=train_df, ws=ws)

    ws = wb['30test_loss']
    print_df_to_excel(df=test_df, ws=ws)

    wb.save('{}/training_curve_results.xlsx'.format(write_dir))
