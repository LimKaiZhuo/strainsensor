from sklearn.svm import SVR
from sklearn.tree import DecisionTreeRegressor
from sklearn.multioutput import MultiOutputRegressor
from sklearn.ensemble import AdaBoostRegressor
from sklearn.metrics import classification_report, confusion_matrix, matthews_corrcoef, mean_squared_error
import pickle, time, gc
import numpy as np
import pandas as pd
from pandas import Series
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import xgboost as xgb

from tensorflow.python.keras.models import Sequential, Model
from tensorflow.python.keras.layers import Dense, Dropout, merge, Input, concatenate, Reshape, Permute, LSTM, \
    TimeDistributed, RepeatVector, MaxPooling1D, BatchNormalization
import tensorflow.python.keras

from own_package.others import print_array_to_excel, create_results_directory
from own_package.models.models import ann
from own_package.models.mimosvr import msvr, kernelmatrix


class ANNmodel:
    def __init__(self, fl, hparams):
        """
        Initialises new DNN model based on input features_dim, labels_dim, hparams
        :param features_dim: Number of input feature nodes. Integer
        :param labels_dim: Number of output label nodes. Integer
        :param hparams: Dict containing hyperparameter information. Dict can be created using create_hparams() function.
        hparams includes: hidden_layers: List containing number of nodes in each hidden layer. [10, 20] means 10 then 20 nodes.
        """
        self.features_dim = fl.features_c_dim
        self.labels_dim = 1
        self.hparams = hparams
        self.normalise_labels = fl.normalise_labels
        self.labels_end_scaler = fl.labels_end_scaler
        features_in = Input(shape=(self.features_dim,), name='main_features_c_input')
        model = ann(self.features_dim, self.labels_dim, self.hparams)
        x = model(features_in)
        self.model = Model(inputs=features_in, outputs=x)
        optimizer = keras.optimizers.adam(clipnorm=1)
        self.model.compile(optimizer=optimizer, loss='mean_squared_error')

    def train_model(self, fl, i_fl,
                    save_name='mt.h5', save_dir='./save/models/',
                    save_mode=False, plot_name=None):
        # Training model
        training_features = fl.features_c_norm
        val_features = i_fl.features_c_norm
        if self.normalise_labels:
            training_labels = fl.labels_end_norm
            val_labels = i_fl.labels_end_norm
        else:
            training_labels = fl.labels_end
            val_labels = i_fl.labels_end
        # labels must come in the matrix with rows of examples and columns are end, p1,p2,p3,...

        if plot_name:
            history = self.model.fit(training_features, training_labels,
                                     validation_data=(val_features, val_labels),
                                     epochs=self.hparams['epochs'],
                                     batch_size=self.hparams['batch_size'],
                                     verbose=self.hparams['verbose'])
            # summarize history for accuracy
            plt.semilogy(history.history['loss'], label=['train'])
            plt.semilogy(history.history['val_loss'], label=['test'])
            plt.plot([], [], ' ', label='Final train: {:.3e}'.format(history.history['loss'][-1]))
            plt.plot([], [], ' ', label='Final val: {:.3e}'.format(history.history['val_loss'][-1]))
            plt.title('model loss')
            plt.ylabel('loss')
            plt.xlabel('epoch')
            plt.legend(loc='upper right')
            plt.savefig(plot_name, bbox_inches='tight')
            plt.close()
        else:
            self.model.fit(training_features, training_labels,
                           epochs=self.hparams['epochs'],
                           batch_size=self.hparams['batch_size'],
                           verbose=self.hparams['verbose'])
        # Debugging check to see features and prediction
        # pprint.pprint(training_features)
        # pprint.pprint(self.model.predict(training_features))
        # pprint.pprint(training_labels)
        # Saving Model
        if save_mode:
            self.model.save(save_dir + save_name)
        return self.model

    def eval(self, eval_fl):
        features = eval_fl.features_c_norm
        if self.normalise_labels:
            predictions = eval_fl.labels_end_scaler.inverse_transform(self.model.predict(features))
        else:
            predictions = self.model.predict(features)
        return predictions


class SVRmodel:
    def __init__(self, fl, epsilon=1, c=1):
        """
        Initialises new DNN model based on input features_dim, labels_dim, hparams
        :param features_dim: Number of input feature nodes. Integer
        :param labels_dim: Number of output label nodes. Integer
        :param hparams: Dict containing hyperparameter information. Dict can be created using create_hparams() function.
        hparams includes: hidden_layers: List containing number of nodes in each hidden layer. [10, 20] means 10 then 20 nodes.
        """
        self.labels_dim = fl.labels_dim  # Assuming that each task has only 1 dimensional output
        self.labels_scaler = fl.labels_scaler
        if self.labels_dim == 1:
            self.model = SVR(kernel='rbf', gamma=gamma, C=c)
        else:
            self.model = MultiOutputRegressor(SVR(kernel='rbf', epsilon=epsilon, C=c))

    def train_model(self, fl, save_mode=False, plot_name=None):
        training_features = fl.features_c_norm
        training_labels = fl.labels_norm

        self.model.fit(training_features, training_labels)

        return self.model

    def eval(self, eval_fl):
        features = eval_fl.features_c_norm
        if self.labels_dim == 1:
            y_pred = self.model.predict(features)[:, None]
        else:
            y_pred = self.model.predict(features)
        mse_norm = mean_squared_error(eval_fl.labels_norm, y_pred)
        mse = mean_squared_error(eval_fl.labels, self.labels_scaler.inverse_transform(y_pred))

        return y_pred, mse, mse_norm


class DTRmodel:
    def __init__(self, fl, max_depth=8, num_est=300):
        """
        Initialises new DNN model based on input features_dim, labels_dim, hparams
        :param features_dim: Number of input feature nodes. Integer
        :param labels_dim: Number of output label nodes. Integer
        :param hparams: Dict containing hyperparameter information. Dict can be created using create_hparams() function.
        hparams includes: hidden_layers: List containing number of nodes in each hidden layer. [10, 20] means 10 then 20 nodes.
        """
        self.labels_dim = fl.labels_dim  # Assuming that each task has only 1 dimensional output
        self.labels_scaler = fl.labels_scaler
        self.model = MultiOutputRegressor(
            AdaBoostRegressor(DecisionTreeRegressor(max_depth=max_depth), n_estimators=num_est))
        self.normalise_labels = fl.normalise_labels

    def train_model(self, fl, save_mode=False, plot_name=None):
        training_features = fl.features_c_norm
        if self.normalise_labels:
            training_labels = fl.labels_norm
        else:
            training_labels = fl.labels

        self.model.fit(training_features, training_labels)

        return self.model

    def eval(self, eval_fl):
        features = eval_fl.features_c_norm
        if self.labels_dim == 1:
            y_pred = self.model.predict(features)[:, None]
        else:
            y_pred = self.model.predict(features)
        if self.normalise_labels:
            mse_norm = mean_squared_error(eval_fl.labels_norm, y_pred)
            mse = mean_squared_error(eval_fl.labels, self.labels_scaler.inverse_transform(y_pred))
        else:
            mse_norm = -1
            mse = mean_squared_error(eval_fl.labels, y_pred)

        return y_pred, mse, mse_norm


class XGBmodel:
    def __init__(self, fl, hparams):
        """
        Initialises new DNN model based on input features_dim, labels_dim, hparams
        :param features_dim: Number of input feature nodes. Integer
        :param labels_dim: Number of output label nodes. Integer
        :param hparams: Dict containing hyperparameter information. Dict can be created using create_hparams() function.
        hparams includes: hidden_layers: List containing number of nodes in each hidden layer. [10, 20] means 10 then 20 nodes.
        """
        self.labels_dim = fl.labels_dim  # Assuming that each task has only 1 dimensional output
        self.labels_scaler = fl.labels_scaler
        default_hparams = {'seed': 42,
                           'booster': 'gbtree',
                           'learning_rate': 0.1,
                           'objective': 'reg:squarederror',
                           'verbosity': 0,
                           'subsample': 1,
                           'num_boost_round': 600,
                           'early_stopping_rounds': 100,
                           # params that will vary
                           'm': 6,
                           'p': 12,
                           'max_depth': 1,
                           'colsample_bytree': 0.5,
                           }
        self.hparams = {**default_hparams, **hparams}
        self.model = MultiOutputRegressor(
            xgb.XGBRegressor(objective=self.hparams['objective'],
                             n_estimators=self.hparams['num_boost_round'],
                             max_depth=self.hparams['max_depth'],
                             booster=self.hparams['booster'],
                             gamma=self.hparams['gamma'],
                             subsample=self.hparams['subsample'],
                             random_state=self.hparams['seed'],))
        self.normalise_labels = fl.normalise_labels

    def train_model(self, fl, i_fl, **kwargs):
        training_features = fl.features_c_norm
        val_features = i_fl.features_c_norm
        if self.normalise_labels:
            training_labels = fl.labels_norm
            val_labels = i_fl.labels_norm
        else:
            training_labels = fl.labels
            val_labels = i_fl.labels

        self.model.fit(X=training_features, y=training_labels, sample_weight=None, **{'early_stopping_rounds': 100,
                                                                                      'eval_set':(val_features, val_labels)})

        return self.model

    def eval(self, eval_fl):
        features = eval_fl.features_c_norm
        if self.labels_dim == 1:
            y_pred = self.model.predict(features)[:, None]
        else:
            y_pred = self.model.predict(features)
        if self.normalise_labels:
            mse_norm = mean_squared_error(eval_fl.labels_norm, y_pred)
            mse = mean_squared_error(eval_fl.labels, self.labels_scaler.inverse_transform(y_pred))
        else:
            mse_norm = -1
            mse = mean_squared_error(eval_fl.labels, y_pred)

        return y_pred, mse, mse_norm


'''
class XGBmodel:
    def __init__(self, fl, hparams):
        """
        """
        self.labels_dim = fl.labels_dim  # Assuming that each task has only 1 dimensional output
        self.labels_scaler = fl.labels_scaler
        self.normalise_labels = fl.normalise_labels
        default_hparams = {'seed': 42,
                           'booster': 'gbtree',
                           'learning_rate': 0.1,
                           'objective': 'reg:squarederror',
                           'verbosity': 0,
                           'subsample': 1,
                           'num_boost_round': 600,
                           'early_stopping_rounds': 100,
                           # params that will vary
                           'm': 6,
                           'p': 12,
                           'max_depth': 1,
                           'colsample_bytree': 0.5,
                           }
        self.hparams = {**default_hparams,**hparams}

    def train_model(self, fl, i_fl, **kwargs):
        training_features = fl.features_c_norm
        val_features = i_fl.features_c_norm
        if self.normalise_labels:
            training_labels = fl.labels_norm
            val_labels = i_fl.labels_norm
        else:
            training_labels = fl.labels
            val_labels = i_fl.labels

        dtrain = xgb.DMatrix(training_features, label=training_labels)
        deval = xgb.DMatrix(data=val_features, label=val_labels)
        self.model = xgb.train(self.hparams, dtrain=dtrain, num_boost_round=self.hparams['num_boost_round'],
                               early_stopping_rounds=self.hparams['early_stopping_rounds'],
                               evals=[(dtrain, 'train'), (deval, 'val')],
                               verbose_eval=False)

        return self.model

    def eval(self, eval_fl):
        features = xgb.DMatrix(eval_fl.features_c_norm)
        if self.labels_dim == 1:
            y_pred = self.model.predict(features)[:, None]
        else:
            y_pred = self.model.predict(features)
        if self.normalise_labels:
            mse_norm = mean_squared_error(eval_fl.labels_norm, y_pred)
            mse = mean_squared_error(eval_fl.labels, self.labels_scaler.inverse_transform(y_pred))
        else:
            mse_norm = -1
            mse = mean_squared_error(eval_fl.labels, y_pred)

        return y_pred, mse, mse_norm
'''


class Predict_SVR_DTR:
    def __init__(self, model, labels_scaler, xgb=False):
        self.model = model
        self.labels_scaler = labels_scaler
        self.xgb = xgb

    def predict(self, features):
        '''
        if self.xgb:
            features = xgb.DMatrix(features)
            try:
                y_pred = self.labels_scaler.inverse_transform(
                    self.model.predict(features, ntree_limit=self.model.best_ntree_limit))
            except:
                y_pred = self.model.predict(features, ntree_limit=self.model.best_ntree_limit)
            return y_pred
        else:
            try:
                y_pred = self.labels_scaler.inverse_transform(self.model.predict(features))
            except:
                y_pred = self.model.predict(features)
            return y_pred
        '''
        try:
            y_pred = self.labels_scaler.inverse_transform(self.model.predict(features))
        except:
            y_pred = self.model.predict(features)
        return y_pred


class MIMOSVRmodel:
    def __init__(self, fl, gamma=1):
        """
        Initialises new DNN model based on input features_dim, labels_dim, hparams
        :param features_dim: Number of input feature nodes. Integer
        :param labels_dim: Number of output label nodes. Integer
        :param hparams: Dict containing hyperparameter information. Dict can be created using create_hparams() function.
        hparams includes: hidden_layers: List containing number of nodes in each hidden layer. [10, 20] means 10 then 20 nodes.
        """
        self.labels_dim = fl.labels_dim  # Assuming that each task has only 1 dimensional output
        self.fl = fl
        self.labels_scaler = fl.labels_scaler
        if self.labels_dim == 1:
            raise TypeError('MultiSVRmodel only accepts multiple labels. Label dimension is 1.')

    def train_model(self, fl, save_mode=False, plot_name=None):
        training_features = fl.features_c_norm
        training_labels = fl.labels_norm

        self.C = 1  # Parametro de regularización (boxConstraint)
        self.paramK = 1  # Parametro para la funcion kernel
        self.tipoK = 'rbf'  # Tipo de kernel a usar en el modelo. Los disponibles son: 'lin', 'poly' y 'rbf'
        self.epsilon = 20  # Parametro que establece el ancho del epsilon tubo o zona de tolerancia (por defecto es 1)
        self.tol = 10 ** -20  # Parametro que necesita el modelo para detener el modelo si el error se vuelve muy bajo

        self.model = msvr(training_features, training_labels, self.tipoK, self.C,
                          self.epsilon,
                          self.paramK, self.tol)

        return self.model

    def eval(self, eval_fl):
        features = eval_fl.features_c_norm

        Beta, numero_VectoresS, kernel_train, indices_vectoresS = self.model

        kernel_test = kernelmatrix(self.tipoK, features.T, self.fl.features_c_norm.T, self.paramK)
        y_pred = np.dot(kernel_test, Beta)

        mse_norm = mean_squared_error(eval_fl.labels_norm, y_pred)
        mse = mean_squared_error(eval_fl.labels, self.labels_scaler.inverse_transform(y_pred))

        return y_pred, mse, mse_norm


def run_svr(fl_store, write_dir, excel_dir, model_selector, gamma=1, hparams=None, save_name=None):
    # Run k model instance to perform skf
    predicted_labels_store = []
    folds = []
    val_idx = []
    val_features = []
    val_labels = []
    for fold, fl_tuple in enumerate(fl_store):
        instance_start = time.time()

        (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl
        if model_selector == 'svr':
            model = SVRmodel(fl=ss_fl, gamma=gamma)
            model.train_model(fl=ss_fl)
        elif model_selector == 'ann':
            model = ANNmodel(fl=ss_fl, hparams=hparams)
            #  plot_name='{}/plots/{}.png'.format(write_dir,fold)
            model.train_model(fl=ss_fl, i_fl=i_ss_fl)
        else:
            raise KeyError('model selector argument is not one of the available models.')

        # Evaluation
        predicted_labels = model.eval(i_ss_fl)
        predicted_labels_store.extend(predicted_labels.flatten().tolist())

        # Saving model
        save_model_name = '{}/models/{}_{}_{}'.format(write_dir, save_name, model_selector, str(fold + 1))
        print('Saving instance {} model in {}'.format(fold + 1, save_model_name))
        if model_selector == 'svr':
            with open(save_model_name, 'wb') as handle:
                pickle.dump(model, handle, protocol=pickle.HIGHEST_PROTOCOL)
        elif model_selector == 'ann':
            model.model.save(save_model_name + '.h5')

        del model
        gc.collect()

        # Preparing data to put into new_df that consists of all the validation dataset and its predicted labels
        folds.extend([fold] * i_ss_fl.count)  # Make a col that contains the fold number for each example
        if len(val_features):
            val_features = np.concatenate((val_features, i_ss_fl.features_c),
                                          axis=0)
        else:
            val_features = i_ss_fl.features_c

        val_labels.extend(i_ss_fl.labels_end.flatten().tolist())
        val_idx.extend(i_ss_fl.idx)

        # Printing one instance summary.
        instance_end = time.time()
        print(
            '\nFor k-fold run {} out of {}. Each fold has {} examples. Time taken for '
            'instance = {}\n'
            '####################################################################################################'
                .format(fold + 1, 10, i_ss_fl.count, instance_end - instance_start))

    # Calculating metrics based on complete validation prediction
    mse = mean_squared_error(y_true=val_labels, y_pred=predicted_labels_store)

    # Creating dataframe to print into excel later.
    new_df = np.concatenate((np.array(folds)[:, None],  # Convert 1d list to col. vector
                             val_features,
                             np.array(val_labels)[:, None],
                             np.array(predicted_labels_store)[:, None])
                            , axis=1)
    headers = ['folds'] + \
              list(map(str, fl_store[0][0].features_c_names)) + \
              ['End', 'P_End']

    # val_idx is the original position of the example in the data_loader
    new_df = pd.DataFrame(data=new_df, columns=headers, index=val_idx)

    skf_file = excel_dir
    print('Writing into' + skf_file)
    wb = load_workbook(skf_file)
    wb.create_sheet(model_selector)
    sheet_name = wb.sheetnames[-1]

    # Writing results dataframe
    pd_writer = pd.ExcelWriter(skf_file, engine='openpyxl')
    pd_writer.book = wb
    pd_writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    new_df.to_excel(pd_writer, sheet_name=sheet_name)
    start_col = len(new_df.columns) + 4

    # Writing other subset split, instance per run, and bounds
    ws = wb.sheetnames
    ws = wb[ws[-1]]
    headers = ['mse']
    values = [mse]
    print_array_to_excel(np.array(headers), (1, start_col + 1), ws, axis=1)
    print_array_to_excel(np.array(values), (2, start_col + 1), ws, axis=1)
    pd_writer.save()
    pd_writer.close()
    wb.close()

    return mse
