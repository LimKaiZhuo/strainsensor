import numpy as np
import numpy.random as rng
import pandas as pd
from openpyxl import load_workbook
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import StratifiedKFold, KFold, LeaveOneOut
import pickle
import os
import pathlib
import warnings
import copy
import xlrd
from own_package.smote.smote_code import produce_smote, produce_invariant
from .others import print_array_to_excel


class Shortcut_fl:
    def __init__(self, features_c, labels, scaler, feature_names, label_names, norm_mask):
        self.features_c_names = feature_names
        self.features_c = features_c
        self.features_c_dim = features_c.shape[1]
        self.count = features_c.shape[0]
        self.norm_mask = norm_mask
        mask = np.array([1] * self.features_c_dim, dtype=np.bool)
        mask[norm_mask] = 0
        if features_c[:, mask].shape[1] == 0:
            self.scaler = 0
            self.features_c_norm = features_c
        else:
            if scaler is None:
                # If scaler is None, means normalize the data with all input data
                self.scaler = MinMaxScaler()
                self.scaler.fit(features_c[:, mask])  # Setting up scaler
            else:
                # If scaler is given, means normalize the data with the given scaler
                self.scaler = scaler
            features_c_norm = self.scaler.transform(features_c[:, mask])  # Normalizing features_c
            self.features_c_norm = np.copy(features_c)
            self.features_c_norm[:, mask] = features_c_norm
        self.labels = labels
        self.labels_names = label_names

def load_testset_to_fl(testset_excel_file, norm_mask, scaler):
    df = pd.read_excel(testset_excel_file, index_col=0)
    features = df.iloc[:,:-3].values
    labels = df.iloc[:,-3:].values
    return Shortcut_fl(features_c=features, labels=labels, scaler=scaler,
                       feature_names=df.columns[:-3], label_names=df.columns[-3:], norm_mask=norm_mask)


def load_data_to_fl(data_loader_excel_file, normalise_labels, label_type, norm_mask=None):
    xls = pd.ExcelFile(data_loader_excel_file)
    df_features = pd.read_excel(xls, sheet_name='features', index_col=0)
    df_features_d = pd.read_excel(xls, sheet_name='features_d', index_col=0)
    df_labels = pd.read_excel(xls, sheet_name=label_type, index_col=0)
    try:
        df_classification = pd.read_excel(xls, sheet_name='class', index_col=0)
        labels_classification = df_classification.values.flatten()
    except:
        df_classification = df_labels
        labels_classification = df_labels.values

    features_c = df_features.values
    features_c_names = df_features.columns.values
    if not df_features_d.empty:
        features_d_names = df_features_d.columns.values.tolist()

        # Lookup for features_d descriptors
        lookup_df_store = []
        features_dc_names = []
        for item in features_d_names:
            try:
                temp_df = pd.read_excel(data_loader_excel_file, sheet_name=item, index_col=0)
                temp_df = temp_df.set_index(item)
                lookup_df_store.append(temp_df)
                features_dc_names.extend(temp_df.columns.values.tolist())
            except xlrd.biffh.XLRDError:
                print('{} features_d descriptors not found in excel. Skipping this feature_d'.format(item))

        # Filling up features descriptors for each discrete feature
        features_dc_store = []
        for idx, rows in df_features_d.iterrows():
            row_list = []
            for header, lookup_df in zip(features_d_names, lookup_df_store):
                row_list.extend(lookup_df.loc[rows[header], :].values.tolist())
            features_dc_store.append(row_list)

        # Combine features_c with features_dc
        features_dc_store = np.array(features_dc_store)
        features_c = np.concatenate((features_c, features_dc_store), axis=1)
        features_c_names = np.concatenate((features_c_names, np.array(features_dc_names)), axis=0)
    else:
        lookup_df_store = None

    if label_type == 'points':
        labels = df_labels.values
        labels_end = labels[:,0][:,None]  # Make 2D array
        labels = labels[:,2:]
        labels_names = df_labels.columns.values
    elif label_type == 'cutoff':
        labels = df_labels.values
        labels_names = df_labels.columns.values
        labels_end = labels[:,-1][:,None]
        remove_idx = np.where(labels[:,0]==-1)[0]
        labels = np.delete(labels, remove_idx, axis=0)
        labels_end = np.delete(labels_end, remove_idx, axis=0)
        features_c = np.delete(features_c, remove_idx, axis=0)
    elif label_type == 'gf20':
        labels = df_labels.values
        labels_end = labels[:,0][:,None]  # Make 2D array
        labels = labels[:,1:]
        labels_names = df_labels.columns.values
    else:
        raise KeyError('label_type {} not recognised'.format(label_type))


    fl = Features_labels(features_c, labels_end, labels, label_type=label_type, features_c_names=features_c_names,
                         labels_names=labels_names, labels_classification=labels_classification,
                         norm_mask=norm_mask,
                         normalise_labels=normalise_labels,
                         features_d_df=df_features_d, lookup_df=lookup_df_store)

    return fl


class Features_labels:
    def __init__(self, features_c, labels_end, labels, label_type, features_c_names=None, labels_names=None, scaler=None,
                 norm_mask=None, normalise_labels=False, labels_scaler=None, labels_end_scaler=None, labels_classification=None,
                 idx=None, features_d_df=None,
                 lookup_df=None):
        """
        Creates fl class with a lot useful attributes
        :param features_c: Continuous features. Np array, no. of examples x continous features
        :param labels: Labels as np array, no. of examples x dim
        :param scaler: Scaler to transform features c. If given, use given MinMax scaler from sklearn,
        else create scaler based on given features c.
        """

        self.features_c_names = features_c_names
        self.label_type = label_type

        if isinstance(features_d_df, pd.DataFrame):
            if not features_d_df.empty:
                self.features_d_df = features_d_df
                self.lookup_df_store = lookup_df
                self.features_d_names = features_d_df.columns.values.tolist()
                self.features_d_space = []
                for lookup_df in self.lookup_df_store:
                    self.features_d_space.append(lookup_df.index.values.tolist())

        # Setting up features
        self.count = features_c.shape[0]
        if isinstance(idx, np.ndarray):
            self.idx = idx
        else:
            self.idx = np.arange(self.count)
        self.features_c = np.copy(features_c)
        self.features_c_dim = features_c.shape[1]

        if norm_mask:
            self.norm_mask = norm_mask
            mask = np.array([1] * self.features_c_dim, dtype=np.bool)
            mask[norm_mask] = 0
            if features_c[:, mask].shape[1] == 0:
                self.scaler = 0
                self.features_c_norm = features_c
            else:
                if scaler is None:
                    # If scaler is None, means normalize the data with all input data
                    self.scaler = MinMaxScaler()
                    self.scaler.fit(features_c[:, mask])  # Setting up scaler
                else:
                    # If scaler is given, means normalize the data with the given scaler
                    self.scaler = scaler
                features_c_norm = self.scaler.transform(features_c[:, mask])  # Normalizing features_c
                self.features_c_norm = features_c
                self.features_c_norm[:, mask] = features_c_norm
        else:
            # Normalizing continuous features
            self.norm_mask = None
            if scaler is None:
                # If scaler is None, means normalize the data with all input data
                self.scaler = MinMaxScaler()
                self.scaler.fit(features_c)  # Setting up scaler
            else:
                # If scaler is given, means normalize the data with the given scaler
                self.scaler = scaler
            self.features_c_norm = self.scaler.transform(features_c)  # Normalizing features_c

        # Setting up labels
        self.labels_end = labels_end
        self.labels = labels
        if len(labels.shape) == 2:
            self.labels_dim = labels.shape[1]
        else:
            self.labels_dim = 1

        # Label name is size 2 larger than self.labels as it includes the col name for end point and the first pt.
        self.labels_names = labels_names
        self.labels_classification = labels_classification

        # Normalizing labels for the 19 labels going into the neural network
        if normalise_labels:
            self.normalise_labels = normalise_labels
            if labels_scaler is None:
                self.labels_scaler = MinMaxScaler(feature_range=(0, 1))
                self.labels_scaler.fit(labels)
                self.labels_end_scaler = MinMaxScaler(feature_range=(0, 1))
                self.labels_end_scaler.fit(labels_end)
            else:
                self.labels_scaler = labels_scaler
                self.labels_end_scaler = labels_end_scaler
            self.labels_norm = self.labels_scaler.transform(labels)
            self.labels_end_norm = self.labels_end_scaler.transform(labels_end)
        else:
            self.normalise_labels = False
            self.labels_scaler = None
            self.labels_norm = None
            self.labels_end_scaler = None
            self.labels_end_norm = None

    def apply_scaling(self, features_c):
        if features_c.ndim == 1:
            features_c = features_c.reshape((1, -1))
        if self.norm_mask:
            norm_mask = self.norm_mask
            mask = np.array([1] * self.features_c_dim, dtype=np.bool)
            mask[norm_mask] = 0
            features_c_norm = np.copy(features_c)
            features_c_norm[:, mask] = self.scaler.transform(features_c[:, mask])
        else:
            features_c_norm = self.scaler.transform(features_c)
        return features_c_norm

    def generate_random_examples(self, numel):
        gen_features_c_norm_a = rng.random_sample((numel, self.features_c_dim))
        gen_features_c_a = self.scaler.inverse_transform(gen_features_c_norm_a)

        # Creating dic for SNN prediction
        gen_dic = {}
        gen_dic = dict(
            zip(('gen_features_c_a', 'gen_features_c_norm_a'), (gen_features_c_a, gen_features_c_norm_a)))
        return gen_dic

    def create_kf(self, k_folds, shuffle=True):
        '''
        Almost the same as skf except can work for regression labels and folds are not stratified.
        Create list of tuples containing (fl_train,fl_val) fl objects for k fold cross validation
        :param k_folds: Number of folds
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        skf = KFold(n_splits=k_folds, shuffle=shuffle)
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(skf.split(self.features_c, self.labels)):
            # Generate batches from indices
            xval_idx = self.idx[val_indices]
            xtrain, xval = self.features_c[train_indices], self.features_c[val_indices]
            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            yendtrain, yendval = self.labels_end[train_indices], self.labels_end[val_indices]
            yclasstrain, yclassval = self.labels_classification[train_indices], self.labels_classification[val_indices]
            fl_store.append(
                (Features_labels(xtrain, yendtrain, ytrain, scaler=self.scaler, normalise_labels=self.normalise_labels,
                                 labels_scaler=self.labels_scaler, labels_end_scaler=self.labels_end_scaler,
                                 labels_classification=yclasstrain, labels_names=self.labels_names,
                                 norm_mask=self.norm_mask, features_c_names=self.features_c_names, label_type=self.label_type),
                 Features_labels(xval, yendval, yval, idx=xval_idx, scaler=self.scaler, normalise_labels=self.normalise_labels,
                                 labels_scaler=self.labels_scaler, labels_end_scaler=self.labels_end_scaler,
                                 labels_classification=yclassval, labels_names=self.labels_names,
                                 norm_mask=self.norm_mask, features_c_names=self.features_c_names, label_type=self.label_type))
            )
        return fl_store

    def smote_kf_augment(self, smote_excel, k_folds, shuffle=True):
        """
        Same as kf above. But appends all smote data to each fold's training examples. Validation examples no change.
        :param smote_excel:
        :param k_folds:
        :param shuffle:
        :return:
        """
        smote = pd.read_excel(smote_excel, index_col=0).values
        smote_features = smote[:, :6]
        smote_labels = smote[:,6:]
        smote_end = smote_labels[:,-1][:, None]  # Make 2D array
        fl_store = []
        # Instantiate the cross validator
        skf = KFold(n_splits=k_folds, shuffle=shuffle)
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(skf.split(self.features_c, self.labels)):
            # Generate batches from indices
            xval_idx = self.idx[val_indices]
            xtrain, xval = np.concatenate((self.features_c[train_indices], smote_features), axis=0), self.features_c[val_indices]
            ytrain, yval = np.concatenate((self.labels[train_indices], smote_labels), axis=0), self.labels[val_indices]
            yendtrain, yendval = np.concatenate((self.labels_end[train_indices], smote_end), axis=0), self.labels_end[val_indices]
            fl_store.append(
                (Features_labels(xtrain, yendtrain, ytrain, scaler=self.scaler, normalise_labels=self.normalise_labels,
                                 labels_scaler=self.labels_scaler, labels_end_scaler=self.labels_end_scaler,
                                 norm_mask=self.norm_mask, features_c_names=self.features_c_names,
                                 label_type=self.label_type, labels_names=self.labels_names),
                 Features_labels(xval, yendval, yval, idx=xval_idx, scaler=self.scaler,
                                 normalise_labels=self.normalise_labels,
                                 labels_scaler=self.labels_scaler, labels_end_scaler=self.labels_end_scaler,
                                 norm_mask=self.norm_mask, features_c_names=self.features_c_names,
                                 label_type=self.label_type, labels_names=self.labels_names))
            )
        return fl_store

    def fold_smote_kf_augment(self, numel, k_folds, shuffle=True):
        """
        Same as kf above. But appends all smote data to each fold's training examples. Validation examples no change.
        :param smote_excel:
        :param k_folds:
        :param shuffle:
        :return:
        """
        fl_store = []
        # Instantiate the cross validator
        skf = KFold(n_splits=k_folds, shuffle=shuffle)
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(skf.split(self.features_c, self.labels)):
            # Generate batches from indices
            xval_idx = self.idx[val_indices]
            smote_features, smote_labels = produce_smote(self.features_c[train_indices],
                                                         self.labels[train_indices], numel=numel)
            smote_end = smote_labels[:,None,-1]
            xtrain, xval = np.concatenate((self.features_c[train_indices], smote_features), axis=0), self.features_c[val_indices]
            ytrain, yval = np.concatenate((self.labels[train_indices], smote_labels), axis=0), self.labels[val_indices]
            yendtrain, yendval = np.concatenate((self.labels_end[train_indices], smote_end), axis=0), self.labels_end[val_indices]
            fl_store.append(
                (Features_labels(xtrain, yendtrain, ytrain, scaler=self.scaler, normalise_labels=self.normalise_labels,
                                 labels_scaler=self.labels_scaler, labels_end_scaler=self.labels_end_scaler,
                                 norm_mask=self.norm_mask, features_c_names=self.features_c_names,labels_names=self.labels_names,
                                 label_type=self.label_type),
                 Features_labels(xval, yendval, yval, idx=xval_idx, scaler=self.scaler,
                                 normalise_labels=self.normalise_labels,
                                 labels_scaler=self.labels_scaler, labels_end_scaler=self.labels_end_scaler,
                                 norm_mask=self.norm_mask, features_c_names=self.features_c_names,labels_names=self.labels_names,
                                 label_type=self.label_type))
            )
        return fl_store

    def fold_invariant_kf_augment(self, numel, k_folds, shuffle=True):
        """
        Same as kf above. But appends all smote data to each fold's training examples. Validation examples no change.
        :param smote_excel:
        :param k_folds:
        :param shuffle:
        :return:
        """
        fl_store = []
        # Instantiate the cross validator
        skf = KFold(n_splits=k_folds, shuffle=shuffle)
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(skf.split(self.features_c, self.labels)):
            # Generate batches from indices
            xval_idx = self.idx[val_indices]
            smote_features, smote_labels = produce_invariant(self.features_c[train_indices],
                                                         self.labels[train_indices], numel=numel)
            smote_end = smote_labels[:,None,-1]
            xtrain, xval = np.concatenate((self.features_c[train_indices], smote_features), axis=0), self.features_c[val_indices]
            ytrain, yval = np.concatenate((self.labels[train_indices], smote_labels), axis=0), self.labels[val_indices]
            yendtrain, yendval = np.concatenate((self.labels_end[train_indices], smote_end), axis=0), self.labels_end[val_indices]
            fl_store.append(
                (Features_labels(xtrain, yendtrain, ytrain, scaler=self.scaler, normalise_labels=self.normalise_labels,
                                 labels_scaler=self.labels_scaler, labels_end_scaler=self.labels_end_scaler,
                                 norm_mask=self.norm_mask, features_c_names=self.features_c_names,labels_names=self.labels_names,
                                 label_type=self.label_type),
                 Features_labels(xval, yendval, yval, idx=xval_idx, scaler=self.scaler,
                                 normalise_labels=self.normalise_labels,
                                 labels_scaler=self.labels_scaler, labels_end_scaler=self.labels_end_scaler,
                                 norm_mask=self.norm_mask, features_c_names=self.features_c_names,labels_names=self.labels_names,
                                 label_type=self.label_type))
            )
        return fl_store

    def create_loocv(self):
        '''
        Create list of tuples containing (fl_train,fl_val) fl objects for leave one out cross validation
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        loocv = LeaveOneOut()
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(loocv.split(self.features_c, self.labels)):
            # Generate batches from indices
            xval_idx = self.idx[val_indices]
            xtrain, xval = self.features_c[train_indices], self.features_c[val_indices]
            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            fl_store.append(
                (Features_labels(xtrain, ytrain, scaler=self.scaler, labels_scaler=self.labels_scaler,
                                 norm_mask=self.norm_mask),
                 Features_labels(xval, yval, idx=xval_idx, scaler=self.scaler, labels_scaler=self.labels_scaler,
                                 norm_mask=self.norm_mask))
            )
        return fl_store

    def write_data_to_excel(self, loader_excel_file='./excel/data_loader.xlsx'):
        # Excel writing part
        wb = load_workbook(loader_excel_file)
        # Setting up subset features and labels sheet
        sheet_name_store = ['temp_features_c', 'temp_labels']
        ss_store = [self.features_c, self.labels]
        axis_store = [2, 0]  # Because feature_c is 2D while labels is col vector, so order of axis is 2,0
        for cnt, sheet_name in enumerate(sheet_name_store):
            if sheet_name in wb.sheetnames:
                # If temp sheet exists, remove it to create a new one. Else skip this.
                idx = wb.sheetnames.index(sheet_name)  # index of temp sheet
                wb.remove(wb.worksheets[idx])  # remove temp
                wb.create_sheet(sheet_name, idx)  # create an empty sheet using old index
            else:
                wb.create_sheet(sheet_name)  # Create the new sheet
            # Print array to the correct worksheet
            print_array_to_excel(ss_store[cnt], (2, 1), wb[sheet_name_store[cnt]], axis=axis_store[cnt])
        wb.save(loader_excel_file)
        wb.close()




class Features_labels_grid:
    def __init__(self, features, labels, idx=None):
        """
        Creates fl class with a lot useful attributes for grid data classification
        :param features:
        :param labels: Labels as np array, no. of examples x dim
        :param scaler: Scaler to transform features c. If given, use given MinMax scaler from sklearn,
        else create scaler based on given features c.
        """
        # Setting up features
        self.count = features.shape[0]
        self.features = np.copy(features)
        self.features_dim = features.shape[1]

        # idx is for when doing k-fold cross validation, to keep track of which examples are in the val. set
        if isinstance(idx, np.ndarray):
            self.idx = idx
        else:
            self.idx = np.arange(self.count)

        # Setting up labels
        self.labels = np.array(labels)
        self.labels_dim = 1

    def create_kf(self, k_folds, shuffle=True):
        '''
        Almost the same as skf except can work for regression labels and folds are not stratified.
        Create list of tuples containing (fl_train,fl_val) fl objects for k fold cross validation
        :param k_folds: Number of folds
        :return: List of tuples
        '''
        fl_store = []
        # Instantiate the cross validator
        skf = KFold(n_splits=k_folds, shuffle=shuffle)
        # Loop through the indices the split() method returns
        for _, (train_indices, val_indices) in enumerate(skf.split(self.features, self.labels)):
            # Generate batches from indices
            xval_idx = self.idx[val_indices]
            xtrain, xval = self.features[train_indices], self.features[val_indices]
            ytrain, yval = self.labels[train_indices], self.labels[val_indices]
            fl_store.append(
                (Features_labels_grid(xtrain, ytrain),
                 Features_labels_grid(xval, yval, idx=xval_idx)
                 )
            )
        return fl_store
