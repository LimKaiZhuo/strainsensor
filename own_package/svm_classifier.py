from sklearn.svm import SVC
from sklearn.metrics import classification_report, confusion_matrix, matthews_corrcoef
import pickle, time, gc
import numpy as np
import pandas as pd
from pandas import Series
import openpyxl
from openpyxl import load_workbook

from own_package.others import print_array_to_excel, create_results_directory


class SVMmodel:
    def __init__(self, fl, gamma=1):
        self.features_dim = fl.features_dim
        self.labels_dim = fl.labels_dim  # Assuming that each task has only 1 dimensional output
        self.model = SVC(kernel='rbf', gamma=gamma,probability=True)

    def train_model(self, fl):
        training_features = fl.features
        training_labels = fl.labels
        self.model.fit(training_features, training_labels)
        return self.model

    def eval(self, eval_fl):
        features = eval_fl.features
        y_pred = self.model.predict(features)
        return y_pred


def run_classification(read_dir, write_dir, gamma=1):
    # Load fl class
    with open(read_dir + '/grid_data', 'rb') as handle:
        fl = pickle.load(handle)

    fl_store = fl.create_kf(k_folds=10, shuffle=True)

    # Run k model instance to perform skf
    predicted_labels_store = []
    folds = []
    val_idx = []
    val_features = []
    val_labels = []
    for fold, fl_tuple in enumerate(fl_store):
        instance_start = time.time()

        (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl

        model = SVMmodel(fl=ss_fl, gamma=gamma)
        model.train_model(fl=ss_fl)

        # Evaluation
        predicted_labels = model.eval(i_ss_fl)
        predicted_labels_store.extend(predicted_labels)

        # Saving model
        save_model_name = write_dir + '/models/svm_' + str(fold + 1)
        print('Saving instance {} model in {}'.format(fold + 1, save_model_name))
        with open(save_model_name, 'wb') as handle:
            pickle.dump(model, handle, protocol=pickle.HIGHEST_PROTOCOL)

        del model
        gc.collect()

        # Preparing data to put into new_df that consists of all the validation dataset and its predicted labels
        folds.extend([fold] * i_ss_fl.count)  # Make a col that contains the fold number for each example
        if len(val_features):
            val_features = np.concatenate((val_features, i_ss_fl.features),
                                            axis=0)
        else:
            val_features = i_ss_fl.features

        val_labels.extend(i_ss_fl.labels)
        val_idx.extend(i_ss_fl.idx)

        # Printing one instance summary.
        instance_end = time.time()
        print(
            '\nFor k-fold run {} out of {}. Each fold has {} examples. Time taken for '
            'instance = {}\n'
            '####################################################################################################'
                .format(fold + 1, 10, i_ss_fl.count, instance_end - instance_start))


    # Calculating metrics based on complete validation prediction
    mcc = matthews_corrcoef(y_true=val_labels, y_pred=predicted_labels_store)

    # Creating dataframe to print into excel later.
    new_df = np.concatenate((np.array(folds)[:, None],  # Convert 1d list to col. vector
                             val_features,
                             np.array(val_labels)[:,None],
                             np.array(predicted_labels_store)[:,None])
                            , axis=1)
    headers = ['folds'] + \
              ['PVA', 'CNT'] + \
              ['Labels'] + \
              ['Prediction']

    # val_idx is the original position of the example in the data_loader
    new_df = pd.DataFrame(data=new_df, columns=headers, index=val_idx)

    skf_file = write_dir + '/classifier results.xlsx'
    print('Writing into' + skf_file)
    wb = openpyxl.Workbook()
    wb.save(skf_file)

    # Writing hparam dataframe first
    pd_writer = pd.ExcelWriter(skf_file, engine='openpyxl')
    pd_writer.book = wb
    pd_writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    new_df.to_excel(pd_writer)
    start_col = len(new_df.columns) + 4

    # Writing other subset split, instance per run, and bounds
    ws = wb.sheetnames
    ws = wb[ws[-1]]
    headers = ['mcc']
    values = [mcc]
    print_array_to_excel(np.array(headers), (1, start_col + 1), ws, axis=1)
    print_array_to_excel(np.array(values), (2, start_col + 1), ws, axis=1)
    pd_writer.save()
    pd_writer.close()
    wb.close()