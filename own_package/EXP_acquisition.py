import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from sklearn.metrics import mean_squared_error
import time, os
import shutil

# Own Scripts
from .cross_validation import run_skf
from .features_labels_setup import load_data_to_fl
from .active_learning.acquisition import load_model_ensemble


def model_ensemble_prediction_exp_acq(model_store, i_ss_fl):
    """
    Run prediction given one set of feactures_c_norm input, using all the models in model store.
    :param model_store: List of keras models returned by the def load_model_ensemble
    :param i_ss_fl: Inverse subset selection fl.
    :return: List of metrics.
    """
    features_c_norm = i_ss_fl.features_c_norm
    labels = i_ss_fl.labels
    predictions_store = []
    for model in model_store:
        predictions = model.predict(features_c_norm)
        predictions = np.concatenate((predictions[0], predictions[1]), axis=1).reshape((-1)).tolist()

        '''
        predictions = model.predict(features_c_norm)
        predictions = [x.item() for x in predictions]
        predictions = np.vstack(predictions).reshape((-1))
        predictions = predictions.tolist()
        '''

        predictions_store.append(predictions)
    predictions_mean = np.mean(predictions_store, axis=0)
    predictions_std = np.std(predictions_store, axis=0)
    predictions_error = np.square(predictions_mean - labels).reshape(-1)

    return predictions_mean, predictions_std, predictions_error


def variance_error_experiement(model_mode, loss_mode, norm_mask, labels_norm, loader_file, model_dir, hparams,
                               results_excel='./excel/exp_training_results.xlsx', ):
    main_fl = load_data_to_fl(loader_file, norm_mask=norm_mask)
    loocv_fl_store = main_fl.create_loocv()
    save_model_list = range(main_fl.count)
    save_model_list = ['{}/{}'.format(model_dir, str(x)) for x in save_model_list]

    for directory in save_model_list:
        if not os.path.exists(directory):
            os.makedirs(directory)
        else:
            shutil.rmtree(directory)
            os.makedirs(directory)


    predictions_mean = []
    predictions_std = []
    predictions_abs_error = []
    for fl_tuple, save_model_dir in zip(loocv_fl_store, save_model_list):
        # ss_fl = subset of fl. i_ss_fl = inverse of subset of fl (Contains the 1 example that was left out)
        (ss_fl, i_ss_fl) = fl_tuple

        # Train model on the training subset

        run_skf(model_mode=model_mode, loss_mode=loss_mode, cv_mode='skf', hparams=hparams,
                norm_mask=norm_mask, labels_norm=labels_norm,
                loader_file=loader_file,
                skf_file=results_excel, skf_sheet=None,
                k_folds=10, k_shuffle=True,
                save_model_name=None, save_model=True,
                save_model_dir=save_model_dir + '/',
                plot_name=None)


        # Load trained models
        model_store = load_model_ensemble(save_model_dir)
        m, s, e = model_ensemble_prediction_exp_acq(model_store, i_ss_fl)
        predictions_mean.append(m.tolist())
        predictions_std.append(s.tolist())
        predictions_abs_error.append(e.tolist())

    new_df = np.concatenate((main_fl.features_c,
                             main_fl.labels,
                             np.array(predictions_mean),
                             np.array(predictions_std),
                             np.array(predictions_abs_error)
                             ), axis=1)
    labels_name = list(map(str, main_fl.labels_names))
    predicted_labels_name = ['P_' + str(x) for x in labels_name]
    predicted_labels_std_name = ['P_S_' + str(x) for x in labels_name]
    predicted_labels_error_name = ['P_E_' + str(x) for x in labels_name]
    headers = list(map(str, main_fl.features_c_names)) + \
              list(map(str, main_fl.labels_names)) + \
              predicted_labels_name + predicted_labels_std_name + predicted_labels_error_name

    new_df = pd.DataFrame(data=new_df, columns=headers)

    wb = load_workbook(results_excel)
    wb.create_sheet('Results')
    pd_writer = pd.ExcelWriter(results_excel, engine='openpyxl')
    pd_writer.book = wb
    pd_writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    new_df.to_excel(pd_writer, 'Results')
    pd_writer.save()
    pd_writer.close()
    wb.close()
