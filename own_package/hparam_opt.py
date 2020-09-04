import pandas as pd
import numpy as np
import gc, pickle, time
import openpyxl
from openpyxl import load_workbook
from skopt import gp_minimize, dummy_minimize
from skopt.space import Real, Integer
from skopt.utils import use_named_args
from skopt.plots import plot_convergence
from sklearn.metrics import matthews_corrcoef, mean_squared_error
# Own Scripts
from own_package.models.models import create_hparams
from own_package.svm_classifier import run_classification, SVMmodel
from own_package.svr import SVRmodel, run_svr
from .cross_validation import run_skf, run_skf_train_val_test_error
from own_package.others import print_array_to_excel, create_excel_file, print_df_to_excel


def hparam_opt(model_mode, loss_mode, norm_mask, fl_in, fl_store_in, write_dir, save_model_dir,
               total_run, instance_per_run=3, save_model=False, scoring='mse',
               plot_dir=None):
    """
     names = ['shared_1_l', 'shared_1_h',
              'shared_2_l', 'shared_2_h',
              'ts_1_l', 'ts_1_h',
              'ts_2_l', 'ts_2_h',
              'epochs_l', 'epochs_h',
              'l1_l', 'l1_h']
    :param model_mode:
    :param loader_file:
    :param total_run:
    :param instance_per_run:
    :param hparam_file:
    :return:
    """

    hparam_file = write_dir + '/skf_results.xlsx'

    global run_count, data_store, fl, fl_store
    run_count = 0
    fl = fl_in
    fl_store = fl_store_in

    if model_mode == 'hps':
        bounds = [[5, 200, ],
                  [0, 200, ],
                  [5, 200, ],
                  [0, 200, ],
                  [10, 2000],
                  [0, 0.3]]
        shared_1 = Integer(low=bounds[0][0], high=bounds[0][1], name='shared_1')
        shared_2 = Integer(low=bounds[1][0], high=bounds[1][1], name='shared_2')
        ts_1 = Integer(low=bounds[2][0], high=bounds[2][1], name='ts_1')
        ts_2 = Integer(low=bounds[3][0], high=bounds[3][1], name='ts_2')
        epochs = Integer(low=bounds[4][0], high=bounds[4][1], name='epochs')
        reg_l1 = Real(low=bounds[5][0], high=bounds[5][1], name='reg_l1')
        dimensions = [shared_1, shared_2, ts_1, ts_2, epochs, reg_l1]
        default_parameters = [30, 30, 30, 30, 300, 0.01]

        @use_named_args(dimensions=dimensions)
        def fitness(shared_1, shared_2, ts_1, ts_2, epochs, reg_l1):
            global run_count, best_loss, data_store, fl, best_hparams
            run_count += 1

            hparams = create_hparams(shared_layers=[shared_1, shared_2], ts_layers=[ts_1, ts_2], epochs=epochs,
                                     reg_l1=reg_l1,
                                     verbose=0)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_{}'.format(plot_dir, model_mode, loss_mode, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, cv_mode='skf', hparams=hparams,
                              norm_mask=norm_mask, labels_norm=labels_norm,
                              loader_file=loader_file,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=20, k_shuffle=True,
                              save_model_name='_' + str(run_count) + '_' + str(cnt), save_model=True,
                              save_model_dir='./save/models',
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            if mse_avg < best_loss:
                best_hparams = hparams
                best_loss = mse_avg
            loss = mse_avg
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run MSE {} \n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, mse_avg))
            print(pd.DataFrame(hparams))
            print()
            return loss
    elif model_mode == 'cs':
        bounds = [[5, 200, ],
                  [5, 200, ],
                  [0, 200, ],
                  [10, 2000],
                  [0, 0.3]]
        shared_1 = Integer(low=bounds[0][0], high=bounds[0][1], name='shared_1')
        shared_2 = Integer(low=bounds[1][0], high=bounds[1][1], name='shared_2')
        shared_3 = Integer(low=bounds[2][0], high=bounds[2][1], name='shared_3')
        epochs = Integer(low=bounds[3][0], high=bounds[3][1], name='epochs')
        reg_l1 = Real(low=bounds[4][0], high=bounds[4][1], name='reg_l1')
        dimensions = [shared_1, shared_2, shared_3, epochs, reg_l1]
        default_parameters = [30, 30, 30, 300, 0.01]

        @use_named_args(dimensions=dimensions)
        def fitness(shared_1, shared_2, shared_3, epochs, reg_l1):
            global run_count, best_loss, data_store, fl, best_hparams
            run_count += 1

            hparams = create_hparams(cs_layers=[shared_1, shared_2, shared_3], epochs=epochs,
                                     reg_l1=reg_l1,
                                     verbose=0)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, cv_mode='skf', hparams=hparams,
                              norm_mask=norm_mask, labels_norm=labels_norm,
                              loader_file=loader_file,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True,
                              save_model_name='_' + str(run_count) + '_' + str(cnt), save_model=True,
                              save_model_dir='./save/models',
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            if mse_avg < best_loss:
                best_hparams = hparams
                best_loss = mse_avg
            loss = mse_avg
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run MSE {} \n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, mse_avg))
            print(pd.DataFrame(hparams))
            print()
            return loss

    elif model_mode == 'conv1':
        start_time = time.time()
        bounds = [[1, 400, ],
                  [1, 50, ],
                  [100, 2000]]

        pre = Integer(low=bounds[0][0], high=bounds[0][1], name='pre')
        filters = Integer(low=bounds[1][0], high=bounds[1][1], name='filters')
        epochs = Integer(low=bounds[2][0], high=bounds[2][1], name='epochs')
        dimensions = [pre, filters, epochs]
        default_parameters = [70, 5, 1000]

        @use_named_args(dimensions=dimensions)
        def fitness(pre, filters, epochs):
            global run_count, data_store, fl, fl_store
            run_count += 1
            hparams = create_hparams(pre=pre, filters=filters, epochs=epochs,
                                     reg_l1=0.05, reg_l2=0.05, loss='mse',
                                     verbose=0)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True,
                              save_model_name='hparam_' + str(run_count) + '_' + str(cnt + 1), save_model=save_model,
                              save_model_dir=save_model_dir,
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            loss = mse_avg
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run MSE {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, mse_avg, end_time - start_time))
            return loss
    elif model_mode == 'ann3':
        start_time = time.time()

        # bounds = [[10, 300, ],
        #          [50, 800]]
        bounds = [[50, 900, ],
                  [50, 1500]]

        pre = Integer(low=bounds[0][0], high=bounds[0][1], name='pre')
        epochs = Integer(low=bounds[1][0], high=bounds[1][1], name='epochs')
        dimensions = [pre, epochs]
        default_parameters = [70, 50]

        @use_named_args(dimensions=dimensions)
        def fitness(pre, epochs):
            global run_count, data_store, fl, fl_store
            run_count += 1
            hparams = create_hparams(pre=pre, epochs=epochs, loss='mse',
                                     reg_l1=0.05, reg_l2=0.05,
                                     verbose=0)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True, scoring=scoring,
                              save_model_name='hparam_' + str(run_count) + '_' + str(cnt + 1), save_model=save_model,
                              save_model_dir=save_model_dir,
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            loss = mse_avg
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run {} {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, scoring, mse_avg, end_time - start_time))
            return loss
    elif model_mode == 'dtr':
        start_time = time.time()
        bounds = [[2, 50, ],
                  [1, 500]]

        depth = Integer(low=bounds[0][0], high=bounds[0][1], name='depth')
        num_est = Integer(low=bounds[1][0], high=bounds[1][1], name='num_est')
        dimensions = [depth, num_est]
        default_parameters = [6, 300]

        @use_named_args(dimensions=dimensions)
        def fitness(depth, num_est):
            global run_count, data_store, fl, fl_store
            run_count += 1
            hparams = create_hparams(max_depth=depth, num_est=num_est)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True, scoring=scoring,
                              save_model_name='hparam_' + str(run_count) + '_' + str(cnt + 1), save_model=save_model,
                              save_model_dir=save_model_dir,
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            loss = mse_avg
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run {} {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, scoring, mse_avg, end_time - start_time))
            return loss
    elif model_mode == 'svr':
        start_time = time.time()
        bounds = [[0.00005, 0.001, ],
                  [0.0005, 0.01]]

        epsilon = Real(low=bounds[0][0], high=bounds[0][1], name='epsilon')
        c = Real(low=bounds[1][0], high=bounds[1][1], name='c')
        dimensions = [epsilon, c]
        default_parameters = [0.0001, 0.001]

        @use_named_args(dimensions=dimensions)
        def fitness(epsilon, c):
            global run_count, data_store, fl, fl_store
            run_count += 1
            hparams = create_hparams(epsilon=epsilon, c=c)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True, scoring=scoring,
                              save_model_name='hparam_' + str(run_count) + '_' + str(cnt + 1), save_model=save_model,
                              save_model_dir=save_model_dir,
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            loss = mse_avg
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run {} {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, scoring, mse_avg, end_time - start_time))
            return loss

    search_result = gp_minimize(func=fitness,
                                dimensions=dimensions,
                                acq_func='EI',  # Expected Improvement.
                                n_calls=total_run,
                                x0=default_parameters)

    wb = load_workbook(write_dir + '/hparam_results.xlsx')
    hparam_store = np.array(search_result.x_iters)
    results = np.array(search_result.func_vals)
    index = np.arange(total_run) + 1
    toprint = np.concatenate((index.reshape(-1, 1), hparam_store, results.reshape(-1, 1)), axis=1)
    if model_mode == 'conv1':
        header = np.array(['index', 'pre', 'filters', 'epochs', 'mse'])
    elif model_mode == 'ann3':
        header = np.array(['index', 'pre', 'epochs', 'mse'])
    elif model_mode == 'dtr':
        header = np.array(['index', 'max_depth', 'num_est', 'mse'])
    elif model_mode == 'svr':
        header = np.array(['index', 'epsilon', 'c', 'mse'])
    toprint = np.concatenate((header.reshape(1, -1), toprint), axis=0)
    sheetname = wb.sheetnames[-1]
    ws = wb[sheetname]
    print_array_to_excel(toprint, (1, 1), ws, axis=2)
    wb.save(write_dir + '/hparam_results.xlsx')
    wb.close()


def hparam_opt_training(model_mode, loss_mode, norm_mask, fl_in, fl_store_in, write_dir, save_model_dir,
                        total_run, instance_per_run=3, save_model=False, scoring='mse',
                        plot_dir=None):
    """
     names = ['shared_1_l', 'shared_1_h',
              'shared_2_l', 'shared_2_h',
              'ts_1_l', 'ts_1_h',
              'ts_2_l', 'ts_2_h',
              'epochs_l', 'epochs_h',
              'l1_l', 'l1_h']
    :param model_mode:
    :param loader_file:
    :param total_run:
    :param instance_per_run:
    :param hparam_file:
    :return:
    """

    hparam_file = write_dir + '/skf_results.xlsx'

    global run_count, data_store, fl, fl_store
    run_count = 0
    fl = fl_in
    fl_store = [[fl, fl]]

    if model_mode == 'hps':
        bounds = [[5, 200, ],
                  [0, 200, ],
                  [5, 200, ],
                  [0, 200, ],
                  [10, 2000],
                  [0, 0.3]]
        shared_1 = Integer(low=bounds[0][0], high=bounds[0][1], name='shared_1')
        shared_2 = Integer(low=bounds[1][0], high=bounds[1][1], name='shared_2')
        ts_1 = Integer(low=bounds[2][0], high=bounds[2][1], name='ts_1')
        ts_2 = Integer(low=bounds[3][0], high=bounds[3][1], name='ts_2')
        epochs = Integer(low=bounds[4][0], high=bounds[4][1], name='epochs')
        reg_l1 = Real(low=bounds[5][0], high=bounds[5][1], name='reg_l1')
        dimensions = [shared_1, shared_2, ts_1, ts_2, epochs, reg_l1]
        default_parameters = [30, 30, 30, 30, 300, 0.01]

        @use_named_args(dimensions=dimensions)
        def fitness(shared_1, shared_2, ts_1, ts_2, epochs, reg_l1):
            global run_count, best_loss, data_store, fl, best_hparams
            run_count += 1

            hparams = create_hparams(shared_layers=[shared_1, shared_2], ts_layers=[ts_1, ts_2], epochs=epochs,
                                     reg_l1=reg_l1,
                                     verbose=0)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_{}'.format(plot_dir, model_mode, loss_mode, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, cv_mode='skf', hparams=hparams,
                              norm_mask=norm_mask, labels_norm=labels_norm,
                              loader_file=loader_file,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=20, k_shuffle=True,
                              save_model_name='_' + str(run_count) + '_' + str(cnt), save_model=True,
                              save_model_dir='./save/models',
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            if mse_avg < best_loss:
                best_hparams = hparams
                best_loss = mse_avg
            loss = mse_avg
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run MSE {} \n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, mse_avg))
            print(pd.DataFrame(hparams))
            print()
            return loss
    elif model_mode == 'cs':
        bounds = [[5, 200, ],
                  [5, 200, ],
                  [0, 200, ],
                  [10, 2000],
                  [0, 0.3]]
        shared_1 = Integer(low=bounds[0][0], high=bounds[0][1], name='shared_1')
        shared_2 = Integer(low=bounds[1][0], high=bounds[1][1], name='shared_2')
        shared_3 = Integer(low=bounds[2][0], high=bounds[2][1], name='shared_3')
        epochs = Integer(low=bounds[3][0], high=bounds[3][1], name='epochs')
        reg_l1 = Real(low=bounds[4][0], high=bounds[4][1], name='reg_l1')
        dimensions = [shared_1, shared_2, shared_3, epochs, reg_l1]
        default_parameters = [30, 30, 30, 300, 0.01]

        @use_named_args(dimensions=dimensions)
        def fitness(shared_1, shared_2, shared_3, epochs, reg_l1):
            global run_count, best_loss, data_store, fl, best_hparams
            run_count += 1

            hparams = create_hparams(cs_layers=[shared_1, shared_2, shared_3], epochs=epochs,
                                     reg_l1=reg_l1,
                                     verbose=0)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, cv_mode='skf', hparams=hparams,
                              norm_mask=norm_mask, labels_norm=labels_norm,
                              loader_file=loader_file,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True,
                              save_model_name='_' + str(run_count) + '_' + str(cnt), save_model=True,
                              save_model_dir='./save/models',
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            if mse_avg < best_loss:
                best_hparams = hparams
                best_loss = mse_avg
            loss = mse_avg
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run MSE {} \n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, mse_avg))
            print(pd.DataFrame(hparams))
            print()
            return loss

    elif model_mode == 'conv1':
        start_time = time.time()
        bounds = [[50, 400, ],
                  [1, 50, ],
                  [100, 600]]

        pre = Integer(low=bounds[0][0], high=bounds[0][1], name='pre')
        filters = Integer(low=bounds[1][0], high=bounds[1][1], name='filters')
        epochs = Integer(low=bounds[2][0], high=bounds[2][1], name='epochs')
        dimensions = [pre, filters, epochs]
        default_parameters = [70, 1, 500]

        @use_named_args(dimensions=dimensions)
        def fitness(pre, filters, epochs):
            global run_count, data_store, fl, fl_store
            run_count += 1
            hparams = create_hparams(pre=pre, filters=filters, epochs=epochs,
                                     reg_l1=0.05, reg_l2=0.05,
                                     verbose=0)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True,
                              save_model_name='hparam_' + str(run_count) + '_' + str(cnt + 1), save_model=save_model,
                              save_model_dir=save_model_dir,
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            loss = mse_avg
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run MSE {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, mse_avg, end_time - start_time))
            return loss
    elif model_mode == 'ann3':
        start_time = time.time()

        # bounds = [[10, 300, ],
        #          [50, 800]]
        bounds = [[10, 2000, ],
                  [1000, 10000]]

        pre = Integer(low=bounds[0][0], high=bounds[0][1], name='pre')
        epochs = Integer(low=bounds[1][0], high=bounds[1][1], name='epochs')
        dimensions = [pre, epochs]
        default_parameters = [70, 4000]

        @use_named_args(dimensions=dimensions)
        def fitness(pre, epochs):
            global run_count, data_store, fl, fl_store
            run_count += 1
            hparams = create_hparams(pre=pre, epochs=epochs,
                                     reg_l1=0.0005, reg_l2=0,
                                     verbose=0)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True, scoring=scoring,
                              save_model_name='hparam_' + str(run_count) + '_' + str(cnt + 1), save_model=save_model,
                              save_model_dir=save_model_dir,
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            loss = mse_avg
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run {} {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, scoring, mse_avg, end_time - start_time))
            return loss
    elif model_mode == 'dtr':
        start_time = time.time()
        bounds = [[2, 100, ],
                  [1, 500]]

        depth = Integer(low=bounds[0][0], high=bounds[0][1], name='depth')
        num_est = Integer(low=bounds[1][0], high=bounds[1][1], name='num_est')
        dimensions = [depth, num_est]
        default_parameters = [6, 300]

        @use_named_args(dimensions=dimensions)
        def fitness(depth, num_est):
            global run_count, data_store, fl, fl_store
            run_count += 1
            hparams = create_hparams(max_depth=depth, num_est=num_est)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True, scoring=scoring,
                              save_model_name='hparam_' + str(run_count) + '_' + str(cnt + 1), save_model=save_model,
                              save_model_dir=save_model_dir,
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            loss = mse_avg
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run {} {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, scoring, mse_avg, end_time - start_time))
            return loss
    elif model_mode == 'svr':
        start_time = time.time()
        bounds = [[0.00005, 0.001, ],
                  [0.0005, 0.01]]

        epsilon = Real(low=bounds[0][0], high=bounds[0][1], name='epsilon')
        c = Real(low=bounds[1][0], high=bounds[1][1], name='c')
        dimensions = [epsilon, c]
        default_parameters = [0.0001, 0.001]

        @use_named_args(dimensions=dimensions)
        def fitness(epsilon, c):
            global run_count, data_store, fl, fl_store
            run_count += 1
            hparams = create_hparams(epsilon=epsilon, c=c)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True, scoring=scoring,
                              save_model_name='hparam_' + str(run_count) + '_' + str(cnt + 1), save_model=save_model,
                              save_model_dir=save_model_dir,
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            loss = mse_avg
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run {} {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, scoring, mse_avg, end_time - start_time))
            return loss

    search_result = gp_minimize(func=fitness,
                                dimensions=dimensions,
                                acq_func='EI',  # Expected Improvement.
                                n_calls=total_run,
                                x0=default_parameters)

    wb = load_workbook(write_dir + '/hparam_results.xlsx')
    hparam_store = np.array(search_result.x_iters)
    results = np.array(search_result.func_vals)
    index = np.arange(total_run) + 1
    toprint = np.concatenate((index.reshape(-1, 1), hparam_store, results.reshape(-1, 1)), axis=1)
    if model_mode == 'conv1':
        header = np.array(['index', 'pre', 'filters', 'epochs', 'mse'])
    elif model_mode == 'ann3':
        header = np.array(['index', 'pre', 'epochs', 'mse'])
    elif model_mode == 'dtr':
        header = np.array(['index', 'max_depth', 'num_est', 'mse'])
    elif model_mode == 'svr':
        header = np.array(['index', 'epsilon', 'c', 'mse'])
    toprint = np.concatenate((header.reshape(1, -1), toprint), axis=0)
    sheetname = wb.sheetnames[-1]
    ws = wb[sheetname]
    print_array_to_excel(toprint, (1, 1), ws, axis=2)
    wb.save(write_dir + '/hparam_results.xlsx')
    wb.close()


def hparam_opt_train_val_test(model_mode, loss_mode, norm_mask, fl_in, fl_store_in, test_fl, ett_fl_store,
                              write_dir, save_model_dir,
                              total_run, instance_per_run=3, save_model=False, scoring='mse',
                              plot_dir=None):
    """
    hi
     names = ['shared_1_l', 'shared_1_h',
              'shared_2_l', 'shared_2_h',
              'ts_1_l', 'ts_1_h',
              'ts_2_l', 'ts_2_h',
              'epochs_l', 'epochs_h',
              'l1_l', 'l1_h']
    :param model_mode:
    :param loader_file:
    :param total_run:
    :param instance_per_run:
    :param hparam_file:
    :return:
    """
    hparam_file = write_dir + '/skf_results.xlsx'
    data_store_dir = write_dir + '/data_store'
    # st_df_excel = create_excel_file(write_dir + '/st_df.xlsx')
    # stt_df_excel = create_excel_file(write_dir + '/stt_df.xlsx')
    # solo_summary_excel = create_excel_file(write_dir + '/solo_summary.xlsx')
    # ot_df_excel = create_excel_file(write_dir + 'ot_df.xlsx')
    # ov_df_excel = create_excel_file(write_dir + 'ov_df.xlsx')
    # ott_df_excel = create_excel_file(write_dir + 'ott_df.xlsx')
    global run_count, data_store, fl, fl_store, data_store, data_store_count, data_store_name
    run_count = 0
    fl = fl_in
    fl_store = fl_store_in
    data_store = []

    if model_mode == 'ann3':
        start_time = time.time()
        # bounds = [[10, 300, ],
        #          [50, 800]]
        bounds = [[30, 3000, ],
                  [100, 4000]]

        pre = Integer(low=bounds[0][0], high=bounds[0][1], name='pre')
        epochs = Integer(low=bounds[1][0], high=bounds[1][1], name='epochs')
        dimensions = [pre, epochs]
        default_parameters = [300, 500]
        data_store_count = 1
        data_store_name = 0

        @use_named_args(dimensions=dimensions)
        def fitness(pre, epochs):
            global run_count, data_store, fl, fl_store, data_store, data_store_count, data_store_name
            run_count += 1
            hparams = create_hparams(pre=pre, epochs=epochs, loss='mse', learning_rate=0.001 / 2,
                                     reg_l1=0.0005, reg_l2=0,
                                     verbose=0)

            if plot_dir:
                plot_name = '{}/{}_{}_run_{}'.format(plot_dir, model_mode, loss_mode, run_count)
            else:
                plot_name = None
            val_score, train_score, data = run_skf_train_val_test_error(model_mode=model_mode, loss_mode=loss_mode,
                                                                        fl=fl,
                                                                        fl_store=fl_store, test_fl=test_fl,
                                                                        ett_fl_store=ett_fl_store,
                                                                        model_name='{}_{}_{}'.format(write_dir,
                                                                                                     model_mode,
                                                                                                     run_count),
                                                                        hparams=hparams,
                                                                        k_folds=10, scoring=scoring,
                                                                        save_model_name='hparam_' + str(
                                                                            run_count) + '_',
                                                                        save_model=save_model,
                                                                        save_model_dir=save_model_dir,
                                                                        plot_name=plot_name)
            if (data_store_count - 1) % 5 == 0:
                data_store = []
                data_store_name += 5
            data.append([pre, epochs])
            data_store.append(data)
            with open('{}_{}.pkl'.format(data_store_dir, data_store_name), "wb") as file:
                pickle.dump(data_store, file)
            data_store_count += 1
            loss = (val_score * 2 + train_score) / 3
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run {} {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, scoring, loss, end_time - start_time))
            return loss
    elif model_mode == 'dtr':
        start_time = time.time()
        bounds = [[1, 200, ],
                  [1, 1000]]

        depth = Integer(low=bounds[0][0], high=bounds[0][1], name='depth')
        num_est = Integer(low=bounds[1][0], high=bounds[1][1], name='num_est')
        dimensions = [depth, num_est]
        default_parameters = [3, 300]
        data_store_count = 1
        data_store_name = 0

        @use_named_args(dimensions=dimensions)
        def fitness(depth, num_est):
            global run_count, data_store, fl, fl_store, data_store_count, data_store_name
            run_count += 1
            hparams = create_hparams(max_depth=depth, num_est=num_est)

            if plot_dir:
                plot_name = '{}/{}_{}_run_{}'.format(plot_dir, model_mode, loss_mode, run_count)
            else:
                plot_name = None
            val_score, train_score, data = run_skf_train_val_test_error(model_mode=model_mode, loss_mode=loss_mode,
                                                                        fl=fl,
                                                                        fl_store=fl_store, test_fl=test_fl,
                                                                        ett_fl_store=ett_fl_store,
                                                                        model_name='{}_{}_{}'.format(write_dir,
                                                                                                     model_mode,
                                                                                                     run_count),
                                                                        hparams=hparams,
                                                                        k_folds=10, scoring=scoring,
                                                                        save_model_name='hparam_' + str(
                                                                            run_count) + '_',
                                                                        save_model=save_model,
                                                                        save_model_dir=save_model_dir,
                                                                        plot_name=plot_name)

            # loss = (val_score + train_score) / 2
            loss = val_score
            if (data_store_count - 1) % 5 == 0:
                data_store = []
                data_store_name += 5
            data.append([depth, num_est])
            data_store.append(data)
            with open('{}_{}.pkl'.format(data_store_dir, data_store_name), "wb") as file:
                pickle.dump(data_store, file)

            data_store_count += 1
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Depth {}, No. Estimators {}\n'
                  'Instance per run {} \n'
                  'Current run {} {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, depth, num_est, instance_per_run, scoring, loss, end_time - start_time))
            return loss
    elif model_mode == 'xgb':
        start_time = time.time()
        hparam_opt_params = {'max_depth': {'type': 'Integer', 'lower': 1, 'upper': 6},
                             'num_boost_round': {'type': 'Integer', 'lower': 10, 'upper': 1200},
                             'subsample': {'type': 'Real', 'lower': 0.5, 'upper': 1},
                             'gamma': {'type': 'Real', 'lower': 0, 'upper': 5}
                             }
        dimensions = []
        for k, v in hparam_opt_params.items():
            if v['type'] == 'Real':
                dimensions.append(Real(v['lower'], v['upper'], name=k))
            elif v['type'] == 'Integer':
                dimensions.append(Integer(v['lower'], v['upper'], name=k))
            else:
                raise TypeError('hparam opt bounds variable type must be Real or Integer only.')

        default_parameters = [3, 0.5, 0.1]
        data_store_count = 1
        data_store_name = 0

        @use_named_args(dimensions=dimensions)
        def fitness(**params):
            global run_count, data_store, fl, fl_store, data_store_count, data_store_name
            run_count += 1
            hparams = params

            if plot_dir:
                plot_name = '{}/{}_{}_run_{}'.format(plot_dir, model_mode, loss_mode, run_count)
            else:
                plot_name = None
            val_score, train_score, data = run_skf_train_val_test_error(model_mode=model_mode, loss_mode=loss_mode,
                                                                        fl=fl,
                                                                        fl_store=fl_store, test_fl=test_fl,
                                                                        ett_fl_store=ett_fl_store,
                                                                        model_name='{}_{}_{}'.format(write_dir,
                                                                                                     model_mode,
                                                                                                     run_count),
                                                                        hparams=hparams,
                                                                        k_folds=10, scoring=scoring,
                                                                        save_model_name='hparam_' + str(
                                                                            run_count) + '_',
                                                                        save_model=save_model,
                                                                        save_model_dir=save_model_dir,
                                                                        plot_name=plot_name)

            # loss = (val_score + train_score) / 2
            loss = val_score
            if (data_store_count - 1) % 5 == 0:
                data_store = []
                data_store_name += 5
            data.append([depth, num_est])
            data_store.append(data)
            with open('{}_{}.pkl'.format(data_store_dir, data_store_name), "wb") as file:
                pickle.dump(data_store, file)

            data_store_count += 1
            end_time = time.time()
            print(f'**************************************************************************************************\n'
                  f'Run Number {run_count} \n'
                  f'{params}\n'
                  f'Instance per run {instance_per_run} \n'
                  f'Current run {scoring} {loss} \n'
                  f'Time Taken: {end_time - start_time}\n'
                  '*********************************************************************************************')
            return loss
    elif model_mode == 'svr':
        start_time = time.time()
        bounds = [[0.00005, 0.001, ],
                  [0.0005, 0.01]]

        epsilon = Real(low=bounds[0][0], high=bounds[0][1], name='epsilon')
        c = Real(low=bounds[1][0], high=bounds[1][1], name='c')
        dimensions = [epsilon, c]
        default_parameters = [0.0001, 0.001]

        @use_named_args(dimensions=dimensions)
        def fitness(epsilon, c):
            global run_count, data_store, fl, fl_store
            run_count += 1
            hparams = create_hparams(epsilon=epsilon, c=c)

            mse_avg = 0

            for cnt in range(instance_per_run):
                if plot_dir:
                    plot_name = '{}/{}_{}_run_{}_count_{}'.format(plot_dir, model_mode, loss_mode, run_count, cnt)
                else:
                    plot_name = None
                mse = run_skf(model_mode=model_mode, loss_mode=loss_mode, fl=fl, fl_store=fl_store, hparams=hparams,
                              skf_file=hparam_file, skf_sheet='_' + str(run_count) + '_' + str(cnt),
                              k_folds=10, k_shuffle=True, scoring=scoring,
                              save_model_name='hparam_' + str(run_count) + '_' + str(cnt + 1), save_model=save_model,
                              save_model_dir=save_model_dir,
                              plot_name=plot_name)
                mse_avg += mse

            mse_avg = mse_avg / instance_per_run
            loss = mse_avg
            end_time = time.time()
            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Instance per run {} \n'
                  'Current run {} {} \n'
                  'Time Taken: {}\n'
                  '*********************************************************************************************'.format(
                run_count, instance_per_run, scoring, mse_avg, end_time - start_time))
            return loss

    search_result = gp_minimize(func=fitness,
                                dimensions=dimensions,
                                acq_func='EI',  # Expected Improvement.
                                n_calls=total_run,
                                x0=default_parameters)

    wb = load_workbook(write_dir + '/hparam_results.xlsx')
    hparam_store = np.array(search_result.x_iters)
    results = np.array(search_result.func_vals)
    index = np.arange(total_run) + 1
    toprint = np.concatenate((index.reshape(-1, 1), hparam_store, results.reshape(-1, 1)), axis=1)
    if model_mode == 'conv1':
        header = np.array(['index', 'pre', 'filters', 'epochs', 'mse'])
    elif model_mode == 'ann3':
        header = np.array(['index', 'pre', 'epochs', 'mse'])
    elif model_mode == 'dtr':
        header = np.array(['index', 'max_depth', 'num_est', 'mse'])
    elif model_mode == 'svr':
        header = np.array(['index', 'epsilon', 'c', 'mse'])
    toprint = np.concatenate((header.reshape(1, -1), toprint), axis=0)
    sheetname = wb.sheetnames[-1]
    ws = wb[sheetname]
    print_array_to_excel(toprint, (1, 1), ws, axis=2)
    wb.save(write_dir + '/hparam_results.xlsx')
    wb.close()


def read_hparam_data(data_store, write_dir, ett_names, print_s_df, trainset_ett_idx=None):
    numel_ett = len(ett_names)
    if print_s_df:
        st_df_excel = create_excel_file(write_dir + '/st_df.xlsx')
        st_df_wb = openpyxl.load_workbook(st_df_excel)
        stt_df_excel = create_excel_file(write_dir + '/stt_df.xlsx')
        stt_df_wb = openpyxl.load_workbook(stt_df_excel)
        sett_df_excel_store = [create_excel_file('{}/sett_df_{}.xlsx'.format(write_dir, ett_name)) for ett_name in
                               ett_names]
        sett_df_wb_store = [openpyxl.load_workbook(ett_df_excel) for ett_df_excel in sett_df_excel_store]

    solo_summary_excel = create_excel_file(write_dir + '/solo_summary.xlsx')
    solo_summary_wb = openpyxl.load_workbook(solo_summary_excel)
    ws = solo_summary_wb['Sheet']
    print_array_to_excel(
        array=['Trial', 'Fold', 'name', 'Train MSE', 'Train MRE', 'Test MSE', 'Test MRE'] \
              + [x + '_MSE' for x in ett_names] + [x + '_MRE' for x in ett_names] + ['hparam1', 'hparam2'],
        first_cell=(1, 1), ws=ws, axis=1)
    solo_row = 2
    ot_df_excel = create_excel_file(write_dir + '/ot_df.xlsx')
    ot_df_wb = openpyxl.load_workbook(ot_df_excel)
    ov_df_excel = create_excel_file(write_dir + '/ov_df.xlsx')
    ov_df_wb = openpyxl.load_workbook(ov_df_excel)
    ott_df_excel = create_excel_file(write_dir + '/ott_df.xlsx')
    ott_df_wb = openpyxl.load_workbook(ott_df_excel)
    oett_df_excel_store = [create_excel_file('{}/oett_df_{}.xlsx'.format(write_dir, ett_name)) for ett_name in
                           ett_names]
    oett_df_wb_store = [openpyxl.load_workbook(ett_df_excel) for ett_df_excel in oett_df_excel_store]
    overall_summary_excel = create_excel_file(write_dir + '/overall_summary.xlsx')
    overall_summary_wb = openpyxl.load_workbook(overall_summary_excel)
    ws = overall_summary_wb['Sheet']
    if trainset_ett_idx:
        print_array_to_excel(
            array=['Trial', 'name', 'Train MSE', 'Train MRE', 'Val MSE', 'Val MRE', 'Test MSE', 'Test MRE',
                   'un125Train MSE', 'un125Train MRE'] \
                  + [x + '_MSE' for x in ett_names] + [x + '_MRE' for x in ett_names] + ['hparam1', 'hparam2'],
            first_cell=(1, 1), ws=ws, axis=1)
    else:
        print_array_to_excel(
            array=['Trial', 'name', 'Train MSE', 'Train MRE', 'Val MSE', 'Val MRE', 'Test MSE', 'Test MRE'] \
                  + [x + '_MSE' for x in ett_names] + [x + '_MRE' for x in ett_names] + ['hparam1', 'hparam2'],
            first_cell=(1, 1), ws=ws, axis=1)
    overall_row = 2

    for trial, data in enumerate(data_store):
        fold_numel = len(data[0][0])
        if print_s_df:
            for fold, (name, st_df, stt_df, sett_df_store) in enumerate(
                    zip(*([data[0][0]] + [data[i] for i in [2, 3, 9]]))):
                # Loop 10 times for 10CV
                name = name.partition('hparams_opt')[-1]
                if len(name) > 30:
                    name = name[-30:]
                st_df_wb.create_sheet(name)
                ws = st_df_wb[name]
                print_df_to_excel(df=st_df, ws=ws)
                stt_df_wb.create_sheet(name)
                ws = stt_df_wb[name]
                print_df_to_excel(df=stt_df, ws=ws)
                for sett_df_wb, sett_df in zip(sett_df_wb_store, sett_df_store):
                    # Loop 15 times for 15 different invariant testset
                    sett_df_wb.create_sheet(name)
                    ws = sett_df_wb[name]
                    print_df_to_excel(df=sett_df, ws=ws)

        ws = solo_summary_wb['Sheet']
        trial_name = [trial] * fold_numel
        fold_name = list(range(1, fold_numel + 1))
        print_array_to_excel(array=trial_name, first_cell=(solo_row, 1), ws=ws, axis=0)
        print_array_to_excel(array=fold_name, first_cell=(solo_row, 2), ws=ws, axis=0)
        print_array_to_excel(array=data[0][0], first_cell=(solo_row, 3), ws=ws, axis=0)
        print_array_to_excel(array=data[0][1], first_cell=(solo_row, 4), ws=ws, axis=0)
        print_array_to_excel(array=data[0][2], first_cell=(solo_row, 5), ws=ws, axis=0)
        print_array_to_excel(array=data[0][3], first_cell=(solo_row, 6), ws=ws, axis=0)
        print_array_to_excel(array=data[0][4], first_cell=(solo_row, 7), ws=ws, axis=0)
        sett_mse = [[x[idx] for x in data[7][0]] for idx in range(numel_ett)]
        sett_mre = [[x[idx] for x in data[7][1]] for idx in range(numel_ett)]
        for mse, mre, col1, col2 in zip(sett_mse, sett_mre, list(range(8, 8 + numel_ett)),
                                        list(range(8 + numel_ett, 8 + 2 * numel_ett))):
            print_array_to_excel(array=mse, first_cell=(solo_row, col1), ws=ws, axis=0)
            print_array_to_excel(array=mre, first_cell=(solo_row, col2), ws=ws, axis=0)
        try:
            print_array_to_excel(array=[data[-1][0]] * fold_numel, first_cell=(solo_row, 8 + 2 * numel_ett), ws=ws,
                                 axis=0)
            print_array_to_excel(array=[data[-1][1]] * fold_numel, first_cell=(solo_row, 9 + 2 * numel_ett), ws=ws,
                                 axis=0)
        except IndexError:
            pass
        solo_row += fold_numel

        name = data[1][0].partition('hparams_opt')[-1]

        if name == "":
            name = data[1][0].partition('/results/')[-1]

        if len(name) > 30:
            name = name[-30:]
        ot_df_wb.create_sheet(name)
        ws = ot_df_wb[name]
        print_df_to_excel(df=data[4], ws=ws)
        ov_df_wb.create_sheet(name)
        ws = ov_df_wb[name]
        print_df_to_excel(df=data[5], ws=ws)
        ott_df_wb.create_sheet(name)
        ws = ott_df_wb[name]
        print_df_to_excel(df=data[6], ws=ws)
        for oett_df_wb, oett_df in zip(oett_df_wb_store, data[10]):
            # Loop 15 times for 15 different invariant testset
            oett_df_wb.create_sheet(name)
            ws = oett_df_wb[name]
            print_df_to_excel(df=oett_df, ws=ws)

        ws = overall_summary_wb['Sheet']
        ws.cell(overall_row, 1).value = trial
        ws.cell(overall_row, 2).value = data[1][0]
        ws.cell(overall_row, 3).value = data[1][1]
        ws.cell(overall_row, 4).value = data[1][2]
        ws.cell(overall_row, 5).value = data[1][3]
        ws.cell(overall_row, 6).value = data[1][4]
        ws.cell(overall_row, 7).value = data[1][5]
        ws.cell(overall_row, 8).value = data[1][6]
        col = 9
        if trainset_ett_idx:
            untrainset_df = data[10][trainset_ett_idx].copy(deep=True)
            ov_df = data[5]
            untrainset_df.iloc[:ov_df.shape[0], -3:] = ov_df.iloc[:, -3:]
            y = untrainset_df.iloc[:, :3].values
            p_y = untrainset_df.iloc[:, -3:].values
            mse = np.mean((y - p_y) ** 2)
            mre = np.mean(np.abs(y - p_y).T / y[:, -1])
            ws.cell(overall_row, col).value = mse
            ws.cell(overall_row, col + 1).value = mre
            col += 2

        print_array_to_excel(array=data[8][0], first_cell=(overall_row, col), ws=ws, axis=1)
        print_array_to_excel(array=data[8][1], first_cell=(overall_row, col + numel_ett), ws=ws, axis=1)
        try:
            ws.cell(overall_row, col + 2 * numel_ett).value = data[-1][0]
            ws.cell(overall_row, col + 1 + 2 * numel_ett).value = data[-1][1]
        except IndexError:
            pass
        overall_row += 1
    if print_s_df:
        st_df_wb.save(st_df_excel)
        stt_df_wb.save(stt_df_excel)
        [sett_df_wb.save(sett_df_excel) for sett_df_wb, sett_df_excel in zip(sett_df_wb_store, sett_df_excel_store)]
    solo_summary_wb.save(solo_summary_excel)
    ot_df_wb.save(ot_df_excel)
    ov_df_wb.save(ov_df_excel)
    ott_df_wb.save(ott_df_excel)
    [sett_df_wb.save(sett_df_excel) for sett_df_wb, sett_df_excel in zip(oett_df_wb_store, oett_df_excel_store)]
    overall_summary_wb.save(overall_summary_excel)


def grid_hparam_opt(fl, total_run):
    """
     names = ['shared_1_l', 'shared_1_h',
              'shared_2_l', 'shared_2_h',
              'ts_1_l', 'ts_1_h',
              'ts_2_l', 'ts_2_h',
              'epochs_l', 'epochs_h',
              'l1_l', 'l1_h']
    :param model_mode:
    :param loader_file:
    :param total_run:
    :param instance_per_run:
    :param hparam_file:
    :return:
    """

    global run_count, best_loss, data_store, fl_store, best_hparams
    run_count = 0
    best_loss = 100000

    gamma = Real(low=0.1, high=300, name='gamma')
    dimensions = [gamma]
    default_parameters = [30]

    fl_store = fl.create_kf(k_folds=10, shuffle=True)

    @use_named_args(dimensions=dimensions)
    def fitness(gamma):
        global run_count, best_loss, fl_store
        run_count += 1

        # Run k model instance to perform skf
        predicted_labels_store = []
        val_labels = []
        for fold, fl_tuple in enumerate(fl_store):
            (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl

            model = SVMmodel(fl=ss_fl, gamma=gamma)
            model.train_model(fl=ss_fl)

            # Evaluation
            predicted_labels = model.eval(i_ss_fl)
            predicted_labels_store.extend(predicted_labels)
            val_labels.extend(i_ss_fl.labels)
            del model
            gc.collect()

        # Calculating metrics based on complete validation prediction
        mcc = matthews_corrcoef(y_true=val_labels, y_pred=predicted_labels_store)

        print('**************************************************************************************************\n'
              'Run Number {} \n'
              'Current run MSE {} \n'
              '*********************************************************************************************'.format(
            run_count, mcc))
        return -mcc

    search_result = gp_minimize(func=fitness,
                                dimensions=dimensions,
                                acq_func='EI',  # Expected Improvement.
                                n_calls=total_run,
                                x0=default_parameters)
    plot_convergence(search_result)
    print('Best Loss = {}'.format(search_result.fun))
    print('Best Gamma = {}'.format(search_result.x[0]))


def svr_hparam_opt(fl, total_run):
    """
     names = ['shared_1_l', 'shared_1_h',
              'shared_2_l', 'shared_2_h',
              'ts_1_l', 'ts_1_h',
              'ts_2_l', 'ts_2_h',
              'epochs_l', 'epochs_h',
              'l1_l', 'l1_h']
    :param model_mode:
    :param loader_file:
    :param total_run:
    :param instance_per_run:
    :param hparam_file:
    :return:
    """

    global run_count, best_loss, data_store, fl_store, best_hparams
    run_count = 0
    best_loss = 100000

    gamma = Real(low=0.1, high=300, name='gamma')
    dimensions = [gamma]
    default_parameters = [30]

    fl_store = fl.create_kf(k_folds=10, shuffle=True)

    @use_named_args(dimensions=dimensions)
    def fitness(gamma):
        global run_count, best_loss, fl_store
        run_count += 1

        # Run k model instance to perform skf
        predicted_labels_store = []
        val_labels = []
        for fold, fl_tuple in enumerate(fl_store):
            (ss_fl, i_ss_fl) = fl_tuple  # ss_fl is training fl, i_ss_fl is validation fl

            model = SVRmodel(fl=ss_fl, gamma=gamma)
            model.train_model(fl=ss_fl)

            # Evaluation
            predicted_labels = model.eval(i_ss_fl).flatten().tolist()
            predicted_labels_store.extend(predicted_labels)
            val_labels.extend(i_ss_fl.labels_end.flatten().tolist())
            del model
            gc.collect()

        # Calculating metrics based on complete validation prediction
        mse = mean_squared_error(y_true=val_labels, y_pred=predicted_labels_store)

        print('**************************************************************************************************\n'
              'Run Number {} \n'
              'Current run MSE {} \n'
              '*********************************************************************************************'.format(
            run_count, mse))
        return mse

    search_result = gp_minimize(func=fitness,
                                dimensions=dimensions,
                                acq_func='EI',  # Expected Improvement.
                                n_calls=total_run,
                                x0=default_parameters)
    plot_convergence(search_result)
    print('Best Loss = {}'.format(search_result.fun))
    print('Best Gamma = {}'.format(search_result.x[0]))


def ann_end_hparam_opt(fl_in, total_run, model_selector, write_dir, excel_dir, hparams_excel_dir):
    """
     names = ['shared_1_l', 'shared_1_h',
              'shared_2_l', 'shared_2_h',
              'ts_1_l', 'ts_1_h',
              'ts_2_l', 'ts_2_h',
              'epochs_l', 'epochs_h',
              'l1_l', 'l1_h']
    :param model_mode:
    :param loader_file:
    :param total_run:
    :param instance_per_run:
    :param hparam_file:
    :return:
    """

    global run_count, best_loss, data_store, fl_store, best_hparams
    fl_store = fl_in
    run_count = 0
    best_loss = 100000

    if model_selector == 'svr':
        gamma = Real(low=0.1, high=300, name='gamma')
        dimensions = [gamma]
        default_parameters = [30]
    elif model_selector == 'ann':
        bounds = [[3, 100, ],
                  [3, 100, ],
                  [80, 2500]]
        shared = Integer(low=bounds[0][0], high=bounds[0][1], name='shared')
        shared2 = Integer(low=bounds[1][0], high=bounds[1][1], name='shared2')
        epochs = Integer(low=bounds[2][0], high=bounds[2][1], name='epochs')
        dimensions = [shared, shared2, epochs]
        default_parameters = [5, 5, 100]
    else:
        raise KeyError('model_selector argument is not a valid model')

    if model_selector == 'svr':
        @use_named_args(dimensions=dimensions)
        def fitness(gamma):
            global run_count, fl_store
            run_count += 1
            mse = run_svr(fl_store, write_dir=write_dir, excel_dir=excel_dir,
                          model_selector=model_selector, gamma=gamma)

            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Current run MSE {} \n'
                  '*********************************************************************************************'.format(
                run_count, mse))
            return mse
    elif model_selector == 'ann':
        @use_named_args(dimensions=dimensions)
        def fitness(shared, shared2, epochs):
            global run_count, fl_store
            run_count += 1
            hparams = create_hparams(shared_layers=[shared, shared2], epochs=epochs,
                                     reg_l1=0.05, reg_l2=0.05,
                                     activation='relu', batch_size=16, verbose=0)
            mse = run_svr(fl_store, write_dir=write_dir, excel_dir=excel_dir,
                          model_selector=model_selector, hparams=hparams, save_name=str(run_count))

            print('**************************************************************************************************\n'
                  'Run Number {} \n'
                  'Current run MSE {} \n'
                  '*********************************************************************************************'.format(
                run_count, mse))
            return mse

    search_result = gp_minimize(func=fitness,
                                dimensions=dimensions,
                                acq_func='EI',  # Expected Improvement.
                                n_calls=total_run,
                                x0=default_parameters)
    plot_convergence(search_result)
    print('Best Loss = {}'.format(search_result.fun))
    print('Best Gamma = {}'.format(search_result.x[0]))

    wb = load_workbook(write_dir + '/hparam_results.xlsx')
    hparam_store = np.array(search_result.x_iters)
    results = np.array(search_result.func_vals)
    index = np.arange(total_run)
    toprint = np.concatenate((index.reshape(-1, 1), hparam_store, results.reshape(-1, 1)), axis=1)
    if model_selector == 'svr':
        header = np.array(['index', 'gamma', 'mse'])
    elif model_selector == 'ann':
        header = np.array(['index', 'shared', 'shared2', 'epochs', 'mse'])
    toprint = np.concatenate((header.reshape(1, -1), toprint), axis=0)
    sheetname = wb.sheetnames[-1]
    ws = wb[sheetname]
    print_array_to_excel(toprint, (1, 1), ws, axis=2)
    wb.save(write_dir + '/hparam_results.xlsx')
    wb.close()
