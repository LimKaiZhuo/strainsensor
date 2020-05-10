import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import openpyxl
from openpyxl import load_workbook
#import umap
import pickle, os
from own_package.others import print_df_to_excel

def draw_umap(data, scale,  n_neighbors=15, min_dist=0.1, n_components=2, metric='euclidean', title=''):
    fit = umap.UMAP(
        n_neighbors=n_neighbors,
        min_dist=min_dist,
        n_components=n_components,
        metric=metric
    )
    u = fit.fit_transform(data)
    fig = plt.figure()
    if n_components == 1:
        ax = fig.add_subplot(111)
        ax.scatter(u[:,0], range(len(u)), c=scale)
    if n_components == 2:
        ax = fig.add_subplot(111)
        ax.scatter(u[:,0], u[:,1], c=scale)
    if n_components == 3:
        ax = fig.add_subplot(111, projection='3d')
        ax.scatter(u[:,0], u[:,1], u[:,2], c=scale, s=100)
    plt.title(title, fontsize=18)
    plt.savefig(plot_directory + '/Exp ' + str(idx + 1) + ' ' + mode + '.png', bbox_inches='tight')
    plt.close()


def read_excel_acquisition_data(write_dir, excel_file):
    results_directory = write_dir + '/acq_fl_data'
    if os.path.exists(results_directory):
        expand = 1
        while True:
            expand += 1
            new_results_directory = results_directory + str(expand)
            if os.path.exists(new_results_directory):
                continue
            else:
                results_directory = new_results_directory
                break
    os.mkdir(results_directory)
    print('Creating new results directory: {}'.format(results_directory))

    xls = pd.ExcelFile(excel_file)
    sheet_names = xls.sheet_names[1:]
    print('Loading the following spreadsheets: {} \n From {}'.format(sheet_names, excel_file))
    features_store = []
    labels_store = []
    for sheet in sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet, usecols=[1,2,3,4,5]).values
        features_store.append(df[:, :-1])
        labels_store.append(df[:, -1])

    with open(results_directory + '/features_store', 'wb') as handle:
        pickle.dump(features_store, handle, protocol=pickle.HIGHEST_PROTOCOL)

    with open(results_directory + '/labels_store', 'wb') as handle:
        pickle.dump(labels_store, handle, protocol=pickle.HIGHEST_PROTOCOL)


def plot_all_umap(read_dir,  n_neighbors=15, min_dist=0.1, n_components=2, metric='euclidean'):
    with open(read_dir + '/features_store', 'rb') as handle:
        features_store = pickle.load(handle)
    with open(read_dir + '/labels_store', 'rb') as handle:
        labels_store = pickle.load(handle)
    for idx, store in enumerate(zip(features_store,labels_store)):
        features, labels = store
        draw_umap(data=features, scale=labels,  n_neighbors=n_neighbors, min_dist=min_dist, n_components=n_components,
                  metric=metric, title='{}/acq_plot{}'.format(read_dir, idx))


def read_hparam_rounds(write_dir, excel_store, rounds):
    df_store = []
    for excel_pair, round in zip(excel_store,rounds):
        # excel_pair is ann, dtr together
        df_ann = pd.read_excel(excel_pair[0])
        df_ann['model'] = 'ann'
        df_dtr=pd.read_excel(excel_pair[1])
        df_dtr['model'] = 'dtr'
        df = pd.concat([df_ann, df_dtr])
        df['round'] = round
        df_store.append(df)

    df = pd.concat(df_store)

    with open('{}/df.pkl'.format(write_dir), 'wb') as handle:
        pickle.dump(df, handle, protocol=pickle.HIGHEST_PROTOCOL)

def plot_hparam_rounds(write_dir, metrics):
    with open('{}/df.pkl'.format(write_dir), 'rb') as handle:
        df = pickle.load(handle)

    df = df.reset_index(drop=True)
    df = df.drop(df[df['round']==1].index)

    for metric in metrics:
        plt.close()
        sns.violinplot(x='round', y=metric, data=df, hue='model', split=True)
        plt.savefig('{}/violin_{}.png'.format(write_dir, metric))
        plt.close()
        sns.boxplot(x='round', y=metric, data=df, hue='model')
        plt.savefig('{}/box_{}.png'.format(write_dir, metric))

def plot_un_hparam_rounds(write_dir, excel_dir):
    whis = 1.5
    wb = openpyxl.Workbook()

    df = pd.read_excel(excel_dir)
    df = df.iloc[:,-(155*2*3):-(30*2*3)]
    new_values = np.array([np.mean(df.iloc[:,x:x+3].values,axis=1) for x in range(0,125*2*3,3)]).T
    se_df = pd.DataFrame(new_values[:,:125])
    wb.create_sheet('Un125 SE')
    ws = wb['Un125 SE']
    print_df_to_excel(df=se_df, ws=ws)
    se_df['round'] = [1, 2, 3, 4, 5, 6, '6e', 7, 8, 9, 10, 11, 12, 13]
    se_df = pd.melt(se_df, id_vars='round', value_name='Squared Error')
    plt.close()
    ax = sns.boxplot(x='round', y='Squared Error', data=se_df, showmeans=True, showfliers=True, whis=whis)
    ax.set_ylim([0,2000])
    plt.savefig('{}/un_se.png'.format(write_dir), bbox_inches = 'tight')
    se_df = se_df.drop(se_df[se_df['round']==1].index)
    plt.close()
    ax = sns.boxplot(x='round', y='Squared Error', data=se_df, showmeans=True, showfliers=True, whis=whis)
    ax.set_ylim([0,2000])
    plt.savefig('{}/un_se without round 1.png'.format(write_dir), bbox_inches = 'tight')

    he_df = pd.DataFrame(new_values[:, -125:])
    wb.create_sheet('Un125 HE')
    ws = wb['Un125 HE']
    print_df_to_excel(df=he_df, ws=ws)
    he_df['round'] = [1, 2, 3, 4, 5, 6, '6e', 7, 8, 9, 10, 11, 12, 13]
    he_df = pd.melt(he_df, id_vars='round', value_name='Haitao Error')

    plt.close()
    ax = sns.boxplot(x='round', y='Haitao Error', data=he_df, showmeans=True, showfliers=True, whis=whis)
    ax.set_ylim([0, 0.8])
    plt.savefig('{}/un_he.png'.format(write_dir), bbox_inches = 'tight')
    he_df = he_df.drop(he_df[he_df['round']==1].index)
    plt.close()
    ax = sns.boxplot(x='round', y='Haitao Error', data=he_df, showmeans=True, showfliers=True, whis=whis)
    ax.set_ylim([0, 0.8])
    plt.savefig('{}/un_he without round 1.png'.format(write_dir), bbox_inches = 'tight')


    # Plot for test
    df = pd.read_excel(excel_dir)
    df = df.iloc[:, -(30 * 2 * 3):]
    new_values = np.array([np.mean(df.iloc[:, x:x + 3].values, axis=1) for x in range(0, 30 * 2 * 3, 3)]).T
    se_df = pd.DataFrame(new_values[:, :30])
    wb.create_sheet('T30 SE')
    ws = wb['T30 SE']
    print_df_to_excel(df=se_df, ws=ws)
    se_df['round'] = [1, 2, 3, 4, 5, 6, '6e', 7, 8, 9, 10, 11, 12, 13]
    se_df = pd.melt(se_df, id_vars='round', value_name='Squared Error')
    plt.close()
    ax = sns.boxplot(x='round', y='Squared Error', data=se_df, showmeans=True, showfliers=True, whis=whis)
    ax.set_ylim([0, 1000])
    plt.savefig('{}/t_se.png'.format(write_dir), bbox_inches='tight')
    se_df = se_df.drop(se_df[se_df['round'] == 1].index)
    plt.close()
    ax = sns.boxplot(x='round', y='Squared Error', data=se_df, showmeans=True, showfliers=True, whis=whis)
    ax.set_ylim([0, 1000])
    plt.savefig('{}/t_se without round 1.png'.format(write_dir), bbox_inches='tight')

    he_df = pd.DataFrame(new_values[:, -30:])
    wb.create_sheet('T30 HE')
    ws = wb['T30 HE']
    print_df_to_excel(df=he_df, ws=ws)
    he_df['round'] = [1, 2, 3, 4, 5, 6, '6e', 7, 8, 9, 10, 11, 12, 13]
    he_df = pd.melt(he_df, id_vars='round', value_name='Haitao Error')

    plt.close()
    ax = sns.boxplot(x='round', y='Haitao Error', data=he_df, showmeans=True, showfliers=True, whis=whis)
    ax.set_ylim([0, 0.8])
    plt.savefig('{}/t_he.png'.format(write_dir), bbox_inches='tight')
    he_df = he_df.drop(he_df[he_df['round'] == 1].index)
    plt.close()
    ax = sns.boxplot(x='round', y='Haitao Error', data=he_df, showmeans=True, showfliers=True, whis=whis)
    ax.set_ylim([0, 0.8])
    plt.savefig('{}/t_he without round 1.png'.format(write_dir), bbox_inches='tight')

    wb.save('{}/rounds error data.xlsx'.format(write_dir))
    pass


def plot_var(excel_dir, combi_names, extra_idx=1):
    df = pd.read_excel(excel_dir, sheet_name='All')
    he_train = df.iloc[1::4, 1:2]
    he_val = df.iloc[1::4, 2:3]
    he_test = df.iloc[1::4, 3:4]

    def plot_single_he(sdf, name):
        sdf['Type'] = combi_names
        sdf = pd.melt(sdf, id_vars='Type', var_name='Dataset', value_name='HE')
        ax = sns.barplot(data=sdf, x='Type', y='HE')
        # ax.tick_params(axis='both', which='major', labelsize=6)
        plt.savefig('./Plots/HE {} barplot.png'.format(name), bbox_inches='tight')
        plt.close()

    plot_single_he(he_train, 'train')
    plot_single_he(he_val, 'val')
    plot_single_he(he_test, 'test')

    df = pd.read_excel(excel_dir, sheet_name='std')
    var = df.iloc[2::4, 5:]
    var.pop('125T')
    var['Type'] = combi_names[extra_idx:]
    var = pd.melt(var, id_vars='Type', var_name='Dataset', value_name='std.')
    ax=sns.barplot(data=var, x='Dataset', y='std.', hue='Type')
    ax.tick_params(axis='both', which='major', labelsize=6)
    plt.savefig('./Plots/std_barplots_all_dataset.png', bbox_inches='tight')
    plt.close()

    df = pd.read_excel(excel_dir, sheet_name='std percentage')
    var = df.iloc[2::4, 5:]
    var.pop('125T')
    var['Type'] = combi_names[1+extra_idx:]
    var = pd.melt(var, id_vars='Type', var_name='Dataset', value_name='std.')
    ax = sns.barplot(data=var, x='Dataset', y='std.', hue='Type')
    ax.tick_params(axis='both', which='major', labelsize=6)
    plt.savefig('./Plots/std_percentage diff barplots_all_dataset.png', bbox_inches='tight')
    plt.close()

    selected_idx = [6,9,11,16,18]
    df = pd.read_excel(excel_dir, sheet_name='std')
    var = df.iloc[2::4, selected_idx]
    var['Type'] = combi_names[extra_idx:]
    var = pd.melt(var, id_vars='Type', var_name='Dataset', value_name='std.')
    ax=sns.barplot(data=var, x='Dataset', y='std.', hue='Type')
    ax.tick_params(axis='both', which='major', labelsize=6)
    plt.savefig('./Plots/std_barplots_selected.png', bbox_inches='tight')
    plt.close()

    df = pd.read_excel(excel_dir, sheet_name='std percentage')
    var = df.iloc[2::4, selected_idx]
    var['Type'] = combi_names[1+extra_idx:]
    var = pd.melt(var, id_vars='Type', var_name='Dataset', value_name='std.')
    ax = sns.barplot(data=var, x='Dataset', y='std.', hue='Type')
    ax.tick_params(axis='both', which='major', labelsize=6)
    plt.savefig('./Plots/std_percentage diff barplots_selected.png', bbox_inches='tight')
    plt.close()



    pass

