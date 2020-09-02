import pandas as pd
import numpy as np
import openpyxl, pickle
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows
from own_package.others import print_array_to_excel, print_df_to_excel, create_excel_file
import os, itertools, pickle
from collections import deque
from deap import algorithms, base, creator, tools


def prepare_grand_data_store(dir_store):
    data_store = []
    for dir in dir_store:
        for filename in os.listdir(dir):
            if filename.endswith(".pkl"):
                with open('{}/{}'.format(dir, filename), 'rb') as handle:
                    data_store.extend(pickle.load(handle))
    return data_store


def get_mse_re(y, p_y):
    return np.mean((y - p_y) ** 2), np.mean(np.abs(y - p_y).T / y[:, -1].T)


def ga_train_val_eval_on_test(results_dir, data_store, hparams):
    # 9, 10 col is the sett ett df.
    # 11 is str but only for HE onwards, before that no str (true training) df
    # -3 is hparams
    # - 2 is unseen mse and he
    # -1 is unseen df


    trainset_ett_idx = -4
    for trial, data in enumerate(data_store):
        untrainset_df = data[10][trainset_ett_idx].copy(deep=True)
        ov_df = data[5]
        untrainset_df.iloc[:ov_df.shape[0], -3:] = ov_df.iloc[:, -3:]
        y = untrainset_df.iloc[:, :3].values
        p_y = untrainset_df.iloc[:, -3:].values
        mse = np.mean((y - p_y) ** 2)
        he = np.mean(np.abs(y - p_y).T / y[:, -1])
        data.append([mse, he])
        data.append([y, p_y])

    p_yt_store = np.array([x[4].iloc[:, -3:].values for x in data_store])
    yt = data_store[0][4].iloc[:, -6:-3].values
    p_yv_store = np.array([x[5].iloc[:, -3:].values for x in data_store])
    yv = data_store[0][5].iloc[:, -6:-3].values
    p_ytt_store = np.array([x[6].iloc[:, -3:].values for x in data_store])
    ytt = data_store[0][6].iloc[:, -6:-3].values
    p_yett_store = [np.array([x[10][idx].iloc[:, -3:].values for x in data_store]) for idx in
                    range(len(data_store[0][10]))]
    yett_store = [data_store[0][10][idx].iloc[:, -6:-3].values for idx in range(len(data_store[0][10]))]
    p_yuns_store = np.array([x[-1][-1] for x in data_store])
    yuns = data_store[0][-1][0]
    # p_y_names = [z for x in data_store for z in x[0][0]]
    p_y_names = [x[1][0] for x in data_store]
    total_models = len(p_y_names)
    creator.create("FitnessMax", base.Fitness, weights=(-1,))
    creator.create("Individual", list, fitness=creator.FitnessMax)

    def eval1(individual):
        selected_mask = [idx for idx, value in enumerate(individual) if value == 1]
        p_yt_selected_mean = np.mean(p_yt_store[selected_mask, :, :], axis=0)
        re_t = np.mean(np.abs(yt - p_yt_selected_mean).T / yt[:, -1].T)
        p_yv_selected_mean = np.mean(p_yv_store[selected_mask, :, :], axis=0)
        re_v = np.mean(np.abs(yv - p_yv_selected_mean).T / yv[:, -1].T)

        re = (re_t + re_v) / 2
        return (re,)

    def eval2(individual):
        selected_mask = [idx for idx, value in enumerate(individual) if value == 1]
        p_yt_selected_mean = np.mean(p_yt_store[selected_mask, :, :], axis=0)
        re_t = np.mean(np.abs(yt - p_yt_selected_mean).T / yt[:, -1].T)

        p_yv_selected_mean = np.mean(p_yv_store[selected_mask, :, :], axis=0)
        re_v = np.mean(np.abs(yv - p_yv_selected_mean).T / yv[:, -1].T)

        re = (re_t + 2 * re_v) / 3
        return (re,)

    def eval3(individual):
        selected_mask = [idx for idx, value in enumerate(individual) if value == 1]
        p_yt_selected_mean = np.mean(p_yt_store[selected_mask, :, :], axis=0)
        re_t = np.mean(np.abs(yt - p_yt_selected_mean).T / yt[:, -1].T)

        p_yv_selected_mean = np.mean(p_yv_store[selected_mask, :, :], axis=0)
        re_v = np.mean(np.abs(yv - p_yv_selected_mean).T / yv[:, -1].T)

        re = re_v
        return (re,)

    toolbox = base.Toolbox()
    toolbox.register("attr_bool", np.random.choice, np.arange(0, 2), p=hparams['init'])
    toolbox.register("individual", tools.initRepeat, creator.Individual, toolbox.attr_bool, n=total_models)
    toolbox.register("population", tools.initRepeat, list, toolbox.individual)
    if hparams['eval_func'] == 'eval1':
        toolbox.register("evaluate", eval1)
    elif hparams['eval_func'] == 'eval2':
        toolbox.register("evaluate", eval2)
    elif hparams['eval_func'] == 'eval3':
        toolbox.register("evaluate", eval3)
    else:
        raise KeyError('eval_func {} is not valid.'.format(hparams['eval_func']))
    toolbox.register("mate", tools.cxTwoPoint)
    toolbox.register("mutate", tools.mutFlipBit, indpb=0.2)
    toolbox.register("select", tools.selTournament, tournsize=3)
    # Logging
    stats = tools.Statistics(key=lambda ind: ind.fitness.values)
    stats.register("avg", np.mean, axis=0)
    stats.register("std", np.std, axis=0)
    stats.register("min", np.min, axis=0)
    stats.register("max", np.max, axis=0)
    pop = toolbox.population(n=hparams['n_pop'])
    hof = tools.HallOfFame(1)
    pop, logbook = algorithms.eaSimple(toolbox=toolbox, population=pop,
                                       cxpb=0.5, mutpb=0.2,
                                       ngen=hparams['n_gen'], halloffame=hof, stats=stats,
                                       verbose=True)

    # Plotting
    gen = logbook.select("gen")
    fit_min = [x.item() for x in logbook.select("min")]
    fit_avg = [x.item() for x in logbook.select("avg")]
    fit_max = [x.item() for x in logbook.select("max")]

    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, fit_min, label="Min MRE")
    line2 = ax1.plot(gen, fit_avg, label="Avg MRE")
    line3 = ax1.plot(gen, fit_max, label="Max MRE")
    plt.legend()
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Relative Error")
    plt.savefig('{}/plots/GA_opt_MRE_all.png'.format(results_dir), bbox_inches="tight")

    fig, ax1 = plt.subplots()
    line1 = ax1.plot(gen, fit_min, label="Min MRE")
    plt.legend()
    ax1.set_xlabel("Generation")
    ax1.set_ylabel("Total Generation Cost")
    plt.savefig('{}/plots/GA_opt_min_only.png'.format(results_dir), bbox_inches="tight")

    # Printing to excel
    excel_name = results_dir + '/results.xlsx'
    wb = openpyxl.Workbook()
    sheetname = wb.sheetnames[-1]
    ws = wb[sheetname]

    # Writing other subset split, instance per run, and bounds
    print_array_to_excel(['n_gen', 'n_pop'], (1, 1), ws, axis=1)
    print_array_to_excel([hparams['n_gen'], hparams['n_pop']], (2, 1), ws, axis=1)
    row = 2
    ws.cell(row + 1, 1).value = 'Best Allocation Value'
    ws.cell(row + 1, 2).value = hof[-1].fitness.values[-1]

    wb.create_sheet('av')
    ws = wb['av']
    ws.cell(1, 1).value = 'Names'
    ws.cell(1, 2).value = 'av'
    print_array_to_excel(p_y_names, (2, 1), ws=ws, axis=0)
    print_array_to_excel(list(hof[-1]), (2, 2), ws=ws, axis=0)

    selected_mask = [idx for idx, value in enumerate(list(hof[-1])) if value == 1]
    p_yt_selected_mean = np.mean(p_yt_store[selected_mask, :, :], axis=0)
    p_yv_selected_mean = np.mean(p_yv_store[selected_mask, :, :], axis=0)
    p_ytt_selected_mean = np.mean(p_ytt_store[selected_mask, :, :], axis=0)
    unseen_missing = False
    try:
        p_yuns_selected_mean = np.mean(p_yuns_store[selected_mask, :, :], axis=0)
        ett_names = ['I01-1', 'I01-2', 'I01-3',
                     'I05-1', 'I05-2', 'I05-3',
                     'I10-1', 'I10-2', 'I10-3',
                     'I30-1', 'I30-2', 'I30-3',
                     'I50-1', 'I50-2', 'I50-3',
                     '125Test', '125Test I01', '125Test I05', '125Test I10']
    except IndexError:
        unseen_missing = True
        ett_names = ['I01-1', 'I01-2', 'I01-3',
                     'I05-1', 'I05-2', 'I05-3',
                     'I10-1', 'I10-2', 'I10-3',
                     'I30-1', 'I30-2', 'I30-3',
                     'I50-1', 'I50-2', 'I50-3',]
    p_yett_store_selected_mean = [np.mean(x[selected_mask, :, :], axis=0) for x in p_yett_store]
    mse_t, re_t = get_mse_re(yt, p_yt_selected_mean)
    mse_v, re_v = get_mse_re(yv, p_yv_selected_mean)
    mse_tt, re_tt = get_mse_re(ytt, p_ytt_selected_mean)
    mse_re_ett_store = [get_mse_re(yett, p_yett) for yett, p_yett in zip(yett_store, p_yett_store_selected_mean)]
    var_ett = []
    if unseen_missing:
        idx_store = [1, 1, 1, 5, 5, 5, 10, 10, 10, 30, 30, 30, 50, 50, 50]
    else:
        idx_store = [1, 1, 1, 5, 5, 5, 10, 10, 10, 30, 30, 30, 50, 50, 50, 0, 1, 5, 10]
        mse_uns, re_uns = get_mse_re(yuns, p_yuns_selected_mean)
    for idx, (invariant, p_y) in enumerate(
            zip(idx_store, p_yett_store_selected_mean)):
        if invariant == 0:
            var_ett.append(0)
        else:

            if idx < 15:
                base_numel = 30
            else:
                base_numel = 125
            var_ett.append(
                np.mean(
                    [np.std(
                        np.concatenate((p_y[i:i+1, :],
                                        p_y[base_numel + invariant * i:base_numel + invariant * i + invariant, :]), axis=0)
                    , axis=0)
                        for i in range(base_numel)])
            )
            #i = 5
            #print('invariant {} idx {} shape {}'.format(invariant, idx, np.concatenate((p_y[i:i+1, :],
            #                            p_y[base_numel + invariant * i:base_numel + invariant * i + invariant, :]), axis=0).shape))




    def print_results(name, y, p_y, mse, re):
        nonlocal wb, ws
        wb.create_sheet(name)
        ws = wb[name]
        df = pd.DataFrame(np.concatenate((y, p_y), axis=1), columns=['y1', 'y2', 'y3', 'P_y1', 'P_y2', 'P_y3'])
        print_df_to_excel(df=df, ws=ws)
        start_col = len(df.columns) + 3
        ws.cell(1, start_col).value = 'MSE'
        ws.cell(2, start_col).value = 'HE'
        ws.cell(1, start_col + 1).value = mse
        ws.cell(2, start_col + 1).value = re

    print_results('Training', yt, p_yt_selected_mean, mse_t, re_t)
    print_results('Val', yv, p_yv_selected_mean, mse_v, re_v)
    print_results('Test', ytt, p_ytt_selected_mean, mse_tt, re_tt)
    if not unseen_missing:
        print_results('Unseen', yuns, p_yuns_selected_mean, mse_uns, re_uns)
        df = pd.DataFrame(data=[[mse_t, mse_v, mse_tt, mse_uns] + [x[0] for x in mse_re_ett_store],
                                [re_t, re_v, re_tt, re_uns] + [x[1] for x in mse_re_ett_store],
                                [0, 0, 0, 0] + var_ett],
                          columns=['Training', 'Val', 'Test', 'Unseen'] + ett_names,
                          index=['MSE', 'HE', 'Var'])
    else:
        df = pd.DataFrame(data=[[mse_t, mse_v, mse_tt] + [x[0] for x in mse_re_ett_store],
                                [re_t, re_v, re_tt] + [x[1] for x in mse_re_ett_store],
                                [0, 0, 0, 0] + var_ett],
                          columns=['Training', 'Val', 'Test', 'Unseen'] + ett_names,
                          index=['MSE', 'HE', 'Var'])
    [print_results(name, yett_store[idx], p_yett_store_selected_mean[idx], mse_re[0], mse_re[1]) for name, idx, mse_re
     in zip(ett_names, range(len(data_store[0][10])), mse_re_ett_store)]

    ws = wb[sheetname]

    print_df_to_excel(df=df, ws=ws, start_row=5)

    wb.save(excel_name)
