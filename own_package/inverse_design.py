import numpy as np
import pandas as pd
import openpyxl, time
from skopt import gp_minimize, gbrt_minimize, forest_minimize, dummy_minimize
from skopt.space import Real, Integer, Categorical
from skopt.utils import use_named_args
from skopt.plots import plot_convergence
from own_package.pso_ga import pso_ga
from own_package.features_labels_setup import load_data_to_fl, load_testset_to_fl
from own_package.others import print_array_to_excel, create_results_directory, print_df_to_excel, create_excel_file
from own_package.active_learning.acquisition import load_svm_ensemble, svm_ensemble_prediction, load_model_ensemble, \
    model_ensemble_prediction


def inverse_design(targets, loss_func, bounds, int_idx, init_guess, model_directory_store, svm_directory, loader_file, write_dir,
                   opt_mode):
    model_store = []
    for model_directory in model_directory_store:
        model_store.extend(load_model_ensemble(model_directory))
    svm_store = load_svm_ensemble(svm_directory)
    fl = load_data_to_fl(loader_file, norm_mask=[0, 1, 3, 4, 5], normalise_labels=False, label_type='cutoff')

    data_store = []
    if opt_mode == 'psoga':
        def fitness(params):
            nonlocal data_store
            features = np.array(params)
            x = features[0]
            y = features[1]
            if x + y > 1:
                u = -y + 1
                v = -x + 1
                features[0:2] = np.array([u, v])

            # SVM Check
            p_class, distance = svm_ensemble_prediction(svm_store, features[0:2])

            if distance.item() < 0:
                # Distance should be negative value when SVM assigns class 0. Hence a_score will be negative.
                # The more negative the a_score is, the further the composition is from the hyperplane,
                # hence, the less likely the optimizer will select examples with class 0.
                mse = 10e5 * distance.item()
                prediction_mean = [-1] * fl.labels_dim
                prediction_std = [-1] * fl.labels_dim
                disagreement = -1
            elif features[0] + features[1] > 1:
                # Distance should be negative value when SVM assigns class 0. Hence a_score will be negative.
                # The more negative the a_score is, the further the composition is from the hyperplane,
                # hence, the less likely the optimizer will select examples with class 0.
                mse = 10e5 * (1 - (features[0] + features[1]))
                prediction_mean = [-1] * fl.labels_dim
                prediction_std = [-1] * fl.labels_dim
                disagreement = -1
            else:
                features_c = features[:-1]
                onehot = features[-1].item()
                if onehot == 0:
                    features_in = np.concatenate((features_c, np.array([1, 0, 0])))
                elif onehot == 1:
                    features_in = np.concatenate((features_c, np.array([0, 1, 0])))
                elif onehot == 2:
                    features_in = np.concatenate((features_c, np.array([0, 0, 1])))
                features_input_norm = fl.apply_scaling(features_in)
                prediction_mean, prediction_std = model_ensemble_prediction(model_store, features_input_norm)
                mse = -loss_func(targets, prediction_mean)
                disagreement = np.mean(prediction_std)
                prediction_mean = prediction_mean.tolist()
                prediction_std = prediction_std.tolist()

            data = list(features) + [-mse, disagreement] + prediction_mean + prediction_std
            data_store.append(data)
            return (-mse,)

        pmin = [x[0] for x in bounds]
        pmax = [x[1] for x in bounds]

        smin = [abs(x - y) * 0.001 for x, y in zip(pmin, pmax)]
        smax = [abs(x - y) * 0.5 for x, y in zip(pmin, pmax)]

        pso_params = {'c1': 1.5, 'c2': 1.5, 'wmin': 0.4, 'wmax': 0.9,
                      'ga_iter_min': 2, 'ga_iter_max': 10, 'iter_gamma': 10,
                      'ga_num_min': 5, 'ga_num_max': 20, 'num_beta': 15,
                      'tourn_size': 3, 'cxpd': 0.9, 'mutpd': 0.05, 'indpd': 0.5, 'eta': 0.5,
                      'pso_iter': 10, 'swarm_size': 300}

        pso_ga(func=fitness, pmin=pmin, pmax=pmax,
               smin=smin, smax=smax,
               int_idx=[3], params=pso_params, ga=True, initial_guess=init_guess)

    elif opt_mode == 'forest' or opt_mode == 'dummy':
        space = [Real(low=bounds[0][0], high=bounds[0][1], name='CNT'),
                 Real(low=bounds[1][0], high=bounds[1][1], name='PVA'),
                 Real(low=bounds[2][0], high=bounds[2][1], name='Thickness'),
                 Categorical(categories=[0, 1, 2], name='Dimension')]

        iter_count = 0
        start = time.time()
        end = 0
        @use_named_args(space)
        def fitness(**params):
            nonlocal data_store, iter_count, start, end
            iter_count +=1
            features = np.array([x for x in params.values()])
            x = features[0]
            y = features[1]
            if x + y > 1:
                u = -y + 1
                v = -x + 1
                features[0:2] = np.array([u, v])
            # SVM Check
            p_class, distance = svm_ensemble_prediction(svm_store, features[0:2])
            if distance.item() < 0:
                # Distance should be negative value when SVM assigns class 0. Hence a_score will be negative.
                # The more negative the a_score is, the further the composition is from the hyperplane,
                # hence, the less likely the optimizer will select examples with class 0.
                mse = 10e5 * distance.item()
                prediction_mean = [-1] * fl.labels_dim
                prediction_std = [-1] * fl.labels_dim
                disagreement = -1
            elif features[0] + features[1] > 1:
                # Sum of composition needs to be less than 1
                mse = 10e5 * (1 - (features[0] + features[1]))
                prediction_mean = [-1] * fl.labels_dim
                prediction_std = [-1] * fl.labels_dim
                disagreement = -1
            else:
                features_c = features[:-1]
                onehot = features[-1].item()
                if onehot == 0:
                    features_in = np.concatenate((features_c, np.array([1, 0, 0])))
                elif onehot == 1:
                    features_in = np.concatenate((features_c, np.array([0, 1, 0])))
                elif onehot == 2:
                    features_in = np.concatenate((features_c, np.array([0, 0, 1])))
                features_input_norm = fl.apply_scaling(features_in)
                prediction_mean, prediction_std = model_ensemble_prediction(model_store, features_input_norm)
                mse = -loss_func(targets, prediction_mean)  # Some negative number
                disagreement = np.mean(prediction_std)
                prediction_mean = prediction_mean.tolist()
                prediction_std = prediction_std.tolist()

            data = list(features) + [-mse, disagreement] + prediction_mean + prediction_std
            data_store.append(data)
            if iter_count % 10 == 0:
                end = time.time()
                print('Current Iteration {}. Time taken for past 10 evals: {}. '.format(iter_count, end-start))
                start = time.time()
            return -mse  # Make negative become positive, and minimizing score towards 0.

        if opt_mode == 'forest':
            forest_minimize(func=fitness,
                            dimensions=space,
                            acq_func='EI',  # Expected Improvement.
                            n_calls=1000,
                            verbose=False)
        else:
            dummy_minimize(func=fitness,
                            dimensions=space,
                            n_calls=5000,
                            verbose=False)

    p_mean_name = np.array(['Pmean_' + str(x) for x in list(map(str, np.arange(1, 4)))])
    p_std_name = np.array(['Pstd_' + str(x) for x in list(map(str, np.arange(1, 4)))])

    columns = np.concatenate((np.array(fl.features_c_names[:-2]),
                              np.array(['mse']),
                              np.array(['Disagreement']),
                              p_mean_name,
                              p_std_name
                              ))

    iter_df = pd.DataFrame(data=data_store,
                           columns=columns)

    iter_df = iter_df.sort_values(by=['mse'], ascending=True)

    excel_dir = create_excel_file('{}/inverse_design_{}_{}.xlsx'.format(write_dir, opt_mode, targets))
    wb = openpyxl.load_workbook(excel_dir)
    ws = wb[wb.sheetnames[-1]]  # Taking the ws name from the back ensures that if SNN1 is the new ws, it works
    ws.cell(1, 1).value = 'Target'
    print_array_to_excel(array=targets, first_cell=(1, 2), axis=1, ws=ws)
    print_df_to_excel(df=iter_df, ws=ws, start_row=3)

    wb.save(excel_dir)
    wb.close()


def eval_model_on_fl(model, fl, return_df=True, label_scaler=None):
    p_y = model.predict(fl.features_c_norm)
    if label_scaler:
        p_y = label_scaler(p_y)
    for row, p_label in enumerate(p_y.tolist()):
        if p_label[1] > p_label[2]:
            p_y[row, 1] = p_y[row, 2]
        if p_label[0] > p_y[row, 1]:
            p_y[row, 0] = p_y[row, 1]
    se_store = (fl.labels - p_y) ** 2
    re_store = (np.abs(fl.labels - p_y).T / fl.labels[:, -1]).T
    if return_df:
        df = pd.DataFrame(data=np.concatenate((fl.labels, p_y), axis=1),
                          index=list(range(1, 1 + fl.count)),
                          columns=list(fl.labels_names)
                                  + ['P_{}'.format(col) for col in fl.labels_names])
        return p_y, df, np.mean(se_store), np.mean(re_store)
    else:
        return p_y, np.mean(se_store), np.mean(re_store)


def eval_models(model_directory_store, results_dir):
    model_store = []
    for model_directory in model_directory_store:
        model_store.extend(load_model_ensemble(model_directory))

    test_excel_dir = './excel/ett_30testset_cut.xlsx'
    ett_names = ['I01-1', 'I01-2', 'I01-3',
                 'I05-1', 'I05-2', 'I05-3',
                 'I10-1', 'I10-2', 'I10-3',
                 'I30-1', 'I30-2', 'I30-3',
                 'I50-1', 'I50-2', 'I50-3',
                 '125Test', '125Test I01', '125Test I05', '125Test I10']
    ett_store = ['./excel/ett_30testset_cut Invariant 1.xlsx',
                 './excel/ett_30testset_cut Invariant 1 - 2.xlsx',
                 './excel/ett_30testset_cut Invariant 1 - 3.xlsx',
                 './excel/ett_30testset_cut Invariant 5.xlsx',
                 './excel/ett_30testset_cut Invariant 5 - 2.xlsx',
                 './excel/ett_30testset_cut Invariant 5 - 3.xlsx',
                 './excel/ett_30testset_cut Invariant 10.xlsx',
                 './excel/ett_30testset_cut Invariant 10 - 2.xlsx',
                 './excel/ett_30testset_cut Invariant 10 - 3.xlsx',
                 './excel/ett_30testset_cut Invariant 30.xlsx',
                 './excel/ett_30testset_cut Invariant 30 - 2.xlsx',
                 './excel/ett_30testset_cut Invariant 30 - 3.xlsx',
                 './excel/ett_30testset_cut Invariant 50.xlsx',
                 './excel/ett_30testset_cut Invariant 50 - 2.xlsx',
                 './excel/ett_30testset_cut Invariant 50 - 3.xlsx',
                 './excel/ett_125trainset_cut.xlsx',
                 './excel/ett_125trainset_cut Invariant 1.xlsx',
                 './excel/ett_125trainset_cut Invariant 5.xlsx',
                 './excel/ett_125trainset_cut Invariant 10.xlsx']

    fl = load_data_to_fl('./excel/Data_loader_spline_full_onehot_R13_cut_CM3.xlsx',
                         label_type='cutoff',
                         normalise_labels=False,
                         norm_mask=[0, 1, 3, 4, 5])
    test_fl = load_testset_to_fl(test_excel_dir, scaler=fl.scaler, norm_mask=[0, 1, 3, 4, 5])
    ett_fl_store = [load_testset_to_fl(x, scaler=fl.scaler, norm_mask=[0, 1, 3, 4, 5]) for x in ett_store]

    ytt = test_fl.labels
    yett_store = [ett_fl.labels for ett_fl in ett_fl_store]

    stt_p_y_store = []
    stt_df_store = []
    stt_mse_store = []
    stt_mre_store = []

    sett_p_y_store = []
    sett_df_store = []
    sett_mse_store = []
    sett_mre_store = []

    for model in model_store:
        stt_p_y, stt_df, stt_mse, stt_mre = eval_model_on_fl(model, test_fl, return_df=True)
        stt_p_y_store.append(stt_p_y)
        stt_df_store.append(stt_df)
        stt_mse_store.append(stt_mse)
        stt_mre_store.append(stt_mre)

        p_y_store = []
        df_store = []
        mse_store = []
        mre_store = []

        for ett_fl in ett_fl_store:
            p_y, df, mse, mre = eval_model_on_fl(model, ett_fl, return_df=True)
            p_y_store.append(p_y)
            df_store.append(df)
            mse_store.append(mse)
            mre_store.append(mre)

        sett_p_y_store.append(p_y_store)
        sett_df_store.append(df_store)
        sett_mse_store.append(mse_store)
        sett_mre_store.append(mre_store)

    p_ytt_selected_mean = np.mean(np.array(stt_p_y_store), axis=0)
    p_yett_store_selected_mean = [np.mean(np.array(p_yett), axis=0) for p_yett in
                                  [list(x) for x in zip(*sett_p_y_store)]]

    def get_mse_re(y, p_y):
        return np.mean((y - p_y) ** 2), np.mean(np.abs(y - p_y).T / y[:, -1].T)

    mse_tt, re_tt = get_mse_re(ytt, p_ytt_selected_mean)
    mse_re_ett_store = [get_mse_re(yett, p_yett) for yett, p_yett in zip(yett_store, p_yett_store_selected_mean)]

    var_ett = []
    idx_store = [1, 1, 1, 5, 5, 5, 10, 10, 10, 30, 30, 30, 50, 50, 50, 0, 1, 5, 10]

    for idx, (invariant, p_y) in enumerate(
            zip(idx_store, p_yett_store_selected_mean)):
        if invariant == 0:
            var_ett.append(0)
        else:
            if idx < 15:
                base_numel = 30
            else:
                base_numel = 125
            var_ett.append(
                np.mean(
                    [np.std(
                        np.concatenate((p_y[i:i + 1, :],
                                        p_y[base_numel + invariant * i:base_numel + invariant * i + invariant, :]),
                                       axis=0)
                        , axis=0)
                        for i in range(base_numel)])
            )

    # Printing to excel
    excel_name = results_dir + '/results.xlsx'
    wb = openpyxl.Workbook()
    wb.create_sheet('main')

    def print_results(name, y, p_y, mse, re):
        nonlocal wb
        wb.create_sheet(name)
        ws = wb[name]
        df = pd.DataFrame(np.concatenate((y, p_y), axis=1), columns=['y1', 'y2', 'y3', 'P_y1', 'P_y2', 'P_y3'])
        print_df_to_excel(df=df, ws=ws)
        start_col = len(df.columns) + 3
        ws.cell(1, start_col).value = 'MSE'
        ws.cell(2, start_col).value = 'HE'
        ws.cell(1, start_col + 1).value = mse
        ws.cell(2, start_col + 1).value = re

    print_results('Test', ytt, p_ytt_selected_mean, mse_tt, re_tt)
    [print_results(name, yett_store[idx], p_yett_store_selected_mean[idx], mse_re[0], mse_re[1]) for name, idx, mse_re
     in zip(ett_names, range(len(yett_store)), mse_re_ett_store)]

    df = pd.DataFrame(data=[[mse_tt] + [x[0] for x in mse_re_ett_store],
                            [re_tt] + [x[1] for x in mse_re_ett_store],
                            [0] + var_ett],
                      columns=['Test'] + ett_names,
                      index=['MSE', 'HE', 'Var'])

    print_df_to_excel(df=df, ws=wb['main'], start_row=5)
    wb.save(excel_name)
