import numpy as np
import pandas as pd
import openpyxl
import os, pickle
from own_package.others import print_df_to_excel


def get_best_trial_from_rounds(dir_store, excel_subname, sort_col, results_excel_dir):
    top_trials_store = []
    for dir in dir_store:
        for filename in os.listdir(dir):
            if filename.__contains__(excel_subname):
                print('Importing data from {} in dir {}'.format(filename, dir))
                df = pd.read_excel('{}/{}'.format(dir, filename), index_col=None)
                df.sort_values(by=sort_col, ascending=True, inplace=True)
                top_trials_store.append(df.iloc[0, :].values)

    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[-1]]
    df = pd.DataFrame(data=top_trials_store, columns=df.columns)
    print_df_to_excel(df=df, ws=ws)
    wb.save(results_excel_dir)


def combine_model(data_store, idx_store, return_test_p_y=False):
    '''
    NOTE idx must be the idx of the data_store nested list, not the trial number since data_store ordering is not the
    same as the trial number ordering.
    :param data_store: [data_for_ann, data_for_dtr]
    :param idx_store: [top_idx_for_ann, top_idx_for_dtr]
    :return:
    '''
    output = ['-']
    name_store = [[x[0][0][0].rpartition('_')[0] for x in data] for data in data_store]
    selected_names = [np.array(name)[idx].tolist() for idx, name in zip(idx_store, name_store)]
    output.append(str(selected_names))  # Names of models combined

    def get_mse_mre(df_idx, return_values=False):
        # 4,5,6 = train, val, test df
        p_y = np.mean(
            np.array(
                [data_store[pair][i][df_idx].iloc[:, -3:].values for pair in range(len(data_store)) for i in idx_store[pair]]),
            axis=0)
        y = data_store[0][0][df_idx].iloc[:, -6:-3].values
        if return_values:
            return np.mean((p_y - y) ** 2), np.mean(np.abs(p_y - y).T / y[:, -1]), p_y, y
        else:
            return np.mean((p_y - y) ** 2), np.mean(np.abs(p_y - y).T / y[:, -1])

    # Train set
    mse, mre = get_mse_mre(4)
    output.append(mse)
    output.append(mre)
    # Val set
    mse, mre = get_mse_mre(5)
    output.append(mse)
    output.append(mre)
    # Test set
    mse, mre, p_y, y = get_mse_mre(6, return_values=True)
    test_p_y = p_y.copy()
    output.append(mse)
    output.append(mre)
    t_se = (y - p_y) ** 2
    t_re = (np.abs(y - p_y).T / y[:, -1]).T
    # Unseen train set
    unseen_idx = 15

    def get_p_y_for_model_idx_unseen_trainset(model_idx, pair):
        untrainset_df = data_store[pair][model_idx][10][unseen_idx].copy(deep=True)
        ov_df = data_store[pair][model_idx][5]
        untrainset_df.iloc[:ov_df.shape[0], -3:] = ov_df.iloc[:, -3:]
        return untrainset_df.iloc[:, -3:].values

    p_y = np.mean(
        np.array([get_p_y_for_model_idx_unseen_trainset(i, pair) for pair in range(len(data_store)) for i in idx_store[pair]]),
        axis=0)
    y = data_store[0][0][10][unseen_idx].iloc[:, :3].values
    un_se = (y - p_y) ** 2
    un_re = (np.abs(y - p_y).T / y[:, -1]).T
    mse = np.mean(un_se)
    mre = np.mean(un_re)
    output.append(mse)
    output.append(mre)
    # Extra Test Sets
    mse_store = []
    mre_store = []
    p_y_ett_store = []
    for ett_idx in range(19):
        p_y = np.mean(np.array(
            [data_store[pair][i][10][ett_idx].iloc[:, -3:].values for pair in range(len(data_store)) for i in
             idx_store[pair]]),
            axis=0)
        p_y_ett_store.append(p_y)
        y = data_store[0][0][10][ett_idx].iloc[:, :3].values
        mse = np.mean((y - p_y) ** 2)
        mre = np.mean(np.abs(y - p_y).T / y[:, -1])
        mse_store.append(mse)
        mre_store.append(mre)
    output.extend(mse_store)
    output.extend(mre_store)

    var_ett = []
    for idx, (invariant, p_y) in enumerate(
            zip([1, 1, 1, 5, 5, 5, 10, 10, 10, 30, 30, 30, 50, 50, 50, 0, 1, 5, 10], p_y_ett_store)):
        if invariant == 0:
            var_ett.append(0)
        else:

            if idx < 15:
                base_numel = 30
            else:
                base_numel = 125
            var_ett.append(
                np.mean(
                    [np.std(
                        np.concatenate((p_y[i:i + 1, :],
                                        p_y[base_numel + invariant * i:base_numel + invariant * i + invariant, :]),
                                       axis=0)
                        , axis=0)
                        for i in range(base_numel)])
            )

    hparams = [[], []]
    for hp, data, top_idx in zip(hparams, data_store, idx_store):
        hp.extend([x[-1] for x in [data[idx] for idx in top_idx]])
    output.extend([str(x) for x in hparams])
    if return_test_p_y:
        return [output, un_se, un_re, t_se, t_re, var_ett, test_p_y]
    else:
        return [output, un_se, un_re, t_se, t_re, var_ett]


def get_best_trial_from_rounds_custom_metric(dir_store, excel_subname, metric_cols, weightage, results_excel_dir,
                                             top_models=1):
    '''

    :param dir_store: directory where data_store.pkl and processed excel results are kept in
    :param excel_subname: name of the excel to base the model selection on.
    Either overall_summary (for 10 fold aggregated results) or solo_summary (for individual fold results)
    :param metric_cols: metric to base comparison on. Can be a list of metric, if so, need to give the weightage list
    :param weightage: list of weightage to multiply the metrics by
    :param results_excel_dir: excel name to print results in
    :param top_models: How many top models to average over for each round.
    :return:
    '''
    top_trials_store = []
    test_p_y_store = []

    for round, dir_pair in enumerate(dir_store):
        top_idx_pair = []
        data_store_pairs = []
        for pair_idx, dir in enumerate(dir_pair):
            for filename in os.listdir(dir):
                if filename == 'overall_summary.xlsx':
                    print('Importing data from {} in dir {}'.format(filename, dir))
                    df = pd.read_excel('{}/{}'.format(dir, filename), index_col=None)
                    column_names = df.columns.tolist()
                    metric_store = np.array([df[metric].values for metric in metric_cols])
                    new_metric = np.sum(metric_store.T * np.array(weightage), axis=1)
                    if top_models == 1:
                        idx = np.argmin(new_metric)
                        top_trials_store.append(df.iloc[idx, :].values.tolist() + [new_metric[idx]])
                    else:
                        data_store = []
                        for filename in os.listdir(dir):
                            if filename.endswith(".pkl"):
                                with open('{}/{}'.format(dir, filename), 'rb') as handle:
                                    data_store.extend(pickle.load(handle))
                        name_store = [x[0][0][0].rpartition('_')[0] for x in data_store]
                        df['new'] = new_metric
                        df.sort_values(by='new', inplace=True)
                        top_rows = list(df.iloc[:top_models]['name'])
                        top_model_idx_store = []
                        if round == 11:
                            top_model_idx_store.extend([0, 1, 2])
                        else:
                            pass
                            #for row in top_rows:
                                #idx = '_{}_1'.format(row.split('_')[-1])
                                #top_model_idx_store.extend([i for i, x in enumerate(name_store) if x.find(idx) > -1])
                        top_model_idx_store = [idx for row in top_rows for idx, s in enumerate(name_store) if row == s]
                        top_idx_pair.append(top_model_idx_store)
                        data_store_pairs.append(data_store)

        [combined_row, un_se, un_re, t_se, t_re, var_ett, test_p_y] = combine_model(data_store=data_store_pairs,
                                                                          idx_store=top_idx_pair, return_test_p_y=True)
        test_p_y_store.append(test_p_y)
        combined_df = pd.DataFrame(np.array(combined_row)[None, :], columns=column_names)
        metric_store = np.array([float(combined_df[metric].values) for metric in metric_cols])
        top_trials_store.append(combined_row + [np.sum(metric_store.T * np.array(
            weightage))] + var_ett + un_se.flatten().tolist() + un_re.flatten().tolist() +
                                t_se.flatten().tolist() + t_re.flatten().tolist() )

    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[-1]]
    ett_names = ['I01-1', 'I01-2', 'I01-3',
                 'I05-1', 'I05-2', 'I05-3',
                 'I10-1', 'I10-2', 'I10-3',
                 'I30-1', 'I30-2', 'I30-3',
                 'I50-1', 'I50-2', 'I50-3',
                 '125Test', '125Test I01', '125Test I05', '125Test I10']
    df = pd.DataFrame(data=top_trials_store,
                      columns=column_names + ['New Metric'] +
                              ['std {}'.format(x) for x in ett_names]+
                              ['UN SE{}_{}'.format(y + 1, x + 1) for x in range(125) for y in range(3)] +
                              ['UN RE{}_{}'.format(y + 1, x + 1) for x in range(125) for y in range(3)] +
                              ['T SE{}_{}'.format(y + 1, x + 1) for x in range(30) for y in range(3)] +
                              ['T RE{}_{}'.format(y + 1, x + 1) for x in range(30) for y in range(3)])
    print_df_to_excel(df=df, ws=ws)
    for idx, p_y in enumerate(test_p_y_store):
        wb.create_sheet(str(idx))
        ws = wb[str(idx)]
        print_df_to_excel(pd.DataFrame(test_p_y), ws=ws)


    wb.save(results_excel_dir)
