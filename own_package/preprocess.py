import math
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import os
import glob
import pickle
from collections import defaultdict
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error
from scipy.optimize import curve_fit
from scipy.integrate import simps, trapz
from scipy.interpolate import CubicSpline, UnivariateSpline, PchipInterpolator, interp1d
from own_package.features_labels_setup import Features_labels_grid


def round_sigfigs(num, sig_figs):
    """Round to specified number of sigfigs.

    >>> round_sigfigs(0, sig_figs=4)
    0
    >>> int(round_sigfigs(12345, sig_figs=2))
    12000
    >>> int(round_sigfigs(-12345, sig_figs=2))
    -12000
    >>> int(round_sigfigs(1, sig_figs=2))
    1
    >>> '{0:.3}'.format(round_sigfigs(3.1415, sig_figs=2))
    '3.1'
    >>> '{0:.3}'.format(round_sigfigs(-3.1415, sig_figs=2))
    '-3.1'
    >>> '{0:.5}'.format(round_sigfigs(0.00098765, sig_figs=2))
    '0.00099'
    >>> '{0:.6}'.format(round_sigfigs(0.00098765, sig_figs=3))
    '0.000988'
    """
    if num != 0:
        return round(num, -int(math.floor(math.log10(abs(num))) - (sig_figs - 1)))
    else:
        return 0  # Can't take the log of 0


def read_excel_data(read_excel_file, write_excel_file, plot_directory, mode, cp0=4, poly=2, normalise_r=False):
    """
    :param mode:
    1) Line mode
    To convert excel raw data which contains data in the form of 2 headers
    Exp     1               2               3
    Data    Strain  R       Strain  R       Strain  R

    Prints summarised metrics into write_excel_file with the following columns:
    ['m1', 'r1', 'm2', 'c2', 'r2', 'x_cp', 'y_cp', 'combined_r2', 'cp']
    m1: slope of first line
    r1: R squared of first line
    m2: slope of second line
    c2: intercept of second line (there is no c1 since c1 is fixed to be 0)
    r2: R squared of second line
    x_cp: x intercept of the 2 best fit lines
    y_cp: y intercept of the 2 best fit lines
    combined_r2: R squared of the first and second line added together
    cp: Crossing point which is the 1st data point of the second line counting from 0

    Also plots the scatter plot and the 2 best fit line and saves the plot in the plot_directory

    :param cp0: cp0 is the initial data point index where the dataset will be split into 2 dataset for line mode.
    eg: cp0 = k ==> [0, 1, ... N] split to [0, 1, ..., k-1] and [k, k+1, ..., N]

    2) Quad mode
    Performs best fit regression based on the following functional form
    y = u(t-c)*(a(x-c)+b(x-c)^2)
    u = heaviside function
    a = scalar multiple for linear term
    b = scalar multiple for quadratic term
    c = delay term


    :param read_excel_file: name of excel file. Remember to add .xlsx at the end

    :return: Nothing
    """
    # read_excel_file part
    df = pd.read_excel(read_excel_file, sheet_name='raw', header=[0, 1], index_col=0)

    strain = df.xs('Strain (%)', level='Data', axis=1)
    strain = strain.values.T.tolist()
    strain_store = []
    for single_exp in strain:
        strain_single = [x for x in single_exp if not np.isnan(x)]
        strain_store.append(strain_single)
    strain_store = [np.array(x) for x in strain_store]

    r = df.xs('R', level='Data', axis=1)
    r = r.values.T.tolist()
    r_store = []
    for single_exp in r:
        r_single = [x for x in single_exp if not np.isnan(x)]
        r_store.append(r_single)
    if normalise_r:
        # Make each column of r be between 0 to 1.
        r_store = [np.array(x) / x[-1] for x in r_store]
    else:
        r_store = [np.array(x) for x in r_store]

    # Calculation of metrics for each experiment
    summary_store = []
    for strain, r in zip(strain_store, r_store):
        if mode == 'line':
            summary = line_optimisation([strain, r], cp0=cp0)
            summary_store.append(summary)
        elif mode == 'poly':
            summary = poly_optimisation([strain, r], poly=poly)
            summary_store.append(summary)
        elif mode == 'poly_cutoff':
            summary = poly_cutoff_optimisation([strain, r], poly=poly)
            summary_store.append(summary)
        elif mode == 'multipoly_cutoff':
            summary = multipoly_cutoff_optimisation([strain, r])
            summary_store.append(summary)
        elif mode == 'line_proxy':
            summary = line_proxy_optimisation([strain, r], poly=poly)
            summary_store.append(summary)

    a_array = np.array(summary_store)[:, 0]
    b_array = np.array(summary_store)[:, 1]
    cp_array = np.array(summary_store)[:, -1]
    c_start_store = []
    for single_cp, single_strain in zip(cp_array, strain_store):
        c_start_store.append(single_strain[int(single_cp)])
    c_start_store = np.array(c_start_store)

    summary_store_name = [[round_sigfigs(x, 3) for x in summary] for summary in summary_store]
    # OLD
    # b_adj = b_array/a_array*(strain_array-c_start_store)**(poly-1)
    # b_adj = b_adj.tolist()

    # Print to Excel and Plotting plots
    if mode == 'line':
        # Print to write_excel_file
        exp_number = np.array(range(np.shape(strain_store)[0])) + 1  # Index to label Exp 1, 2, 3, ...
        df_write = pd.DataFrame(summary_store, index=exp_number, columns=['m1', 'r1', 'm2', 'c2', 'r2', 'x_cp', 'y_cp',
                                                                          'combined_r2', 'cp'])
        df_write.to_excel(write_excel_file)

        # Plotting
        for idx, [strain, r, summary] in enumerate(zip(strain_store, r_store, summary_store)):
            m1, r1, m2, c2, r2, x_cp, y_cp, combined_r2, cp = summary
            plt.scatter(strain[0:cp], r[0:cp], c='b', marker='x', label='1st Line Scatter')
            plt.scatter(strain[cp:], r[cp:], c='r', marker='s', label='2nd Line Scatter')
            first_x = [strain[0], strain[cp - 1]]
            first_y = [x * m1 for x in first_x]
            second_x = [strain[cp], strain[-1]]
            second_y = [x * m2 + c2 for x in second_x]
            plt.plot(first_x, first_y, label='1st Best Fit Line')
            plt.plot(second_x, second_y, label='2nd Best Fit Line')
            # Create empty plot with blank marker containing the extra label
            plt.plot([], [], ' ', label='Combined R Squared = ' + str(round_sigfigs(combined_r2, 3)))
            plt.plot([], [], ' ', label='Crossing Pt. =' + str(cp + 1))
            plt.legend(loc='upper left')
            plt.title('Exp ' + str(idx + 1))
            plt.savefig(plot_directory + '/Exp ' + str(idx + 1) + ' dlr' + '.png', bbox_inches='tight')
            plt.close()
    elif any([mode == 'poly', mode == 'poly_cutoff']):
        # Print to write_excel_file
        exp_number = np.array(range(np.shape(strain_store)[0])) + 1  # Index to label Exp 1, 2, 3, ...
        df_write = pd.DataFrame(summary_store, index=exp_number, columns=['a', 'b', 'c', 'mse', 'e', 'ey', 'cp'])
        df_write.to_excel(write_excel_file)
        # Plotting
        for idx, [strain, r, summary, summary_name] in enumerate(
                zip(strain_store, r_store, summary_store, summary_store_name)):
            a, b, c, mse, e, ey, cp = summary

            # First Segment
            if cp != 0:
                plt.scatter(strain[0:cp], r[0:cp], c='b', marker='x', label='1st Segment Scatter')
                first_x = [strain[0], strain[cp - 1]]
                first_y = [0 for x in first_x]
                plt.plot(first_x, first_y, c='b', label='1st Segment (Line at y=0)')

            # Second segment
            def func(x):
                return a * (x - c) + b * (x - c) ** poly

            plt.scatter(strain[cp:], r[cp:], c='r', marker='s', label='2nd Segment Scatter')
            second_x = strain[cp:]
            second_y = func(second_x)

            plt.plot(second_x, second_y, c='r', label='2nd Segment (Poly ' + str(poly) + ')')
            # Create empty plot with blank marker containing the extra label
            plt.plot([], [], ' ', label='Crossing Pt. =' + str(cp + 1))

            plt.plot([], [], ' ', label='a, b, c, ms, e, ey, cp = ' + str(summary_name))
            plt.legend(loc='upper left')
            plt.title('Exp ' + str(idx + 1))
            plt.savefig(plot_directory + '/Exp ' + str(idx + 1) + ' ' + mode + '.png', bbox_inches='tight')
            plt.close()

    elif mode == 'multipoly_cutoff':
        # Print to write_excel_file
        exp_number = np.array(range(np.shape(strain_store)[0])) + 1  # Index to label Exp 1, 2, 3, ...
        df_write = pd.DataFrame(summary_store, index=exp_number, columns=['a1', 'a2', 'a3', 'c', 'mse', 'e', 'ey', 'cp'])
        df_write.to_excel(write_excel_file)
        # Plotting
        for idx, [strain, r, summary, summary_name] in enumerate(
                zip(strain_store, r_store, summary_store, summary_store_name)):
            a1, a2, a3, c, mse, e, ey, cp = summary

            # First Segment
            if cp != 0:
                plt.scatter(strain[0:cp], r[0:cp], c='b', marker='x', label='1st Segment Scatter')
                first_x = [strain[0], strain[cp - 1]]
                first_y = [0 for x in first_x]
                plt.plot(first_x, first_y, c='b', label='1st Segment (Line at y=0)')

            # Second segment
            def func(x):
                return a1 * (x - c) + a2 * (x - c) ** 2 + a3 * (x - c) ** 3

            plt.scatter(strain[cp:], r[cp:], c='r', marker='s', label='2nd Segment Scatter')
            second_x = strain[cp:]
            second_y = func(second_x)

            plt.plot(second_x, second_y, c='r', label='2nd Segment (Poly ' + str(poly) + ')')
            # Create empty plot with blank marker containing the extra label

            plt.plot([], [], ' ', label='a1, a2, a3, c, ms, e, ey, cp = ' + str(summary_name))
            plt.legend(loc='upper left')
            plt.title('Exp ' + str(idx + 1))
            plt.savefig(plot_directory + '/Exp ' + str(idx + 1) + ' ' + mode + '.png', bbox_inches='tight')
            plt.close()

    elif mode == 'line_proxy':
        # Print to write_excel_file
        exp_number = np.array(range(np.shape(strain_store)[0])) + 1  # Index to label Exp 1, 2, 3, ...
        df_write = pd.DataFrame(summary_store, index=exp_number, columns=['a', 'b', 'c', 'mse', 'm', 'ea',
                                                                          'e', 'ey', 'cp'])
        df_write.to_excel(write_excel_file)

        # Plotting
        for idx, [strain, r, summary, summary_name] in enumerate(
                zip(strain_store, r_store, summary_store, summary_store_name)):
            a, b, c, mse, m, ea, e, ey, cp = summary

            # First Segment
            if cp != 0:
                plt.scatter(strain[0:cp], r[0:cp], c='b', marker='x', label='1st Segment Scatter')
                first_x = [strain[0], strain[cp - 1]]
                first_y = [0 for x in first_x]
                plt.plot(first_x, first_y, c='b', label='1st Segment (Line at y=0)')

            # Second segment
            def func(x):
                return a * (x - c) + b * (x - c) ** poly

            plt.scatter(strain[cp:], r[cp:], c='r', marker='s', label='2nd Segment Scatter')
            second_x = strain[cp:]
            second_y = func(second_x)
            plt.plot(second_x, second_y, c='r', label='2nd Segment (Poly ' + str(poly) + ')')

            # Plot line proxy
            g = e - 0.2 * (e - c)
            gy = func(g)
            x0 = g - gy / m
            plt.plot([x0, e], [0, m * e - m * g + gy], c='g', label='Line Proxy')
            # Create empty plot with blank marker containing the extra label
            plt.plot([], [], ' ', label='Crossing Pt. =' + str(cp + 1))
            plt.plot([], [], ' ', label='a, b, c, mse, m, ea, e, cp =\n ' + str(summary_name))
            plt.legend(loc='upper left')
            plt.title('Exp ' + str(idx + 1))
            plt.savefig(plot_directory + '/Exp ' + str(idx + 1) + ' ' + mode + '.png', bbox_inches='tight')
            plt.close()


def read_excel_data_to_spline(read_excel_file, write_dir, discrete_points, spline_selector):
    cutoff = [10,100]
    # read_excel_file part
    df = pd.read_excel(read_excel_file, sheet_name='raw', header=[0, 1], index_col=0)

    # take only strain columns and make into a new df
    strain = df.xs('Strain (%)', level='Data', axis=1)
    strain = strain.values.T.tolist()  # .T to ensure that each sub list is along the column rather than rows of the df

    # strain store is a list of 1d ndarray, with each inner list being one set of strain data for one experiment
    strain_store = []
    for single_exp in strain:
        strain_single = [x for x in single_exp if not np.isnan(x)]
        strain_store.append(np.array(strain_single))

    # Similar to the above process, but repeat for relative resistance instead
    r = df.xs('R', level='Data', axis=1)
    r = r.values.T.tolist()
    r_store = []
    for single_exp in r:
        r_single = [x for x in single_exp if not np.isnan(x)]
        r_store.append(r_single)

    # Calculation of metrics for each experiment
    summary_store = []
    cutoff_store = []
    cutoff_store2 = []  # second method to do cutoff
    plot_store = []
    gf20_store = []
    for strain, r in zip(strain_store, r_store):
        eval_x = np.linspace(0, strain[-1], num=discrete_points)  # To evaluate spline at
        eval_x_plot = np.linspace(0, strain[-1], num=100)  # For plotting

        if spline_selector == 1:
            spline = PchipInterpolator(strain, r)
            linear_spline = interp1d(strain, r)
            # Getting GF for 20 points
            gf = np.minimum(np.concatenate(
                (
                    (linear_spline.__call__(eval_x[:-1]+0.01)-linear_spline.__call__(eval_x[:-1]))*100/0.01,
                                [(linear_spline.__call__(eval_x[-1])-linear_spline.__call__(eval_x[-1]-0.01))*100/0.01]
                )
            ), 1e6)
            # Store the processed labels. Labels for one example is 1d ndarray of
            # [End_point of strain curve, r1, r2, r3, ... , r_end]
            y_discrete = spline.__call__(eval_x)
            gradient_store = (y_discrete[1:] - y_discrete[:-1])/(eval_x[1]-eval_x[0]) * 100
            # If indexError occurs ==> np.where found nothing ==> there is no points with the required gradient
            # So just put the end point as the answer.
            try:
                cutoff_one = eval_x[np.where(gradient_store>=cutoff[0])[0][0]]
            except IndexError:
                cutoff_one = eval_x[-1]
            try:
                cutoff_two = eval_x[np.where(gradient_store >=cutoff[1])[0][0]]
            except IndexError:
                cutoff_two = eval_x[-1]
            cutoff_store.append([cutoff_one, cutoff_two, strain[-1]])
            summary_store.append(np.concatenate([[strain[-1]], y_discrete]))
            plot_store.append([eval_x_plot, spline.__call__(eval_x_plot)])
            gf20_store.append(gf)
        elif spline_selector == 2:
            # NOT IN USE
            # csaps implementation
            fitted_curve = csaps.UnivariateCubicSmoothingSpline(strain, r, smooth=0.7)
            summary_store.append(np.concatenate([[strain[-1]], fitted_curve(eval_x)]))
            plot_store.append([eval_x_plot, fitted_curve(eval_x_plot)])
        elif spline_selector == 3:
            # NOT IN USE
            # Scripy implementation
            spline = CubicSpline(strain, r)
            # Store the processed labels. Labels for one example is 1d ndarray of
            # [End_point of strain curve, r1, r2, r3, ... , r_end]
            summary_store.append(np.concatenate([[strain[-1]], spline.__call__(eval_x)]))
            plot_store.append([eval_x_plot, spline.__call__(eval_x_plot)])

        # Second cutoff method
        r = np.array(r)
        strain = np.array(strain)
        gf_store = (r[1:] - r[:-1])/(strain[1:] - strain[:-1]) * 100
        if gf_store[-1]<-1:
            cutoff_one = -1
            cutoff_two = -1
        else:
            try:
                cut_idx = np.where(gf_store>=cutoff[0])[0][0]
                if strain[cut_idx] > 0:
                    cutoff_one = strain[cut_idx]
                else:
                    cutoff_one = strain[cut_idx+1]
            except IndexError:
                cutoff_one = strain[-1]
            try:
                cut_idx = np.where(gf_store>=cutoff[1])[0][0]
                if strain[cut_idx] > 0:
                    cutoff_two = strain[cut_idx]
                else:
                    cutoff_two = strain[cut_idx+1]
            except IndexError:
                cutoff_two = strain[-1]
        cutoff_store2.append([cutoff_one, cutoff_two, strain[-1]])



    # Print to write_excel_file


    excel_name = write_dir + '/results.xlsx'
    wb = openpyxl.Workbook()
    wb.create_sheet('points')
    ws = wb['points']

    header = np.array(range(np.shape(strain_store)[0])) + 1  # Index to label Exp 1, 2, 3, ...
    columns = list(range(0,1+discrete_points))
    columns[0] = 'END'
    header = list(header)
    df_write = pd.DataFrame(summary_store, index=header, columns=columns)

    rows = dataframe_to_rows(df_write)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)

    wb.create_sheet('gf20')
    ws = wb['gf20']

    header = np.array(range(np.shape(strain_store)[0])) + 1  # Index to label Exp 1, 2, 3, ...
    columns = list(range(1,1+discrete_points))
    header = list(header)
    df_write = pd.DataFrame(gf20_store, index=header, columns=columns)

    rows = dataframe_to_rows(df_write)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)

    wb.create_sheet('cutoff')
    ws = wb['cutoff']

    header = np.array(range(np.shape(strain_store)[0])) + 1  # Index to label Exp 1, 2, 3, ...
    columns = ['cut={}'.format(x) for x in cutoff] + ['END']
    header = list(header)
    df_write = pd.DataFrame(cutoff_store, index=header, columns=columns)

    rows = dataframe_to_rows(df_write)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)

    wb.save(excel_name)
    wb.close()

    wb.create_sheet('cutoff2')
    ws = wb['cutoff2']

    header = np.array(range(np.shape(strain_store)[0])) + 1  # Index to label Exp 1, 2, 3, ...
    columns = ['cut={}'.format(x) for x in cutoff] + ['END']
    header = list(header)
    df_write = pd.DataFrame(cutoff_store2, index=header, columns=columns)

    rows = dataframe_to_rows(df_write)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)

    wb.save(excel_name)
    wb.close()

    # Plotting
    for idx, [strain, r, plot, summary] in enumerate(zip(strain_store, r_store, plot_store, summary_store)):
        plt.scatter(strain, r, c='r', marker='x', label='Expt. Points')
        plt.plot(plot[0], plot[1], label='Spline Fit')
        eval_x = np.linspace(0, summary[0], num=discrete_points)
        plt.scatter(eval_x, summary[1:], marker='+', label='Discrete Points')
        plt.legend(loc='upper left')
        plt.title('Expt. ' + str(idx + 1))
        plt.savefig(write_dir + '/plots' + '/Expt ' + str(idx + 1) + ' _spline' + '.png', bbox_inches='tight')
        plt.close()


def read_grid_data(read_excel_file, write_dir):
    # read_excel_file part
    df = pd.read_excel(read_excel_file, index_col=0)
    df = df.replace('A',1)
    df = df.replace(['B', 'C', 'D'], 0)

    features = []
    labels = []

    for i in df.index:
        for j in list(df):
            val = df.get_value(i, j)
            if val == 1:
                features.append([i,j])
                labels.append(1)
            elif val == 0:
                features.append([i,j])
                labels.append(0)

    features = np.array(features)/100

    fl = Features_labels_grid(features=features, labels=labels, idx=None)

    with open(write_dir + '/grid_data', 'wb') as handle:
        pickle.dump(fl, handle, protocol=pickle.HIGHEST_PROTOCOL)

    return fl


def line_optimisation(dataset, cp0):
    """
    Given a data set which contains all the data points for 1 experiment, this func will find the best way to cut
    the data set into 2 separate data set such that it can be represented by 2 linear lines with the highest
    sum of R2 value.
    :param dataset: List of two elements the x and y values of one experiment,
    where each x and y is a 1D np array containing those values.
    :param cp0: cp0 is the initial data point index where the dataset will be split into 2 dataset.
    eg: cp0 = k ==> [0, 1, ... N] split to [0, 1, ..., k-1] and [k, k+1, ..., N]
    :return: summary = [Param List, Best Summed R2, cp]
    [Param List] = The optimal set of metric [m1, r1, m2, c2, r2, x_cp, y_cp] which is maximised based on the dataset
    split that maximises r1 and r2.
    Best Summed R2 = Highest sum of R2 scores for line 1 and line 2 after searching through all possible combination
    cp = the data point where the split occurs.
    """

    def dlr(set1, set2):
        """
        Helper function to for the function line_optimisation.
        dlr = Double Linear Regression
        Purpose is to compute two linear regression on dataset 1 and dataset 2.
        :param set1: List [x, y] where x and y are 1D numpy array of the same size, containing the data for
        strain vs resistance. x and y are the datapoints for [0:k] (0,1,...,k-1) pts
        Linear regression on dataset 1 would be fixed intercept at 0.
        :param set2: Same as set1 but x and y are the datapoints [k:N] [k, k+1,..., N]
        Linear regression on dataset 2 would not have a fixed intercept
        :return: List of values [m1, r1, m2, c2, r2, x_cp, y_cp] which stands for [slope of 1st line, R squared of
        first line, slope of second line, intercept of second line, R squared of second line, x- intercept between the
        two lines, y-intercept between the two lines]
        """
        # Line 1
        x = set1[0]
        y = set1[1]
        assert type(x) == np.ndarray and x.ndim == 2 and np.shape(x)[
            1] == 1, 'set1 x value is not a ndarray of shape(-1,1)'
        assert type(y) == np.ndarray and y.ndim == 2 and np.shape(y)[
            1] == 1, 'set1 y value is not a ndarray of shape(-1,1)'
        reg1 = LinearRegression(fit_intercept=False).fit(x, y)  # False such that intercept is fixed at 0
        r1 = reg1.score(x, y)
        m1 = reg1.coef_.item()  # .item() changes ndarray of [[1]] to scalar 1

        # Line 2
        x = set2[0]
        y = set2[1]
        assert x.ndim == 2 and np.shape(x)[1] == 1, 'set2 x value is not a ndarray of shape(-1,1)'
        assert y.ndim == 2 and np.shape(y)[1] == 1, 'set2 y value is not a ndarray of shape(-1,1)'
        reg2 = LinearRegression().fit(x, y)
        r2 = reg2.score(x, y)
        m2 = reg2.coef_.item()
        c2 = reg2.intercept_.item()

        # Intercept
        x_cp = c2 / (m1 - m2)
        y_cp = m1 * x_cp

        return [m1, r1, m2, c2, r2, x_cp, y_cp]

    x = dataset[0]
    y = dataset[1]
    # Total number of data points in x and y (should be same)
    N = np.shape(x)[0]
    assert N == np.shape(y)[0], \
        'Length of x not equal to y ==> Not the same number of data points in x and y.'
    assert N >= cp0 * 2 - 1, \
        'Number of data points is less than cp0 * 2 ==> Not enough to split N into 2 sets with at least cp0 elements' \
        ' each.'

    # Iterating through all possible split combinations from 4 to N-4+1 such that both data set have at least 4
    # data points in it
    score_store = []
    cp_store = range(cp0, N - cp0 + 1)
    for cp in cp_store:
        x1 = x[0:cp].reshape(-1, 1)
        x2 = x[cp:N].reshape(-1, 1)
        y1 = y[0:cp].reshape(-1, 1)
        y2 = y[cp:N].reshape(-1, 1)
        score_store.append(dlr([x1, y1], [x2, y2]))

    # Summing R2 scores of line 1 and line 2 for each combination. Then choose the highest summed R2 value.
    r2_score_store = [score[1] + score[4] for score in score_store]
    max_r2_value = max(r2_score_store)
    max_r2_idx = r2_score_store.index(max_r2_value)
    best_cp = cp_store[max_r2_idx]
    best_score = score_store[max_r2_idx]
    print('############################################################################################################'
          '\nR2 Scores Result: {} \nBest R2 Score Value: {}\nSecond Line Start Pt. Counting from 0: {}\nBest R2 '
          'Score Metric List:\n'
          '[m1, r1, m2, c2, r2, x_cp, y_cp]\n'
          '{}\n'
          '############################################################################################################'
          ''.format(r2_score_store, max_r2_value, best_cp, best_score))
    summary = best_score + [max_r2_value] + [best_cp]

    return summary


def poly_optimisation(dataset, poly):
    """
    Given a data set which contains all the data points for 1 experiment, this func will find the best way to cut
    the data set into 2 separate data set such that it can be represented by:
    1) A flat line at y=0 up till data point c
    2) A quadratic best fit line

    y = u(t-c)*(a(x-c)+b(x-c)^2)
    u = heaviside function
    a = scalar multiple for linear term
    b = scalar multiple for quadratic term
    c = delay term

    :param dataset: List of two elements the x and y values of one experiment,
    where each x and y is a 1D np array containing those values.
    :return: summary = [a, b, c, mse, cp]
    a, b, c is defined above.
    mse = Mean square error
    cp = the data point where the split occurs.
    """

    def quad(set1, set2):
        """
        Helper function to for the function line_optimisation.
        dlr = Double Linear Regression
        Purpose is to compute two linear regression on dataset 1 and dataset 2.
        :param set1: List [x, y] where x and y are 1D numpy array of the same size, containing the data for
        strain vs resistance. x and y are the datapoints for [0:k] (0,1,...,k-1) pts
        Linear regression on dataset 1 would be fixed intercept at 0.
        :param set2: Same as set1 but x and y are the datapoints [k:N] [k, k+1,..., N]
        Linear regression on dataset 2 would not have a fixed intercept
        :return: List of values [m1, r1, m2, c2, r2, x_cp, y_cp] which stands for [slope of 1st line, R squared of
        first line, slope of second line, intercept of second line, R squared of second line, x- intercept between the
        two lines, y-intercept between the two lines]
        """
        # Segement 1
        x = set1[0]
        y = set1[1]
        assert type(x) == np.ndarray and x.ndim == 2 and np.shape(x)[
            1] == 1, 'set1 x value is not a ndarray of shape(-1,1)'
        assert type(y) == np.ndarray and y.ndim == 2 and np.shape(y)[
            1] == 1, 'set1 y value is not a ndarray of shape(-1,1)'
        se1 = np.sum(y ** 2, axis=0)  # Square all y values and sum them, since the line is flat at y=0

        # Segment 2
        x = set2[0]
        y = set2[1]
        assert x.ndim == 2 and np.shape(x)[1] == 1, 'set2 x value is not a ndarray of shape(-1,1)'
        assert y.ndim == 2 and np.shape(y)[1] == 1, 'set2 y value is not a ndarray of shape(-1,1)'

        # Curve fitting for segment 2
        c = x[0, 0]

        def func(x, a, b):
            return a * (x - c) + b * (x - c) ** poly

        popt, _ = curve_fit(func, x[:, 0], y[:, 0], bounds=(0, np.inf))
        a, b = popt
        y_pred = [func(single_x, a, b) for single_x in x]
        se2 = mean_squared_error(y, y_pred) * len(y)

        mse = (se1 + se2) / (np.shape(set1[0])[0] + np.shape(set2[0])[0])
        mse = mse.item()

        return [a, b, c, mse]

    x = dataset[0]
    y = dataset[1]
    # Total number of data points in x and y (should be same)
    N = np.shape(x)[0]
    assert N == np.shape(y)[0], \
        'Length of x not equal to y ==> Not the same number of data points in x and y.'

    # Iterating through all possible split combinations from 4 to N-4+1 such that both data set have at least 4
    # data points in it
    score_store = []
    cp_store = range(0, N - 2)
    for cp in cp_store:
        x1 = x[0:cp].reshape(-1, 1)
        x2 = x[cp:N].reshape(-1, 1)
        y1 = y[0:cp].reshape(-1, 1)
        y2 = y[cp:N].reshape(-1, 1)
        score_store.append(quad([x1, y1], [x2, y2]))

    # Summing R2 scores of line 1 and line 2 for each combination. Then choose the highest summed R2 value.
    mse_store = [score[3] for score in score_store]
    max_mse_value = min(mse_store)
    max_mse_idx = mse_store.index(max_mse_value)
    best_cp = cp_store[max_mse_idx]
    best_score = score_store[max_mse_idx]
    e = x[-1]
    best_score.append(e)
    best_score.append(y[-1])
    best_score.append(best_cp)

    print('############################################################################################################'
          '\nScore Metric List:\n'
          '[a, b, c, mse, e, ey, cp]\n'
          '{}\n'
          '############################################################################################################'
          ''.format(best_score))

    return best_score


def poly_cutoff_optimisation(dataset, poly):
    """
    Given a data set which contains all the data points for 1 experiment, this func will find the best way to cut
    the data set into 2 separate data set such that it can be represented by:
    1) A flat line at y=0 up till data point c
    2) A quadratic best fit line

    y = u(t-c)*(a(x-c)+b(x-c)^2)
    u = heaviside function
    a = scalar multiple for linear term
    b = scalar multiple for quadratic term
    c = delay term

    :param dataset: List of two elements the x and y values of one experiment,
    where each x and y is a 1D np array containing those values.
    :return: summary = [a, b, c, mse, cp]
    a, b, c is defined above.
    mse = Mean square error
    cp = the data point where the split occurs.
    """

    def quad(set1, set2):
        """
        Helper function to for the function line_optimisation.
        dlr = Double Linear Regression
        Purpose is to compute two linear regression on dataset 1 and dataset 2.
        :param set1: List [x, y] where x and y are 1D numpy array of the same size, containing the data for
        strain vs resistance. x and y are the datapoints for [0:k] (0,1,...,k-1) pts
        Linear regression on dataset 1 would be fixed intercept at 0.
        :param set2: Same as set1 but x and y are the datapoints [k:N] [k, k+1,..., N]
        Linear regression on dataset 2 would not have a fixed intercept
        :return: List of values [m1, r1, m2, c2, r2, x_cp, y_cp] which stands for [slope of 1st line, R squared of
        first line, slope of second line, intercept of second line, R squared of second line, x- intercept between the
        two lines, y-intercept between the two lines]
        """
        # Segement 1
        x = set1[0]
        y = set1[1]
        assert type(x) == np.ndarray and x.ndim == 2 and np.shape(x)[
            1] == 1, 'set1 x value is not a ndarray of shape(-1,1)'
        assert type(y) == np.ndarray and y.ndim == 2 and np.shape(y)[
            1] == 1, 'set1 y value is not a ndarray of shape(-1,1)'
        se1 = np.sum(y ** 2, axis=0)  # Square all y values and sum them, since the line is flat at y=0

        # Segment 2
        x = set2[0]
        y = set2[1]
        assert x.ndim == 2 and np.shape(x)[1] == 1, 'set2 x value is not a ndarray of shape(-1,1)'
        assert y.ndim == 2 and np.shape(y)[1] == 1, 'set2 y value is not a ndarray of shape(-1,1)'

        # Curve fitting for segment 2
        c = x[0, 0]

        def func(x, a, b):
            return a * (x - c) + b * (x - c) ** poly

        popt, _ = curve_fit(func, x[:, 0], y[:, 0], bounds=(0, np.inf))
        a, b = popt
        y_pred = [func(single_x, a, b) for single_x in x]
        se2 = mean_squared_error(y, y_pred) * len(y)

        mse = (se1 + se2) / (np.shape(set1[0])[0] + np.shape(set2[0])[0])
        mse = mse.item()

        return [a, b, c, mse]

    x = dataset[0]
    y = dataset[1]
    # Total number of data points in x and y (should be same)
    N = np.shape(x)[0]
    assert N == np.shape(y)[0], \
        'Length of x not equal to y ==> Not the same number of data points in x and y.'

    # Find the cutoff based on 5% of max y value
    y_cutoff = y[-1] * 0.05
    temp_y = y * 0
    mask = (y - y_cutoff) < 0
    temp_y[mask] = (y-y_cutoff)[mask]
    temp_y[temp_y == 0] = -np.inf
    y_cutoff = np.argmax(temp_y)

    x1 = x[0:y_cutoff].reshape(-1, 1)
    x2 = x[y_cutoff :N].reshape(-1, 1)
    y1 = y[0:y_cutoff].reshape(-1, 1)
    y2 = y[y_cutoff:N].reshape(-1, 1)
    score = quad([x1, y1], [x2, y2])

    e = x[-1]
    score.append(e)
    score.append(y[-1])
    score.append(y_cutoff)

    print('############################################################################################################'
          '\nScore Metric List:\n'
          '[a, b, c, mse, e, ey, cp]\n'
          '{}\n'
          '############################################################################################################'
          ''.format(score))

    return score


def multipoly_cutoff_optimisation(dataset):
    """
    Given a data set which contains all the data points for 1 experiment, this func will find the best way to cut
    the data set into 2 separate data set such that it can be represented by:
    1) A flat line at y=0 up till data point c
    2) A quadratic best fit line

    y = u(t-c)*(a(x-c)+b(x-c)^2)
    u = heaviside function
    a = scalar multiple for linear term
    b = scalar multiple for quadratic term
    c = delay term

    :param dataset: List of two elements the x and y values of one experiment,
    where each x and y is a 1D np array containing those values.
    :return: summary = [a, b, c, mse, cp]
    a, b, c is defined above.
    mse = Mean square error
    cp = the data point where the split occurs.
    """

    def quad(set1, set2):
        """
        Helper function to for the function line_optimisation.
        dlr = Double Linear Regression
        Purpose is to compute two linear regression on dataset 1 and dataset 2.
        :param set1: List [x, y] where x and y are 1D numpy array of the same size, containing the data for
        strain vs resistance. x and y are the datapoints for [0:k] (0,1,...,k-1) pts
        Linear regression on dataset 1 would be fixed intercept at 0.
        :param set2: Same as set1 but x and y are the datapoints [k:N] [k, k+1,..., N]
        Linear regression on dataset 2 would not have a fixed intercept
        :return: List of values [m1, r1, m2, c2, r2, x_cp, y_cp] which stands for [slope of 1st line, R squared of
        first line, slope of second line, intercept of second line, R squared of second line, x- intercept between the
        two lines, y-intercept between the two lines]
        """
        # Segement 1
        x = set1[0]
        y = set1[1]
        assert type(x) == np.ndarray and x.ndim == 2 and np.shape(x)[
            1] == 1, 'set1 x value is not a ndarray of shape(-1,1)'
        assert type(y) == np.ndarray and y.ndim == 2 and np.shape(y)[
            1] == 1, 'set1 y value is not a ndarray of shape(-1,1)'
        se1 = np.sum(y ** 2, axis=0)  # Square all y values and sum them, since the line is flat at y=0

        # Segment 2
        x = set2[0]
        y = set2[1]
        assert x.ndim == 2 and np.shape(x)[1] == 1, 'set2 x value is not a ndarray of shape(-1,1)'
        assert y.ndim == 2 and np.shape(y)[1] == 1, 'set2 y value is not a ndarray of shape(-1,1)'

        # Curve fitting for segment 2
        c = x[0, 0]

        def func(x, a1, a2, a3):
            return a1 * (x - c) + a2 * (x - c) ** 2 + a3 * (x - c) ** 3

        popt, _ = curve_fit(func, x[:, 0], y[:, 0], bounds=(0, np.inf))
        a1, a2, a3 = popt
        y_pred = [func(single_x, a1, a2, a3) for single_x in x]
        se2 = mean_squared_error(y, y_pred) * len(y)

        mse = (se1 + se2) / (np.shape(set1[0])[0] + np.shape(set2[0])[0])
        mse = mse.item()

        return [a1, a2, a3, c, mse]

    x = dataset[0]
    y = dataset[1]
    # Total number of data points in x and y (should be same)
    N = np.shape(x)[0]
    assert N == np.shape(y)[0], \
        'Length of x not equal to y ==> Not the same number of data points in x and y.'

    # Find the cutoff based on 5% of max y value
    y_cutoff = y[-1] * 0.05
    temp_y = y * 0
    mask = (y - y_cutoff) < 0
    temp_y[mask] = (y-y_cutoff)[mask]
    temp_y[temp_y == 0] = -np.inf
    y_cutoff = np.argmax(temp_y)

    x1 = x[0:y_cutoff].reshape(-1, 1)
    x2 = x[y_cutoff :N].reshape(-1, 1)
    y1 = y[0:y_cutoff].reshape(-1, 1)
    y2 = y[y_cutoff:N].reshape(-1, 1)
    score = quad([x1, y1], [x2, y2])

    e = x[-1]
    score.append(e)
    score.append(y[-1])
    score.append(y_cutoff)

    print('############################################################################################################'
          '\nScore Metric List:\n'
          '[a1, a2, a3, c, mse, e, ey, cp]\n'
          '{}\n'
          '############################################################################################################'
          ''.format(score))

    return score


def line_proxy_optimisation(dataset, poly):
    """
    Given a data set which contains all the data points for 1 experiment, this func will find the best way to cut
    the data set into 2 separate data set such that it can be represented by:
    1) A flat line at y=0 up till data point c
    2) A quadratic best fit line

    y = u(t-c)*(a(x-c)+b(x-c)^2)
    u = heaviside function
    a = scalar multiple for linear term
    b = scalar multiple for quadratic term
    c = delay term

    :param dataset: List of two elements the x and y values of one experiment,
    where each x and y is a 1D np array containing those values.
    :return: summary = [a, b, c, mse, cp]
    a, b, c is defined above.
    mse = Mean square error
    cp = the data point where the split occurs.
    """

    def quad(set1, set2):
        """
        Helper function to for the function line_optimisation.
        dlr = Double Linear Regression
        Purpose is to compute two linear regression on dataset 1 and dataset 2.
        :param set1: List [x, y] where x and y are 1D numpy array of the same size, containing the data for
        strain vs resistance. x and y are the datapoints for [0:k] (0,1,...,k-1) pts
        Linear regression on dataset 1 would be fixed intercept at 0.
        :param set2: Same as set1 but x and y are the datapoints [k:N] [k, k+1,..., N]
        Linear regression on dataset 2 would not have a fixed intercept
        :return: List of values [m1, r1, m2, c2, r2, x_cp, y_cp] which stands for [slope of 1st line, R squared of
        first line, slope of second line, intercept of second line, R squared of second line, x- intercept between the
        two lines, y-intercept between the two lines]
        """
        # Segement 1
        x = set1[0]
        y = set1[1]
        assert type(x) == np.ndarray and x.ndim == 2 and np.shape(x)[
            1] == 1, 'set1 x value is not a ndarray of shape(-1,1)'
        assert type(y) == np.ndarray and y.ndim == 2 and np.shape(y)[
            1] == 1, 'set1 y value is not a ndarray of shape(-1,1)'
        se1 = np.sum(y ** 2, axis=0)  # Square all y values and sum them, since the line is flat at y=0

        # Segment 2
        x = set2[0]
        y = set2[1]
        assert x.ndim == 2 and np.shape(x)[1] == 1, 'set2 x value is not a ndarray of shape(-1,1)'
        assert y.ndim == 2 and np.shape(y)[1] == 1, 'set2 y value is not a ndarray of shape(-1,1)'

        # Curve fitting for segment 2
        c = x[0, 0]

        def func(x, a, b):
            return a * (x - c) + b * (x - c) ** poly

        popt, _ = curve_fit(func, x[:, 0], y[:, 0], bounds=(0, np.inf))
        a, b = popt
        y_pred = [func(single_x, a, b) for single_x in x]
        se2 = mean_squared_error(y, y_pred) * len(y)

        mse = (se1 + se2) / (np.shape(set1[0])[0] + np.shape(set2[0])[0])
        mse = mse.item()

        return [a, b, c, mse]

    x = dataset[0]
    y = dataset[1]
    # Total number of data points in x and y (should be same)
    N = np.shape(x)[0]
    assert N == np.shape(y)[0], \
        'Length of x not equal to y ==> Not the same number of data points in x and y.'

    # Iterating through all possible split combinations from 4 to N-4+1 such that both data set have at least 4
    # data points in it
    score_store = []
    cp_store = range(0, N - 2)
    for cp in cp_store:
        x1 = x[0:cp].reshape(-1, 1)
        x2 = x[cp:N].reshape(-1, 1)
        y1 = y[0:cp].reshape(-1, 1)
        y2 = y[cp:N].reshape(-1, 1)
        score_store.append(quad([x1, y1], [x2, y2]))

    # Summing R2 scores of line 1 and line 2 for each combination. Then choose the highest summed R2 value.
    mse_store = [score[3] for score in score_store]
    max_mse_value = min(mse_store)
    max_mse_idx = mse_store.index(max_mse_value)
    best_cp = cp_store[max_mse_idx]
    best_score = score_store[max_mse_idx]
    best_score.append(best_cp)

    a, b, c, mse, cp = best_score
    e = x[-1]
    ey = y[-1]

    # Line approx at 90% of interval between c and e
    # Let g be that 90% point
    g = e - 0.2 * (e - c)
    gy = a * (g - c) + b * (g - c) ** poly
    m = a + poly * b * (g - c) ** (poly - 1)

    # Integrating along dy to get error. Note that x and y are flipped since integrating along y axis
    def difference(yp, x):
        return abs(x - (1 / m * yp + g - gy / m))

    x_set = difference(y, x)

    # Remove repeated R = y = 0 values
    last_zero_idx = np.where(y == 0)[0][-1]
    x_set = np.copy(x_set[last_zero_idx:])
    y = np.copy(y[last_zero_idx:])

    ea = trapz(y=x_set, x=y)

    best_score = [a, b, c, mse, m, ea, e, ey, cp]

    print('############################################################################################################'
          '\nScore Metric List:\n'
          '[a, b, c, mse, m, ea, e, ey, cp]\n'
          '{}\n'
          '############################################################################################################'
          ''.format(best_score))

    return best_score
